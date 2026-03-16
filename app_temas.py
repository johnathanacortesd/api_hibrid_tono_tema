# ======================================
# Importaciones
# ======================================
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle
from collections import defaultdict, Counter
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import openai
import re
import time
from unidecode import unidecode
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
import json
import asyncio
import hashlib
from typing import List, Dict, Tuple, Optional, Any
import joblib
import gc

# ======================================
# Configuracion general
# ======================================
st.set_page_config(
    page_title="Análisis de Noticias · IA",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

OPENAI_MODEL_EMBEDDING    = "text-embedding-3-small"
OPENAI_MODEL_CLASIFICACION = "gpt-4.1-nano-2025-04-14"

CONCURRENT_REQUESTS          = 50
SIMILARITY_THRESHOLD_TONO    = 0.92
SIMILARITY_THRESHOLD_TITULOS = 0.95

# ─── Umbrales de clustering ────────────────────────────────────────────────────
# SUBTEMA: dos noticias se consideran "mismo subtema" si su similitud coseno ≥ este valor.
# Sube este valor → menos grupos, etiquetas más generales.
# Baja este valor → más grupos, etiquetas más específicas.
# Rango recomendado: 0.78 – 0.88
UMBRAL_SUBTEMA = 0.82

# TEMA: similitud mínima entre centroides de subtemas para fusionarlos en el mismo tema.
# Valor más bajo = más fusión = menos temas finales.
# Rango recomendado: 0.72 – 0.85
UMBRAL_TEMA    = 0.78

# Número máximo de temas principales tras consolidación
NUM_TEMAS_MAX  = 20

PRICE_INPUT_1M    = 0.10
PRICE_OUTPUT_1M   = 0.40
PRICE_EMBEDDING_1M = 0.02

if 'tokens_input'     not in st.session_state: st.session_state['tokens_input']     = 0
if 'tokens_output'    not in st.session_state: st.session_state['tokens_output']    = 0
if 'tokens_embedding' not in st.session_state: st.session_state['tokens_embedding'] = 0

# ── Geografía ─────────────────────────────────────────────────────────────────
CIUDADES_COLOMBIA = {
    "bogotá","bogota","medellín","medellin","cali","barranquilla","cartagena","cúcuta","cucuta",
    "bucaramanga","pereira","manizales","armenia","ibagué","ibague","villavicencio","montería",
    "monteria","neiva","pasto","valledupar","popayán","popayan","tunja","florencia","sincelejo",
    "riohacha","yopal","santa marta","santamarta","quibdó","quibdo","leticia","mocoa","mitú","mitu",
    "puerto carreño","inírida","inirida","san josé del guaviare","antioquia","atlántico","atlantico",
    "bolívar","bolivar","boyacá","boyaca","caldas","caquetá","caqueta","casanare","cauca","cesar",
    "chocó","choco","córdoba","cordoba","cundinamarca","guainía","guainia","guaviare","huila",
    "la guajira","magdalena","meta","nariño","narino","norte de santander","putumayo","quindío",
    "quindio","risaralda","san andrés","san andres","santander","sucre","tolima","valle del cauca",
    "vaupés","vaupes","vichada",
}
GENTILICIOS_COLOMBIA = {
    "bogotano","bogotanos","bogotana","bogotanas","capitalino","capitalinos","capitalina","capitalinas",
    "antioqueño","antioqueños","antioqueña","antioqueñas","paisa","paisas","medellense","medellenses",
    "caleño","caleños","caleña","caleñas","valluno","vallunos","valluna","vallunas","vallecaucano",
    "vallecaucanos","barranquillero","barranquilleros","cartagenero","cartageneros","costeño","costeños",
    "costeña","costeñas","cucuteño","cucuteños","bumangués","santandereano","santandereanos",
    "boyacense","boyacenses","tolimense","tolimenses","huilense","huilenses","nariñense","nariñenses",
    "pastuso","pastusas","cordobés","cordobeses","cauca","caucano","caucanos","chocoano","chocoanos",
    "casanareño","casanareños","caqueteño","caqueteños","guajiro","guajiros","llanero","llaneros",
    "amazonense","amazonenses","colombiano","colombianos","colombiana","colombianas",
}

STOPWORDS_ES = set("""
a ante bajo cabe con contra de desde durante en entre hacia hasta mediante para por segun sin so
sobre tras y o u e la el los las un una unos unas lo al del se su sus le les mi mis tu tus nuestro
nuestros vuestra vuestras este esta estos estas ese esa esos esas aquel aquella aquellos aquellas
que cual cuales quien quienes cuyo cuya cuyos cuyas como cuando donde cual es son fue fueron era
eran sera seran seria serian he ha han habia habian hay hubo habra habria estoy esta estan estaba
estaban estamos estan estar estare estaria estuvieron estarian estuvo asi ya mas menos tan tanto
cada
""".split())

POS_VARIANTS = [
    r"lanz(a(r|ra|ria|o|on|an|ando)?|amiento)s?", r"prepar(a|ando)",
    r"nuev[oa]\s+(servicio|tienda|plataforma|app|aplicacion|funcion|canal|portal|producto|iniciativa|proyecto)",
    r"apertur(a|ar|ara|o|an)",r"estren(a|o|ara|an|ando)",r"habilit(a|o|ara|an|ando)",
    r"disponible",r"mejor(a|o|an|ando)",r"optimiza|amplia|expande",
    r"alianz(a|as)|acuerd(o|a|os)|convenio(s)?|memorando(s)?|joint\s+venture|colaboraci[oó]n(es)?|asociaci[oó]n(es)?|partnership(s)?|fusi[oó]n(es)?|integraci[oó]n(es)?",
    r"crecimi?ento|aument(a|o|an|ando)",r"gananci(a|as)|utilidad(es)?|benefici(o|os)",
    r"expansion|crece|crecer",r"inversion|invierte|invertir",
    r"innova(cion|dor|ndo)|moderniza",r"exito(so|sa)?|logr(o|os|a|an|ando)",
    r"reconoci(miento|do|da)|premi(o|os|ada)",r"lidera(zgo)?|lider",
    r"consolida|fortalece",r"oportunidad(es)?|potencial",r"solucion(es)?|resuelve",
    r"eficien(te|cia)",r"calidad|excelencia",r"satisfaccion|complace",
    r"confianza|credibilidad",r"sostenible|responsable",r"compromiso|apoya|apoyar",
    r"patrocin(io|a|ador|an|ando)|auspic(ia|io|iador)",r"gana(r|dor|dora|ndo)?|triunf(a|o|ar|ando)",
    r"destaca(r|do|da|ndo)?",r"supera(r|ndo|cion)?",r"record|hito|milestone",
    r"avanza(r|do|da|ndo)?",r"benefici(a|o|ando|ar|ando)",r"importante(s)?",
    r"prioridad",r"bienestar",r"garantizar",r"seguridad",r"atencion",
    r"expres(o|ó|ando)",r"señala(r|do|ando)",r"ratific(a|o|ando|ar)",
]
NEG_VARIANTS = [
    r"demanda|denuncia|sanciona|multa|investiga|critica",
    r"cae|baja|pierde|crisis|quiebra|default",
    r"fraude|escandalo|irregularidad",
    r"fall(a|o|os)|interrumpe|suspende|cierra|renuncia|huelga",
    r"filtracion|ataque|phishing|hackeo|incumple|boicot|queja|reclamo|deteriora",
    r"problema(s|tica|ico)?|dificultad(es)?",r"retras(o|a|ar|ado)",r"perdida(s)?|deficit",
    r"conflict(o|os)?|disputa(s)?",r"rechaz(a|o|ar|ado)",r"negativ(o|a|os|as)",
    r"preocupa(cion|nte|do)?",r"alarma(nte)?|alerta",r"riesgo(s)?|amenaza(s)?",
]
CRISIS_KEYWORDS  = re.compile(r"\b(crisis|emergencia|desastre|deslizamiento|inundaci[oó]n|afectaciones|damnificados|tragedia|zozobra|alerta)\b", re.IGNORECASE)
RESPONSE_VERBS   = re.compile(r"\b(atiend(e|en|iendo)|activ(a|o|ando)|decret(a|o|ando)|responde(r|iendo)|trabaj(a|ando)|lidera(ndo)?|enfrenta(ndo)?|gestiona(ndo)?|declar(o|a|ando)|anunci(a|o|ando))\b", re.IGNORECASE)
POS_PATTERNS     = [re.compile(rf"\b(?:{p})\b", re.IGNORECASE) for p in POS_VARIANTS]
NEG_PATTERNS     = [re.compile(rf"\b(?:{p})\b", re.IGNORECASE) for p in NEG_VARIANTS]

# ======================================
# CSS — Editorial oscuro, tipografía serif + mono
# ======================================
def load_custom_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=JetBrains+Mono:wght@300;400;500&family=Lato:wght@300;400;600&display=swap');

:root {
    --bg:       #0e0e10;
    --s1:       #15151a;
    --s2:       #1d1d24;
    --s3:       #26262f;
    --border:   #2e2e3a;
    --border2:  #3e3e50;
    --text:     #e4e2dc;
    --text2:    #928f8a;
    --text3:    #52504e;
    --gold:     #c8a96a;
    --gold2:    #8b6f3a;
    --green:    #5c9e78;
    --red:      #b85c5c;
    --blue:     #6a8db0;
    --r:        8px;
    --r2:       14px;
}

html, body, [data-testid="stApp"] {
    background: var(--bg) !important;
    color: var(--text) !important;
    font-family: 'Lato', sans-serif;
    font-weight: 300;
}

/* ── Header ── */
.app-header {
    padding: 2.5rem 0 1.5rem;
    border-bottom: 1px solid var(--border);
    margin-bottom: 2.2rem;
    display: flex;
    align-items: flex-end;
    gap: 1.2rem;
}
.app-header-mark {
    font-family: 'Playfair Display', serif;
    font-size: 3rem;
    color: var(--gold);
    line-height: 1;
    user-select: none;
}
.app-header-text {}
.app-header-title {
    font-family: 'Playfair Display', serif;
    font-size: 1.85rem;
    font-weight: 700;
    color: var(--text);
    line-height: 1;
    letter-spacing: -0.01em;
}
.app-header-version {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.68rem;
    color: var(--text3);
    letter-spacing: 0.12em;
    text-transform: uppercase;
    margin-top: 0.35rem;
}

/* ── Tabs ── */
[data-testid="stTabs"] [data-testid="stTabsList"] {
    background: var(--s1) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    padding: 3px !important;
    gap: 3px !important;
}
[data-testid="stTabs"] button[data-baseweb="tab"] {
    font-family: 'Lato', sans-serif !important;
    font-size: 0.88rem !important;
    font-weight: 400 !important;
    color: var(--text2) !important;
    border-radius: 6px !important;
    padding: 0.55rem 1.4rem !important;
    border: none !important;
    background: transparent !important;
    transition: all 0.18s !important;
    letter-spacing: 0.02em !important;
}
[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"] {
    background: var(--s2) !important;
    color: var(--gold) !important;
    border: 1px solid var(--border2) !important;
    font-weight: 600 !important;
}

/* ── Métricas ── */
.metrics-row {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 0.9rem;
    margin: 1.8rem 0;
}
.metric-card {
    background: var(--s1);
    border: 1px solid var(--border);
    border-radius: var(--r2);
    padding: 1.3rem 1rem;
    text-align: center;
    transition: border-color 0.2s, transform 0.2s;
}
.metric-card:hover { border-color: var(--border2); transform: translateY(-1px); }
.metric-val {
    font-family: 'Playfair Display', serif;
    font-size: 2rem;
    font-weight: 700;
    line-height: 1;
    margin-bottom: 0.5rem;
}
.metric-lbl {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.65rem;
    color: var(--text3);
    text-transform: uppercase;
    letter-spacing: 0.14em;
}

/* ── Formularios ── */
[data-testid="stForm"] {
    background: var(--s1) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r2) !important;
    padding: 2rem !important;
}

/* ── Section labels ── */
.sec-label {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.67rem;
    color: var(--text3);
    letter-spacing: 0.18em;
    text-transform: uppercase;
    padding-bottom: 0.6rem;
    border-bottom: 1px solid var(--border);
    margin: 1.8rem 0 1rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}
.sec-label::before {
    content: '';
    display: inline-block;
    width: 3px;
    height: 10px;
    background: var(--gold2);
    border-radius: 2px;
}

/* ── Inputs ── */
[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea {
    background: var(--s2) !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    border-radius: var(--r) !important;
    font-family: 'Lato', sans-serif !important;
    font-weight: 300 !important;
    transition: border-color 0.18s, box-shadow 0.18s !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stTextArea"] textarea:focus {
    border-color: var(--gold2) !important;
    box-shadow: 0 0 0 2px rgba(200,169,106,0.10) !important;
    outline: none !important;
}
label[data-testid="stWidgetLabel"] p {
    color: var(--text2) !important;
    font-size: 0.85rem !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    background: var(--s2) !important;
    border: 1px dashed var(--border2) !important;
    border-radius: var(--r) !important;
    transition: border-color 0.18s !important;
}
[data-testid="stFileUploader"]:hover { border-color: var(--gold2) !important; }

/* ── Botones ── */
.stButton > button,
[data-testid="stDownloadButton"] > button {
    background: var(--s2) !important;
    border: 1px solid var(--border2) !important;
    color: var(--text) !important;
    border-radius: var(--r) !important;
    font-family: 'Lato', sans-serif !important;
    font-weight: 400 !important;
    font-size: 0.9rem !important;
    letter-spacing: 0.04em !important;
    transition: all 0.18s !important;
    padding: 0.55rem 1.4rem !important;
}
.stButton > button:hover,
[data-testid="stDownloadButton"] > button:hover {
    border-color: var(--gold) !important;
    color: var(--gold) !important;
    background: var(--s1) !important;
}
.stButton > button[kind="primary"],
[data-testid="stDownloadButton"] > button[kind="primary"] {
    background: linear-gradient(135deg, var(--gold2) 0%, var(--gold) 100%) !important;
    border: none !important;
    color: #0e0e10 !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 12px rgba(200,169,106,0.20) !important;
}
.stButton > button[kind="primary"]:hover,
[data-testid="stDownloadButton"] > button[kind="primary"]:hover {
    box-shadow: 0 4px 20px rgba(200,169,106,0.35) !important;
    transform: translateY(-1px) !important;
    color: #0e0e10 !important;
}

/* ── Radio ── */
[data-testid="stRadio"] label { color: var(--text2) !important; font-size: 0.88rem !important; }
[data-testid="stRadio"] [aria-checked="true"] + div label { color: var(--gold) !important; }

/* ── Status ── */
[data-testid="stStatus"] {
    background: var(--s1) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.82rem !important;
}

/* ── Alerts ── */
[data-testid="stAlert"] {
    background: var(--s2) !important;
    border: 1px solid var(--border2) !important;
    border-radius: var(--r) !important;
    color: var(--text2) !important;
    font-size: 0.85rem !important;
}

/* ── Success banner ── */
.success-banner {
    background: linear-gradient(135deg, var(--s1) 0%, var(--s2) 100%);
    border: 1px solid var(--green);
    border-left: 3px solid var(--green);
    border-radius: var(--r2);
    padding: 1.6rem 1.8rem;
    margin: 1rem 0 1.6rem;
}
.success-title {
    font-family: 'Playfair Display', serif;
    font-size: 1.4rem;
    color: var(--green);
    margin-bottom: 0.2rem;
}
.success-sub {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.7rem;
    color: var(--text3);
    letter-spacing: 0.08em;
}

/* ── Auth screen ── */
.auth-wrap { max-width: 380px; margin: 10vh auto 0; }
.auth-title {
    font-family: 'Playfair Display', serif;
    font-size: 2.6rem;
    color: var(--gold);
    text-align: center;
    margin-bottom: 0.3rem;
}
.auth-sub {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.7rem;
    color: var(--text3);
    text-align: center;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    margin-bottom: 2.2rem;
}

/* ── Progress ── */
[data-testid="stProgressBar"] > div > div { background: var(--gold) !important; border-radius: 4px !important; }

/* ── Dataframe ── */
[data-testid="stDataFrame"] { border: 1px solid var(--border) !important; border-radius: var(--r) !important; }

/* ── Info box cluster settings ── */
.cluster-info {
    background: var(--s2);
    border: 1px solid var(--border2);
    border-left: 3px solid var(--gold2);
    border-radius: var(--r);
    padding: 1rem 1.2rem;
    margin: 0.8rem 0;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.78rem;
    color: var(--text2);
    line-height: 1.7;
}

hr { border-color: var(--border) !important; }

::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--border2); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--gold2); }

.footer {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.65rem;
    color: var(--text3);
    text-align: center;
    padding: 1.8rem 0 1rem;
    letter-spacing: 0.1em;
    border-top: 1px solid var(--border);
    margin-top: 3rem;
}
</style>
""", unsafe_allow_html=True)


# ======================================
# Utilidades
# ======================================
def check_password() -> bool:
    if st.session_state.get("password_correct", False): return True
    st.markdown("""
    <div class="auth-wrap">
        <div class="auth-title">◈</div>
        <div class="auth-sub">Acceso restringido · Sistema IA</div>
    </div>
    """, unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("password_form"):
            password = st.text_input("Contraseña", type="password", placeholder="···")
            if st.form_submit_button("Ingresar →", use_container_width=True, type="primary"):
                if password == st.secrets.get("APP_PASSWORD", "INVALID"):
                    st.session_state["password_correct"] = True; st.rerun()
                else:
                    st.error("Contraseña incorrecta")
    return False

def call_with_retries(fn, *args, **kwargs):
    delay = 1
    for attempt in range(3):
        try: return fn(*args, **kwargs)
        except Exception as e:
            if attempt == 2: raise e
            time.sleep(delay); delay *= 2

async def acall_with_retries(fn, *args, **kwargs):
    delay = 1
    for attempt in range(3):
        try: return await fn(*args, **kwargs)
        except Exception as e:
            if attempt == 2: raise e
            await asyncio.sleep(delay); delay *= 2

def norm_key(text: Any) -> str:
    if text is None: return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))

def limpiar_tema(tema: str) -> str:
    if not tema: return "Sin tema"
    tema = tema.strip().strip('"\'')
    if tema: tema = tema[0].upper() + tema[1:]
    invalid_end = {"en","de","del","la","el","y","o","con","sin","por","para","sobre"}
    palabras = tema.split()
    while palabras and palabras[-1].lower() in invalid_end: palabras.pop()
    tema = " ".join(palabras[:6])
    return tema if tema else "Sin tema"

def limpiar_tema_geografico(tema: str, marca: str, aliases: List[str]) -> str:
    if not tema: return "Sin tema"
    tl = tema.lower()
    for name in [marca] + [a for a in aliases if a]:
        tl = re.sub(rf'\b{re.escape(unidecode(name.lower()))}\b', '', tl)
    for ciudad in CIUDADES_COLOMBIA:
        tl = re.sub(rf'\b{re.escape(ciudad)}\b', '', tl)
    for gent in GENTILICIOS_COLOMBIA:
        tl = re.sub(rf'\b{re.escape(gent)}\b', '', tl)
    for frase in ["en colombia","de colombia","del pais","en el pais","nacional",
                  "colombiano","colombiana","colombianos","colombianas","territorio nacional"]:
        tl = re.sub(rf'\b{re.escape(frase)}\b', '', tl)
    palabras = [p.strip() for p in tl.split() if p.strip()]
    if not palabras: return "Sin tema"
    tl = palabras[0].upper() + (" ".join(palabras))[len(palabras[0]):]
    return limpiar_tema(tl)

def string_norm_label(s: str) -> str:
    if not s: return ""
    s = unidecode(s.lower())
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return " ".join(t for t in s.split() if t not in STOPWORDS_ES)

def extract_link(cell):
    if hasattr(cell, "hyperlink") and cell.hyperlink:
        return {"value": "Link", "url": cell.hyperlink.target}
    if isinstance(cell.value, str) and "=HYPERLINK" in cell.value:
        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value)
        if m: return {"value": "Link", "url": m.group(1)}
    return {"value": cell.value, "url": None}

def normalize_title_for_comparison(title: Any) -> str:
    if not isinstance(title, str): return ""
    tmp = re.split(r"\s*[:|-]\s*", title, 1)
    return re.sub(r"\W+", " ", tmp[0]).lower().strip()

def clean_title_for_output(title: Any) -> str:
    return re.sub(r"\s*\|\s*[\w\s]+$", "", str(title)).strip()

def corregir_texto(text: Any) -> Any:
    if not isinstance(text, str): return text
    text = re.sub(r"(<br>|\[\.\.\.\]|\s+)", " ", text).strip()
    m = re.search(r"[A-ZÁÉÍÓÚÑ]", text)
    if m: text = text[m.start():]
    if text and not text.endswith("..."): text = text.rstrip(".") + "..."
    return text

def normalizar_tipo_medio(tipo_raw: str) -> str:
    if not isinstance(tipo_raw, str): return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    return {
        "fm":"Radio","am":"Radio","radio":"Radio",
        "aire":"Televisión","cable":"Televisión","tv":"Televisión",
        "television":"Televisión","televisión":"Televisión",
        "senal abierta":"Televisión","señal abierta":"Televisión",
        "diario":"Prensa","prensa":"Prensa",
        "revista":"Revista","revistas":"Revista",
        "online":"Internet","internet":"Internet","digital":"Internet","web":"Internet",
    }.get(t, str(tipo_raw).strip().title() or "Otro")

# ======================================
# Embeddings
# ======================================
def get_embeddings_batch(textos: List[str], batch_size: int = 100) -> List[Optional[List[float]]]:
    if not textos: return []
    resultados: List[Optional[List[float]]] = [None] * len(textos)
    for i in range(0, len(textos), batch_size):
        batch = [t[:2000] if t else "" for t in textos[i:i+batch_size]]
        try:
            resp = call_with_retries(openai.Embedding.create, input=batch, model=OPENAI_MODEL_EMBEDDING)
            usage = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if usage:
                total = usage.get('total_tokens') if isinstance(usage, dict) else getattr(usage, 'total_tokens', 0)
                st.session_state['tokens_embedding'] += (total or 0)
            for j, d in enumerate(resp["data"]): resultados[i+j] = d["embedding"]
        except:
            for j, t in enumerate(batch):
                try:
                    r = openai.Embedding.create(input=[t], model=OPENAI_MODEL_EMBEDDING)
                    resultados[i+j] = r["data"][0]["embedding"]
                except: pass
    return resultados

# ======================================
# DSU (Union-Find)
# ======================================
class DSU:
    def __init__(self, n: int):
        self.p    = list(range(n))
        self.rank = [0] * n

    def find(self, i: int) -> int:
        path = []
        while self.p[i] != i: path.append(i); i = self.p[i]
        for node in path: self.p[node] = i
        return i

    def union(self, i: int, j: int):
        ri, rj = self.find(i), self.find(j)
        if ri == rj: return
        if self.rank[ri] < self.rank[rj]: ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]: self.rank[ri] += 1

    def grupos(self, n: int) -> Dict[int, List[int]]:
        comp: Dict[int, List[int]] = defaultdict(list)
        for i in range(n): comp[self.find(i)].append(i)
        return dict(comp)

# ======================================
# Agrupación de tono
# ======================================
def agrupar_textos_similares(textos: List[str], umbral: float) -> Dict[int, List[int]]:
    if not textos: return {}
    embs = get_embeddings_batch(textos)
    valid = [(i, e) for i, e in enumerate(embs) if e is not None]
    if len(valid) < 2: return {}
    idxs, M = zip(*valid)
    labels = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1-umbral, metric="cosine", linkage="average"
    ).fit(np.array(M)).labels_
    g: Dict[int, List[int]] = defaultdict(list)
    for k, lbl in enumerate(labels): g[lbl].append(idxs[k])
    return dict(enumerate(g.values()))

def agrupar_por_titulo_similar(titulos: List[str]) -> Dict[int, List[int]]:
    gid, grupos, used = 0, {}, set()
    norm = [normalize_title_for_comparison(t) for t in titulos]
    for i in range(len(norm)):
        if i in used or not norm[i]: continue
        grp = [i]; used.add(i)
        for j in range(i+1, len(norm)):
            if j in used or not norm[j]: continue
            if SequenceMatcher(None, norm[i], norm[j]).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                grp.append(j); used.add(j)
        if len(grp) >= 2: grupos[gid] = grp; gid += 1
    return grupos

def seleccionar_representante(indices: List[int], textos: List[str]) -> Tuple[int, str]:
    embs = get_embeddings_batch([textos[i] for i in indices])
    validos = [(indices[k], e) for k, e in enumerate(embs) if e is not None]
    if not validos: return indices[0], textos[indices[0]]
    idxs, M = zip(*validos)
    centro = np.mean(M, axis=0, keepdims=True)
    best = int(np.argmax(cosine_similarity(np.array(M), centro)))
    return idxs[best], textos[idxs[best]]

# ======================================
# CLASIFICADOR DE TONO
# ======================================
class ClasificadorTono:
    def __init__(self, marca: str, aliases: List[str]):
        self.marca   = marca
        self.aliases = aliases or []
        names    = [marca] + [a for a in self.aliases if a]
        patterns = [re.escape(unidecode(n.strip().lower())) for n in names if n.strip()]
        self.brand_re = re.compile(
            r"\b(" + "|".join(patterns) + r")\b" if patterns else r"(a^b)",
            re.IGNORECASE
        )

    def _contextos(self, texto: str) -> List[str]:
        tl = unidecode(texto.lower())
        matches = list(self.brand_re.finditer(tl))
        if not matches: return [texto[:600]]
        out = []
        for i, m in enumerate(matches):
            win = 250 if i == 0 else 150
            s   = max(0, m.start() - win)
            e   = min(len(texto), m.end() + win)
            while e < len(texto) and texto[e] not in '.!?': e += 1
            out.append(texto[s:e+1].strip())
        return list(dict.fromkeys(out))[:4]

    def _reglas(self, contextos: List[str]) -> Optional[str]:
        pos, neg = 0, 0
        for ctx in contextos:
            t = unidecode(ctx.lower())
            neg_present = bool(re.search(
                r'\b(no|sin|nunca|jamás|niega|rechaza|desmiente)\b.{0,30}',
                t, re.IGNORECASE
            ))
            ph = sum(1 for p in POS_PATTERNS if p.search(t))
            nh = sum(1 for p in NEG_PATTERNS if p.search(t))
            if CRISIS_KEYWORDS.search(t) and RESPONSE_VERBS.search(t): pos += 3; continue
            if neg_present: neg += ph
            else: pos += ph; neg += nh
        if pos >= 3 and pos > neg * 1.5: return "Positivo"
        if neg >= 3 and neg > pos * 1.5: return "Negativo"
        return None

    async def _llm(self, contextos: List[str]) -> Dict[str, str]:
        aliases_str = ", ".join(self.aliases) or "ninguno"
        prompt = (
            f"Analiza el sentimiento hacia '{self.marca}' (alias: {aliases_str}).\n"
            f"Positivo: logros, lanzamientos, reconocimientos.\n"
            f"Negativo: críticas, sanciones, pérdidas.\n"
            f"Neutro: menciones informativas.\n"
            f"Fragmentos:\n---\n{chr(10).join(contextos[:3])}\n---\n"
            f'Responde SOLO JSON: {{"tono":"Positivo|Negativo|Neutro"}}'
        )
        try:
            resp = await acall_with_retries(
                openai.ChatCompletion.acreate,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role":"user","content":prompt}],
                max_tokens=50, temperature=0.0,
                response_format={"type":"json_object"}
            )
            usage = resp.get('usage',{}) if isinstance(resp,dict) else getattr(resp,'usage',{})
            if usage:
                pt = usage.get('prompt_tokens') if isinstance(usage,dict) else getattr(usage,'prompt_tokens',0)
                ct = usage.get('completion_tokens') if isinstance(usage,dict) else getattr(usage,'completion_tokens',0)
                st.session_state['tokens_input']  += (pt or 0)
                st.session_state['tokens_output'] += (ct or 0)
            tono = str(json.loads(resp.choices[0].message.content).get("tono","Neutro")).title()
            return {"tono": tono if tono in ("Positivo","Negativo","Neutro") else "Neutro"}
        except: return {"tono":"Neutro"}

    async def _clasificar_async(self, texto: str, sem: asyncio.Semaphore):
        async with sem:
            ctx = self._contextos(texto)
            r   = self._reglas(ctx)
            if r: return {"tono": r}
            return await self._llm(ctx)

    async def procesar_lote_async(self, textos: pd.Series, pbar, resumenes: pd.Series, titulos: pd.Series):
        n     = len(textos)
        txts  = textos.tolist()
        pbar.progress(0.05, "Agrupando para análisis de tono…")
        dsu = DSU(n)
        for g in [agrupar_textos_similares(txts, SIMILARITY_THRESHOLD_TONO),
                  agrupar_por_titulo_similar(titulos.tolist())]:
            for _, idxs in g.items():
                for j in idxs[1:]: dsu.union(idxs[0], j)
        grupos   = dsu.grupos(n)
        reps     = {cid: seleccionar_representante(idxs, txts)[1] for cid, idxs in grupos.items()}
        sem      = asyncio.Semaphore(CONCURRENT_REQUESTS)
        tasks    = [self._clasificar_async(rep, sem) for rep in reps.values()]
        resultados = []
        for i, f in enumerate(asyncio.as_completed(tasks)):
            resultados.append(await f)
            pbar.progress(0.1 + 0.85*(i+1)/len(tasks), f"Analizando tono {i+1}/{len(tasks)}")
        res_por_grupo = {list(reps.keys())[i]: r for i, r in enumerate(resultados)}
        final = [None] * n
        for cid, idxs in grupos.items():
            r = res_por_grupo.get(cid, {"tono":"Neutro"})
            for i in idxs: final[i] = r
        pbar.progress(1.0, "Tono completado")
        return final

def analizar_tono_con_pkl(textos, pkl_file):
    try:
        pipeline = joblib.load(pkl_file)
        TONO_MAP = {1:"Positivo","1":"Positivo",0:"Neutro","0":"Neutro",-1:"Negativo","-1":"Negativo"}
        return [{"tono": TONO_MAP.get(p, str(p).title())} for p in pipeline.predict(textos)]
    except Exception as e:
        st.error(f"Error pkl sentimiento: {e}"); return None

# ======================================
# CLASIFICADOR DE SUBTEMAS — flujo invertido
# ======================================
# PRINCIPIO CENTRAL:
# 1. Agrupar TODO el contenido primero (hash + títulos + semántica)
# 2. Por cada grupo final, generar UNA sola etiqueta
# 3. Nunca generar dos etiquetas para fusionarlas después
#
# Esto garantiza que:
# a) Noticias en el mismo grupo → siempre el mismo subtema (determinista)
# b) No hay etiquetas redundantes que "deberían" haberse fusionado
# c) El número de subtemas = número de clusters semánticos reales

class ClasificadorSubtema:
    def __init__(self, marca: str, aliases: List[str]):
        self.marca   = marca
        self.aliases = aliases or []
        # Caché: clave → etiqueta. Permite reutilizar entre runs si el objeto persiste.
        self._cache: Dict[str, str] = {}

    # ── Paso 1: hash exacto ───────────────────────────────────────────────────
    def _paso1_hash_exacto(self, titulos: List[str], resumenes: List[str], dsu: DSU):
        """Noticias con título o inicio de resumen idéntico → mismo grupo."""
        def norm(t: str) -> str:
            t = unidecode(str(t).lower())
            return re.sub(r'[^a-z0-9\s]', '', t).split()
        def tok_40(t): return ' '.join(norm(t)[:40])
        def tok_15(t): return ' '.join(norm(t)[:15])

        bkt_tit: Dict[str, List[int]] = defaultdict(list)
        bkt_res: Dict[str, List[int]] = defaultdict(list)
        for i, (tit, res) in enumerate(zip(titulos, resumenes)):
            nt, nr = tok_40(tit), tok_15(res)
            if nt: bkt_tit[hashlib.md5(nt.encode()).hexdigest()].append(i)
            if nr: bkt_res[hashlib.md5(nr.encode()).hexdigest()].append(i)
        for bkt in (bkt_tit, bkt_res):
            for idxs in bkt.values():
                for j in idxs[1:]: dsu.union(idxs[0], j)

    # ── Paso 2: similitud de títulos (SequenceMatcher) ───────────────────────
    def _paso2_titulos_similares(self, titulos: List[str], dsu: DSU):
        norm = [normalize_title_for_comparison(t) for t in titulos]
        n    = len(norm)
        for i in range(n):
            if not norm[i]: continue
            for j in range(i+1, n):
                if not norm[j] or dsu.find(i) == dsu.find(j): continue
                if SequenceMatcher(None, norm[i], norm[j]).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    dsu.union(i, j)

    # ── Paso 3: clustering semántico ─────────────────────────────────────────
    def _paso3_semantico(self, textos: List[str], titulos: List[str],
                          indices_sueltos: List[int], dsu: DSU, pbar, p_start: float):
        """
        Agrupa los índices que aún están solos usando AgglomerativeClustering.
        UMBRAL_SUBTEMA controla directamente cuántos subtemas distintos habrá:
        - Más alto (0.88) → menos grupos, etiquetas más amplias
        - Más bajo (0.75)  → más grupos, etiquetas más específicas
        """
        if len(indices_sueltos) < 2: return
        BATCH = 400
        total_batches = max(1, len(indices_sueltos) // BATCH + 1)
        for b_num, b_start in enumerate(range(0, len(indices_sueltos), BATCH)):
            batch = indices_sueltos[b_start:b_start+BATCH]
            # Texto combinado: título con más peso que el resumen
            txts  = [f"{titulos[i][:150]} {titulos[i][:80]} {textos[i][:1000]}" for i in batch]
            embs  = get_embeddings_batch(txts)
            ok    = [(batch[k], e) for k, e in enumerate(embs) if e is not None]
            if len(ok) < 2: continue
            idxs, M = zip(*ok)
            sim = cosine_similarity(np.array(M))
            labels = AgglomerativeClustering(
                n_clusters=None,
                distance_threshold=1 - UMBRAL_SUBTEMA,
                metric='precomputed',
                linkage='average'
            ).fit(1 - sim).labels_
            g: Dict[int, List[int]] = defaultdict(list)
            for k, lbl in enumerate(labels): g[lbl].append(idxs[k])
            for cluster in g.values():
                if len(cluster) >= 2:
                    for j in cluster[1:]: dsu.union(cluster[0], j)
            pbar.progress(
                p_start + (1-p_start) * (b_num+1) / total_batches,
                f"Clustering semántico… lote {b_num+1}/{total_batches}"
            )

    # ── Paso 4: generar etiqueta para un grupo ────────────────────────────────
    def _generar_etiqueta(self, textos_grp: List[str], titulos_grp: List[str]) -> str:
        """
        Genera UNA etiqueta por grupo.
        Clave de caché = hash de títulos normalizados + ordenados.
        Si el mismo conjunto de títulos llega por cualquier ruta, produce la misma etiqueta.
        """
        titulos_norm = sorted(set(normalize_title_for_comparison(t) for t in titulos_grp if t))
        cache_key    = hashlib.md5("|".join(titulos_norm[:5]).encode()).hexdigest()
        if cache_key in self._cache: return self._cache[cache_key]

        # Keywords del contenido
        palabras = []
        for t in titulos_grp[:8]:
            palabras.extend(w for w in string_norm_label(t).split() if len(w) > 3)
        keywords = " · ".join(w for w, _ in Counter(palabras).most_common(6))

        # Muestra de títulos representativos (sin duplicados)
        titulos_muestra = list(dict.fromkeys(t[:100] for t in titulos_grp if t))[:6]

        prompt = (
            "Genera un SUBTEMA periodístico en español (3-5 palabras) que describa con precisión"
            " el asunto central de estas noticias.\n\n"
            f"TÍTULOS:\n" + "\n".join(f"  · {t}" for t in titulos_muestra) + "\n\n"
            f"PALABRAS CLAVE: {keywords}\n\n"
            "REGLAS:\n"
            "  - No uses el nombre de la empresa ni ciudades ni gentilicios\n"
            "  - No uses verbos vagos ('actividades', 'gestión', 'acciones')\n"
            "  - El subtema debe describir EL ASUNTO específico, no el actor\n"
            "  - Ejemplos correctos: 'Resultados Tercer Trimestre', 'Programa Becas Universitarias',"
            " 'Apertura Sucursal Centro', 'Sanción Regulatoria Financiera'\n"
            '  - Responde SOLO JSON: {"subtema":"..."}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role":"user","content":prompt}],
                max_tokens=40, temperature=0.0,   # temperatura 0 = máximo determinismo
                response_format={"type":"json_object"}
            )
            usage = resp.get('usage',{}) if isinstance(resp,dict) else getattr(resp,'usage',{})
            if usage:
                pt = usage.get('prompt_tokens') if isinstance(usage,dict) else getattr(usage,'prompt_tokens',0)
                ct = usage.get('completion_tokens') if isinstance(usage,dict) else getattr(usage,'completion_tokens',0)
                st.session_state['tokens_input']  += (pt or 0)
                st.session_state['tokens_output'] += (ct or 0)
            raw = json.loads(resp.choices[0].message.content).get("subtema","Varios")
            etiqueta = limpiar_tema_geografico(limpiar_tema(raw), self.marca, self.aliases)
        except:
            etiqueta = "Actividad Corporativa"

        self._cache[cache_key] = etiqueta
        return etiqueta

    # ── Método principal ──────────────────────────────────────────────────────
    def procesar_lote(self, col_resumen: pd.Series, pbar,
                      resumenes_puros: pd.Series, titulos_puros: pd.Series) -> List[str]:
        textos   = col_resumen.tolist()
        titulos  = titulos_puros.tolist()
        resumenes = resumenes_puros.tolist()
        n        = len(textos)

        # ── Fase 1: agrupación determinista ───────────────────────────────────
        pbar.progress(0.05, "Fase 1 · Agrupando noticias idénticas…")
        dsu = DSU(n)
        self._paso1_hash_exacto(titulos, resumenes, dsu)

        pbar.progress(0.15, "Fase 2 · Similitud de títulos…")
        self._paso2_titulos_similares(titulos, dsu)

        # Identificar sueltos para clustering semántico
        grupos_previos   = dsu.grupos(n)
        indices_sueltos  = [i for idxs in grupos_previos.values() if len(idxs)==1 for i in idxs]

        # ── Fase 3: clustering semántico sobre sueltos ────────────────────────
        pbar.progress(0.25, "Fase 3 · Clustering semántico…")
        self._paso3_semantico(textos, titulos, indices_sueltos, dsu, pbar, p_start=0.25)

        # Grupos finales
        grupos_finales = dsu.grupos(n)
        n_grupos       = len(grupos_finales)

        # ── Fase 4: generar UNA etiqueta por grupo y propagar ─────────────────
        pbar.progress(0.55, f"Fase 4 · Etiquetando {n_grupos} grupos…")
        mapa: Dict[int, str] = {}
        for k, (lid, idxs) in enumerate(grupos_finales.items()):
            if k % 15 == 0:
                pbar.progress(0.55 + 0.40*(k/n_grupos), f"Etiquetando grupo {k+1}/{n_grupos}…")
            etiqueta = self._generar_etiqueta([textos[i] for i in idxs], [titulos[i] for i in idxs])
            # GARANTÍA: todos los índices del mismo grupo reciben exactamente la misma cadena
            for i in idxs: mapa[i] = etiqueta

        subtemas = [mapa.get(i, "Varios") for i in range(n)]
        pbar.progress(1.0, f"✓ {len(set(subtemas))} subtemas en {n_grupos} grupos")
        st.info(f"Subtemas únicos: **{len(set(subtemas))}** · Grupos semánticos: **{n_grupos}**")
        return subtemas


# ======================================
# CONSOLIDACIÓN DE TEMAS — también flujo invertido
# ======================================
def consolidar_temas(subtemas: List[str], textos: List[str], pbar) -> List[str]:
    """
    FLUJO:
    1. Calcular centroide semántico de cada subtema (media de embeddings de sus textos)
    2. Clustering de centroides con umbral UMBRAL_TEMA
    3. Generar UNA etiqueta de tema por cluster de subtemas
    4. Propagar: cada subtema mapea a su tema

    Esto produce exactamente N clusters de temas, donde N depende solo de UMBRAL_TEMA.
    No hay etiquetas generadas dos veces.
    """
    pbar.progress(0.05, "Calculando centroides de subtemas…")
    df = pd.DataFrame({'subtema': subtemas, 'texto': textos})
    unique_subs = list(df['subtema'].unique())

    if len(unique_subs) <= 1:
        pbar.progress(1.0, "Un solo tema")
        return subtemas

    # Embeddings de todos los textos
    todos_embs = get_embeddings_batch(textos)

    # Centroide semántico por subtema (hasta 40 textos de muestra)
    centroids: Dict[str, np.ndarray] = {}
    for sub in unique_subs:
        idxs   = df.index[df['subtema'] == sub].tolist()[:40]
        vecs   = [todos_embs[i] for i in idxs if todos_embs[i] is not None]
        if vecs: centroids[sub] = np.mean(vecs, axis=0)

    valid_subs = [s for s in unique_subs if s in centroids]
    if len(valid_subs) < 2:
        pbar.progress(1.0, "Sin agrupación posible")
        return subtemas

    pbar.progress(0.45, "Clustering de subtemas en temas…")
    M      = np.array([centroids[s] for s in valid_subs])
    sim    = cosine_similarity(M)

    # NUM_TEMAS_MAX como tope, pero si el umbral produce menos → respetarlo
    n_clusters = min(NUM_TEMAS_MAX, len(valid_subs))
    clustering  = AgglomerativeClustering(
        n_clusters=n_clusters, metric='precomputed', linkage='average'
    ).fit(1 - sim)

    # Por cada cluster de subtemas, generar UNA etiqueta de tema
    clusters_subs: Dict[int, List[str]] = defaultdict(list)
    for i, lbl in enumerate(clustering.labels_): clusters_subs[lbl].append(valid_subs[i])

    mapa_tema: Dict[str, str] = {}
    total_clusters = len(clusters_subs)

    for k, (cid, lista_subs) in enumerate(clusters_subs.items()):
        pbar.progress(0.55 + 0.40*(k/total_clusters), f"Generando tema {k+1}/{total_clusters}…")

        # Muestra de títulos representativos del cluster
        titulos_muestra = lista_subs[:8]
        prompt = (
            "Genera UNA categoría temática general (2-3 palabras, en español) para agrupar"
            " estos subtemas periodísticos.\n\n"
            f"SUBTEMAS: {', '.join(titulos_muestra)}\n\n"
            "REGLAS:\n"
            "  - Sin nombres de empresas ni ciudades\n"
            "  - Sin verbos ni artículos iniciales\n"
            "  - Sustantivos o adjetivo+sustantivo\n"
            "  - Ejemplos: 'Resultados Financieros', 'Sostenibilidad Ambiental',"
            " 'Innovación Tecnológica', 'Responsabilidad Social'\n"
            "  - Responde SOLO el nombre del tema, sin explicaciones"
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role":"user","content":prompt}],
                max_tokens=12, temperature=0.0
            )
            usage = resp.get('usage',{}) if isinstance(resp,dict) else getattr(resp,'usage',{})
            if usage:
                pt = usage.get('prompt_tokens') if isinstance(usage,dict) else getattr(usage,'prompt_tokens',0)
                ct = usage.get('completion_tokens') if isinstance(usage,dict) else getattr(usage,'completion_tokens',0)
                st.session_state['tokens_input']  += (pt or 0)
                st.session_state['tokens_output'] += (ct or 0)
            nombre = limpiar_tema(resp.choices[0].message.content.strip().replace('"','').replace('.',''))
        except:
            nombre = lista_subs[0]

        for sub in lista_subs: mapa_tema[sub] = nombre

    temas_final = [mapa_tema.get(sub, sub) for sub in subtemas]
    n_temas     = len(set(temas_final))
    st.info(f"Temas consolidados: **{n_temas}** (máximo configurado: {NUM_TEMAS_MAX})")
    pbar.progress(1.0, "✓ Temas finalizados")
    return temas_final


def analizar_temas_con_pkl(textos, pkl_file):
    try:
        pipeline = joblib.load(pkl_file)
        return [str(p) for p in pipeline.predict(textos)]
    except Exception as e:
        st.error(f"Error pkl temas: {e}"); return None

# ======================================
# Duplicados y Excel (sin cambios lógicos)
# ======================================
def detectar_duplicados_avanzado(rows: List[Dict], key_map: Dict[str, str]) -> List[Dict]:
    processed = deepcopy(rows)
    seen_url, seen_bcast = {}, {}
    title_buckets: Dict[tuple, List[int]] = defaultdict(list)
    for i, row in enumerate(processed):
        if row.get("is_duplicate"): continue
        tipo    = normalizar_tipo_medio(str(row.get(key_map.get("tipodemedio",""))))
        mencion = norm_key(row.get(key_map.get("menciones","")))
        medio   = norm_key(row.get(key_map.get("medio","")))
        if tipo == "Internet":
            li = row.get(key_map.get("link_nota",{})) or {}
            url = li.get("url") if isinstance(li, dict) else None
            if url and mencion:
                k = (url, mencion)
                if k in seen_url: row["is_duplicate"]=True; row["idduplicada"]=processed[seen_url[k]].get(key_map.get("idnoticia",""),""); continue
                seen_url[k] = i
            if medio and mencion: title_buckets[(medio, mencion)].append(i)
        elif tipo in ("Radio","Televisión"):
            hora = str(row.get(key_map.get("hora",""),"")).strip()
            if mencion and medio and hora:
                k = (mencion, medio, hora)
                if k in seen_bcast: row["is_duplicate"]=True; row["idduplicada"]=processed[seen_bcast[k]].get(key_map.get("idnoticia",""),"")
                else: seen_bcast[k] = i
    for idxs in title_buckets.values():
        if len(idxs) < 2: continue
        for i in range(len(idxs)):
            for j in range(i+1, len(idxs)):
                a, b = idxs[i], idxs[j]
                if processed[a].get("is_duplicate") or processed[b].get("is_duplicate"): continue
                ta = normalize_title_for_comparison(processed[a].get(key_map.get("titulo","")))
                tb = normalize_title_for_comparison(processed[b].get(key_map.get("titulo","")))
                if ta and tb and SequenceMatcher(None,ta,tb).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    if len(ta) < len(tb): processed[a]["is_duplicate"]=True; processed[a]["idduplicada"]=processed[b].get(key_map.get("idnoticia",""),"")
                    else:                 processed[b]["is_duplicate"]=True; processed[b]["idduplicada"]=processed[a].get(key_map.get("idnoticia",""),"")
    return processed

def run_dossier_logic(sheet):
    headers   = [c.value for c in sheet[1] if c.value]
    norm_keys = [norm_key(h) for h in headers]
    key_map   = {nk: nk for nk in norm_keys}
    key_map.update({
        "titulo":      norm_key("Titulo"),
        "resumen":     norm_key("Resumen - Aclaracion"),
        "menciones":   norm_key("Menciones - Empresa"),
        "medio":       norm_key("Medio"),
        "tonoiai":     norm_key("Tono IA"),
        "tema":        norm_key("Tema"),
        "subtema":     norm_key("Subtema"),
        "idnoticia":   norm_key("ID Noticia"),
        "idduplicada": norm_key("ID duplicada"),
        "tipodemedio": norm_key("Tipo de Medio"),
        "hora":        norm_key("Hora"),
        "link_nota":   norm_key("Link Nota"),
        "link_streaming": norm_key("Link (Streaming - Imagen)"),
        "region":      norm_key("Region"),
    })
    rows, split_rows = [], []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row): continue
        rows.append({norm_keys[i]: c for i, c in enumerate(row) if i < len(norm_keys)})
    for r_cells in rows:
        base = {
            k: (extract_link(v) if k in (key_map["link_nota"], key_map["link_streaming"]) else v.value)
            for k, v in r_cells.items()
        }
        if key_map.get("tipodemedio") in base:
            base[key_map["tipodemedio"]] = normalizar_tipo_medio(base.get(key_map["tipodemedio"]))
        ml = [m.strip() for m in str(base.get(key_map["menciones"],"")).split(";") if m.strip()]
        for m in ml or [None]:
            nr = deepcopy(base)
            if m: nr[key_map["menciones"]] = m
            split_rows.append(nr)
    for idx, row in enumerate(split_rows): row.update({"original_index": idx, "is_duplicate": False})
    processed = detectar_duplicados_avanzado(split_rows, key_map)
    for row in processed:
        if row["is_duplicate"]:
            row.update({key_map["tonoiai"]:"Duplicada", key_map["tema"]:"Duplicada", key_map["subtema"]:"Duplicada"})
    return processed, key_map

def fix_links_by_media_type(row, key_map):
    tkey, ln_key, ls_key = key_map.get("tipodemedio"), key_map.get("link_nota"), key_map.get("link_streaming")
    if not (tkey and ln_key and ls_key): return
    tipo = row.get(tkey,"")
    ln   = row.get(ln_key) or {"value":"","url":None}
    ls   = row.get(ls_key) or {"value":"","url":None}
    hurl = lambda x: isinstance(x,dict) and bool(x.get("url"))
    if   tipo in ("Radio","Televisión"):  row[ls_key] = {"value":"","url":None}
    elif tipo == "Internet":              row[ln_key], row[ls_key] = ls, ln
    elif tipo in ("Prensa","Revista"):
        if not hurl(ln) and hurl(ls): row[ln_key] = ls
        row[ls_key] = {"value":"","url":None}

def generate_output_excel(rows, key_map):
    wb    = Workbook(); ws = wb.active; ws.title = "Resultado"
    ORDER = ["ID Noticia","Fecha","Hora","Medio","Tipo de Medio","Region","Seccion - Programa",
             "Titulo","Autor - Conductor","Nro. Pagina","Dimension","Duracion - Nro. Caracteres",
             "CPE","Audiencia","Tier","Tono","Tono IA","Tema","Subtema","Link Nota",
             "Resumen - Aclaracion","Link (Streaming - Imagen)","Menciones - Empresa","ID duplicada"]
    NUM   = {"ID Noticia","Nro. Pagina","Dimension","Duracion - Nro. Caracteres","CPE","Tier","Audiencia"}
    ws.append(ORDER)
    ls = NamedStyle(name="HL", font=Font(color="0000FF", underline="single"))
    if "HL" not in wb.style_names: wb.add_named_style(ls)
    for row in rows:
        tk = key_map.get("titulo")
        if tk and tk in row: row[tk] = clean_title_for_output(row.get(tk))
        rk = key_map.get("resumen")
        if rk and rk in row: row[rk] = corregir_texto(row.get(rk))
        out, links = [], {}
        for ci, h in enumerate(ORDER, 1):
            dk  = key_map.get(norm_key(h), norm_key(h))
            val = row.get(dk)
            cv  = None
            if h in NUM:
                try: cv = float(val) if val is not None and str(val).strip() != "" else None
                except: cv = str(val) if val is not None else None
            elif isinstance(val, dict) and "url" in val:
                cv = val.get("value","Link")
                if val.get("url"): links[ci] = val["url"]
            elif val is not None: cv = str(val)
            out.append(cv)
        ws.append(out)
        for ci, url in links.items():
            cell = ws.cell(row=ws.max_row, column=ci)
            cell.hyperlink = url; cell.style = "HL"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ======================================
# Proceso principal
# ======================================
async def run_full_process_async(dossier_file, region_file, internet_file, brand_name,
                                  brand_aliases, tono_pkl_file, tema_pkl_file, analysis_mode):
    st.session_state.update({'tokens_input':0,'tokens_output':0,'tokens_embedding':0})
    t0 = time.time()
    if "API" in analysis_mode:
        try: openai.api_key = st.secrets["OPENAI_API_KEY"]; openai.aiosession.set(None)
        except: st.error("OPENAI_API_KEY no encontrado."); st.stop()

    with st.status("Paso 1 · Limpieza y duplicados", expanded=True) as s:
        rows, key_map = run_dossier_logic(load_workbook(dossier_file, data_only=True).active)
        s.update(label="✓ Paso 1 · Limpieza completada", state="complete")

    with st.status("Paso 2 · Mapeos", expanded=True) as s:
        df_r = pd.read_excel(region_file)
        rmap = {str(k).lower().strip(): v for k,v in pd.Series(df_r.iloc[:,1].values, index=df_r.iloc[:,0]).to_dict().items()}
        df_i = pd.read_excel(internet_file)
        imap = {str(k).lower().strip(): v for k,v in pd.Series(df_i.iloc[:,1].values, index=df_i.iloc[:,0]).to_dict().items()}
        for row in rows:
            mk = str(row.get(key_map.get("medio",""),"")).lower().strip()
            row[key_map.get("region")] = rmap.get(mk,"N/A")
            if mk in imap: row[key_map.get("medio")]=imap[mk]; row[key_map.get("tipodemedio")]="Internet"
            fix_links_by_media_type(row, key_map)
        s.update(label="✓ Paso 2 · Mapeos aplicados", state="complete")

    gc.collect()
    to_analyze = [r for r in rows if not r.get("is_duplicate")]

    if to_analyze:
        df = pd.DataFrame(to_analyze)
        df["_txt"] = df[key_map["titulo"]].fillna("").astype(str) + ". " + df[key_map["resumen"]].fillna("").astype(str)

        with st.status("Paso 3 · Tono", expanded=True) as s:
            pb = st.progress(0)
            if "PKL" in analysis_mode and tono_pkl_file:
                res = analizar_tono_con_pkl(df["_txt"].tolist(), tono_pkl_file)
                if res is None: st.stop()
            elif "API" in analysis_mode:
                res = await ClasificadorTono(brand_name, brand_aliases).procesar_lote_async(
                    df["_txt"], pb, df[key_map["resumen"]], df[key_map["titulo"]]
                )
            else: res = [{"tono":"N/A"}] * len(to_analyze)
            df[key_map["tonoiai"]] = [r["tono"] for r in res]
            s.update(label="✓ Paso 3 · Tono analizado", state="complete")

        with st.status("Paso 4 · Tema y Subtema", expanded=True) as s:
            pb = st.progress(0)
            if "Solo Modelos PKL" in analysis_mode:
                subtemas = ["N/A"] * len(to_analyze)
                temas    = ["N/A"] * len(to_analyze)
            else:
                subtemas = ClasificadorSubtema(brand_name, brand_aliases).procesar_lote(
                    df["_txt"], pb, df[key_map["resumen"]], df[key_map["titulo"]]
                )
                temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb)

            df[key_map["subtema"]] = subtemas
            if "PKL" in analysis_mode and tema_pkl_file:
                tp = analizar_temas_con_pkl(df["_txt"].tolist(), tema_pkl_file)
                if tp: df[key_map["tema"]] = tp
            else: df[key_map["tema"]] = temas
            s.update(label="✓ Paso 4 · Clasificación completada", state="complete")

        rmap2 = df.set_index("original_index").to_dict("index")
        for row in rows:
            if not row.get("is_duplicate"): row.update(rmap2.get(row["original_index"],{}))

    gc.collect()
    ci  = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co  = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    cem = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M

    with st.status("Paso 5 · Generando informe", expanded=True) as s:
        st.session_state["output_data"]     = generate_output_excel(rows, key_map)
        st.session_state["output_filename"] = f"Informe_IA_{brand_name.replace(' ','_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.session_state["processing_complete"] = True
        st.session_state.update({
            "brand_name":      brand_name,
            "brand_aliases":   brand_aliases,
            "total_rows":      len(rows),
            "unique_rows":     len(to_analyze),
            "duplicates":      len(rows) - len(to_analyze),
            "process_duration":f"{time.time()-t0:.0f}s",
            "process_cost":    f"${ci+co+cem:.4f} USD",
        })
        s.update(label="✓ Proceso completado", state="complete")

# ======================================
# Análisis Rápido
# ======================================
async def run_quick_analysis_async(df, title_col, summary_col, brand_name, aliases):
    st.session_state.update({'tokens_input':0,'tokens_output':0,'tokens_embedding':0})
    df['_txt'] = df[title_col].fillna('').astype(str) + ". " + df[summary_col].fillna('').astype(str)
    with st.status("Paso 1/2 · Tono…", expanded=True) as s:
        pb  = st.progress(0)
        res = await ClasificadorTono(brand_name, aliases).procesar_lote_async(
            df["_txt"], pb, df[summary_col].fillna(''), df[title_col].fillna('')
        )
        df['Tono IA'] = [r["tono"] for r in res]
        s.update(label="✓ Paso 1/2 · Tono analizado", state="complete")
    with st.status("Paso 2/2 · Tema y Subtema…", expanded=True) as s:
        pb       = st.progress(0)
        subtemas = ClasificadorSubtema(brand_name, aliases).procesar_lote(
            df["_txt"], pb, df[summary_col].fillna(''), df[title_col].fillna('')
        )
        df['Subtema'] = subtemas
        temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb)
        df['Tema'] = temas
        s.update(label="✓ Paso 2/2 · Clasificación completada", state="complete")
    df.drop(columns=['_txt'], inplace=True)
    ci  = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co  = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    cem = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    st.session_state['quick_cost'] = f"${ci+co+cem:.4f} USD"
    return df

def gen_quick_excel(df) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w: df.to_excel(w, index=False, sheet_name='Analisis')
    return buf.getvalue()

def render_quick_tab():
    st.markdown('<div class="sec-label">Análisis rápido</div>', unsafe_allow_html=True)
    if 'quick_result' in st.session_state:
        st.markdown('<div class="success-banner"><div class="success-title">Análisis completado</div><div class="success-sub">Los resultados están listos para descargar</div></div>', unsafe_allow_html=True)
        st.metric("Costo estimado", st.session_state.get('quick_cost',"$0.00"))
        st.dataframe(st.session_state.quick_result.head(10), use_container_width=True)
        st.download_button("Descargar resultados →", data=gen_quick_excel(st.session_state.quick_result),
                           file_name="Analisis_Rapido_IA.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")
        if st.button("← Nuevo análisis"):
            for k in ('quick_result','quick_df','quick_name','quick_cost'):
                if k in st.session_state: del st.session_state[k]
            st.rerun()
        return
    if 'quick_df' not in st.session_state:
        st.markdown("Sube un archivo Excel con columnas de título y resumen.")
        f = st.file_uploader("Archivo Excel", type=["xlsx"], label_visibility="collapsed", key="qu")
        if f:
            try: st.session_state.quick_df=pd.read_excel(f); st.session_state.quick_name=f.name; st.rerun()
            except Exception as e: st.error(f"Error: {e}"); st.stop()
    else:
        st.success(f"Archivo **{st.session_state.quick_name}** listo")
        with st.form("qf"):
            cols = st.session_state.quick_df.columns.tolist()
            c1, c2 = st.columns(2)
            tc = c1.selectbox("Columna Título",   cols, 0)
            sc = c2.selectbox("Columna Resumen",  cols, 1 if len(cols)>1 else 0)
            st.write("---")
            bn  = st.text_input("Marca principal", placeholder="Ej: Bancolombia")
            bat = st.text_area("Alias (sep. ;)",    placeholder="Ej: Grupo Bancolombia;Ban", height=70)
            if st.form_submit_button("Analizar →", use_container_width=True, type="primary"):
                if not bn: st.error("Indica la marca.")
                else:
                    try: openai.api_key=st.secrets["OPENAI_API_KEY"]; openai.aiosession.set(None)
                    except: st.error("OPENAI_API_KEY no encontrada."); st.stop()
                    aliases = [a.strip() for a in bat.split(";") if a.strip()]
                    with st.spinner("Analizando…"):
                        st.session_state.quick_result = asyncio.run(
                            run_quick_analysis_async(st.session_state.quick_df.copy(), tc, sc, bn, aliases)
                        )
                    st.rerun()
        if st.button("← Otro archivo"):
            for k in ('quick_df','quick_name','quick_result','quick_cost'):
                if k in st.session_state: del st.session_state[k]
            st.rerun()

# ======================================
# Main
# ======================================
def main():
    load_custom_css()
    if not check_password(): return

    st.markdown("""
    <div class="app-header">
        <div class="app-header-mark">◈</div>
        <div class="app-header-text">
            <div class="app-header-title">Sistema de Análisis de Noticias</div>
            <div class="app-header-version">v9.0 · cluster-first · label-once · OpenAI + PKL</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["Análisis Completo", "Análisis Rápido"])

    with tab1:
        if not st.session_state.get("processing_complete", False):
            with st.form("main_form"):
                st.markdown('<div class="sec-label">Archivos de entrada</div>', unsafe_allow_html=True)
                c1, c2, c3 = st.columns(3)
                df_file  = c1.file_uploader("Dossier (.xlsx)",  type=["xlsx"])
                reg_file = c2.file_uploader("Región (.xlsx)",   type=["xlsx"])
                int_file = c3.file_uploader("Internet (.xlsx)", type=["xlsx"])

                st.markdown('<div class="sec-label">Marca</div>', unsafe_allow_html=True)
                bn  = st.text_input("Nombre principal", placeholder="Ej: Bancolombia", key="bn")
                bat = st.text_area("Alias (sep. ;)",     placeholder="Ej: Grupo Bancolombia;Ban", height=70, key="ba")

                st.markdown('<div class="sec-label">Modo de análisis</div>', unsafe_allow_html=True)
                mode = st.radio("", ["Híbrido (PKL + API)","Solo Modelos PKL","API de OpenAI"], index=0, key="mode")
                tpkl, epkl = None, None
                if "PKL" in mode:
                    p1, p2 = st.columns(2)
                    tpkl = p1.file_uploader("pipeline_sentimiento.pkl", type=["pkl"])
                    epkl = p2.file_uploader("pipeline_tema.pkl",        type=["pkl"])

                # Información de parámetros de clustering
                st.markdown(f"""
                <div class="cluster-info">
                  <b>Parámetros de clustering actuales</b><br>
                  UMBRAL_SUBTEMA = {UMBRAL_SUBTEMA} &nbsp;·&nbsp;
                  UMBRAL_TEMA = {UMBRAL_TEMA} &nbsp;·&nbsp;
                  NUM_TEMAS_MAX = {NUM_TEMAS_MAX}<br>
                  <span style="color:#52504e">
                  Aumenta UMBRAL_SUBTEMA → menos subtemas, más generales &nbsp;|&nbsp;
                  Disminúyelo → más subtemas, más específicos
                  </span>
                </div>
                """, unsafe_allow_html=True)

                if st.form_submit_button("Iniciar análisis →", use_container_width=True, type="primary"):
                    if not all([df_file, reg_file, int_file, bn.strip()]):
                        st.error("Completa todos los campos y archivos.")
                    else:
                        aliases = [a.strip() for a in bat.split(";") if a.strip()]
                        asyncio.run(run_full_process_async(df_file, reg_file, int_file, bn, aliases, tpkl, epkl, mode))
                        st.rerun()
        else:
            total = st.session_state.total_rows
            uniq  = st.session_state.unique_rows
            dups  = st.session_state.duplicates
            dur   = st.session_state.process_duration
            cost  = st.session_state.get("process_cost","$0.00")

            st.markdown('<div class="success-banner"><div class="success-title">Análisis completado</div><div class="success-sub">El informe está listo para descargar</div></div>', unsafe_allow_html=True)
            st.markdown(f"""
            <div class="metrics-row">
              <div class="metric-card"><div class="metric-val" style="color:var(--text)">{total}</div><div class="metric-lbl">Total filas</div></div>
              <div class="metric-card"><div class="metric-val" style="color:var(--green)">{uniq}</div><div class="metric-lbl">Únicas</div></div>
              <div class="metric-card"><div class="metric-val" style="color:var(--gold)">{dups}</div><div class="metric-lbl">Duplicados</div></div>
              <div class="metric-card"><div class="metric-val" style="color:var(--blue)">{dur}</div><div class="metric-lbl">Tiempo</div></div>
              <div class="metric-card"><div class="metric-val" style="color:var(--red)">{cost}</div><div class="metric-lbl">Costo est.</div></div>
            </div>
            """, unsafe_allow_html=True)
            st.download_button("Descargar informe →",
                               data=st.session_state.output_data,
                               file_name=st.session_state.output_filename,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True, type="primary")
            if st.button("← Nuevo análisis", use_container_width=True):
                pwd = st.session_state.get("password_correct")
                st.session_state.clear()
                st.session_state.password_correct = pwd
                st.rerun()

    with tab2: render_quick_tab()

    st.markdown('<div class="footer">v9.0.0 · Realizado por Johnathan Cortés ©</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    main()

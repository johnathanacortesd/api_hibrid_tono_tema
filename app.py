# ======================================
# Importaciones
# ======================================
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle
from collections import defaultdict, Counter
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import openai
import re
import time
from unidecode import unidecode
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
import json
import asyncio
import hashlib
from typing import List, Dict, Tuple, Optional, Any
import joblib
import gc
import requests
import os
import zipfile
import xml.etree.ElementTree as ET
import html
import hashlib

# ======================================
# Configuracion general
# ======================================
st.set_page_config(
    page_title="Análisis de Noticias · IA",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

OPENAI_MODEL_EMBEDDING     = "text-embedding-3-small"
OPENAI_MODEL_CLASIFICACION = "gpt-4.1-nano-2025-04-14"

CONCURRENT_REQUESTS          = 50
SIMILARITY_THRESHOLD_TONO    = 0.82
SIMILARITY_THRESHOLD_TITULOS = 0.93

# ── Umbrales base (corpus grande ≥ 20 noticias) ──────────────────────────────
UMBRAL_SUBTEMA = 0.78
UMBRAL_TEMA    = 0.72
NUM_TEMAS_MAX  = 15

UMBRAL_DEDUP_LABEL           = 0.78
UMBRAL_FUSION_SUBTEMAS       = 0.78
UMBRAL_FUSION_INTERGRUPO     = 0.84
MAX_ITER_FUSION              = 5

UMBRAL_MIN_PERTENENCIA_SUBTEMA = 0.60
UMBRAL_MIN_PERTENENCIA_TEMA    = 0.52

UMBRAL_COHERENCIA_ETIQUETA   = 0.35

MAX_GRUPO_ETIQUETA           = 40

# ── Umbrales mínimos de similitud REAL para agrupar ──────────────────────────
# Una noticia sólo entra en un grupo si su similitud con el representante
# supera este umbral, sin importar lo que diga el clustering.
SIM_MINIMA_AGRUPACION_SUBTEMA = 0.82   # nunca agrupar por debajo de esto
SIM_MINIMA_KEYWORDS_RARAS     = 0.78   # para _paso2b_keywords
SIM_MINIMA_FUSION_INTER       = 0.88   # más estricto que el base

PRICE_INPUT_1M     = 0.10
PRICE_OUTPUT_1M    = 0.40
PRICE_EMBEDDING_1M = 0.02

if 'tokens_input'     not in st.session_state: st.session_state['tokens_input']     = 0
if 'tokens_output'    not in st.session_state: st.session_state['tokens_output']    = 0
if 'tokens_embedding' not in st.session_state: st.session_state['tokens_embedding'] = 0

CIUDADES_COLOMBIA = {
    "bogotá","bogota","medellín","medellin","cali","barranquilla","cartagena",
    "cúcuta","cucuta","bucaramanga","pereira","manizales","armenia","ibagué",
    "ibague","villavicencio","montería","monteria","neiva","pasto","valledupar",
    "popayán","popayan","tunja","florencia","sincelejo","riohacha","yopal",
    "santa marta","santamarta","quibdó","quibdo","leticia","mocoa","mitú","mitu",
    "puerto carreño","inírida","inirida","san josé del guaviare","antioquia",
    "atlántico","atlantico","bolívar","bolivar","boyacá","boyaca","caldas",
    "caquetá","caqueta","casanare","cauca","cesar","chocó","choco","córdoba",
    "cordoba","cundinamarca","guainía","guainia","guaviare","huila","la guajira",
    "magdalena","meta","nariño","narino","norte de santander","putumayo",
    "quindío","quindio","risaralda","san andrés","san andres","santander",
    "sucre","tolima","valle del cauca","vaupés","vaupes","vichada",
}
GENTILICIOS_COLOMBIA = {
    "bogotano","bogotanos","bogotana","bogotanas","capitalino","capitalinos",
    "capitalina","capitalinas","antioqueño","antioqueños","antioqueña",
    "antioqueñas","paisa","paisas","medellense","medellenses","caleño",
    "caleños","caleña","caleñas","valluno","vallunos","valluna","vallunas",
    "vallecaucano","vallecaucanos","barranquillero","barranquilleros",
    "cartagenero","cartageneros","costeño","costeños","costeña","costeñas",
    "cucuteño","cucuteños","bumangués","santandereano","santandereanos",
    "boyacense","boyacenses","tolimense","tolimenses","huilense","huilenses",
    "nariñense","nariñenses","pastuso","pastusas","cordobés","cordobeses",
    "cauca","caucano","caucanos","chocoano","chocoanos","casanareño",
    "casanareños","caqueteño","caqueteños","guajiro","guajiros","llanero",
    "llaneros","amazonense","amazonenses","colombiano","colombianos",
    "colombiana","colombianas",
}

STOPWORDS_ES = set("""
a ante bajo cabe con contra de desde durante en entre hacia hasta mediante
para por segun sin so sobre tras y o u e la el los las un una unos unas lo
al del se su sus le les mi mis tu tus nuestro nuestros vuestra vuestras este
esta estos estas ese esa esos esas aquel aquella aquellos aquellas que cual
cuales quien quienes cuyo cuya cuyos cuyas como cuando donde cual es son fue
fueron era eran sera seran seria serian he ha han habia habian hay hubo habra
habria estoy esta estan estaba estaban estamos estan estar estare estaria
estuvieron estarian estuvo asi ya mas menos tan tanto cada muy todo toda todos
todas ser haber hacer tener poder deber ir dar ver saber querer llegar pasar
encontrar creer decir poner salir volver seguir llevar sentir cambiar
""".split())

_TRAILING_INCOMPLETE = {
    "de","del","la","el","los","las","un","una","unos","unas","al","su","sus",
    "en","con","sin","por","para","sobre","ante","bajo","contra","desde",
    "entre","hacia","hasta","mediante","tras","y","o","u","e","lo","que","se",
    "como","donde","cuando","cual","cuyo","cuya","cuyos","cuyas",
    "este","esta","estos","estas","ese","esa","esos","esas",
    "aquel","aquella","aquellos","aquellas","cada","todo","toda","todos","todas",
    "otro","otra","otros","otras","nuevo","nueva","nuevos","nuevas",
    "gran","grandes","mayor","mayores","menor","menores","mejor","mejores",
    "peor","peores","primer","primera","segundo","segunda","tercer","tercera",
    "más","mas","muy","tan","tanto","tanta","tantos","tantas",
    "mi","mis","tu","tus","nuestro","nuestra","nuestros","nuestras",
    "a","ha","he","ser","estar","haber","hacer","tener","poder","deber",
    "ir","dar","ver","saber","querer","llegar","pasar","decir","poner",
}

POS_VARIANTS = [
    r"lanz(a(r|ra|ria|o|on|an|ando)?|amiento)s?", r"prepar(a|ando)",
    r"nuev[oa]\s+(servicio|tienda|plataforma|app|aplicacion|funcion|canal|portal|producto|iniciativa|proyecto)",
    r"apertur(a|ar|ara|o|an)",r"estren(a|o|ara|an|ando)",r"habilit(a|o|ara|an|ando)",
    r"disponible",r"mejor(a|o|an|ando)",r"optimiza|amplia|expande",
    r"alianz(a|as)|acuerd(o|a|os)|convenio(s)?|memorando(s)?|joint\s+venture|colaboraci[oó]n(es)?|asociaci[oó]n(es)?|partnership(s)?|fusi[oó]n(es)?|integraci[oó]n(es)?",
    r"crecimi?ento|aument(a|o|an|ando)",r"gananci(a|as)|utilidad(es)?|benefici(o|os)",
    r"expansion|crece|crecer",r"inversion|invierte|invertir",
    r"innova(cion|dor|ndo)|moderniza",r"exito(so|sa)?|logr(o|os|a|an|ando)",
    r"reconoci(miento|do|da)|premi(o|os|ada)",r"lidera(zgo)?|lider",
    r"consolida|fortalece",r"oportunidad(es)?|potencial",r"solucion(es)?|resuelve",
    r"eficien(te|cia)",r"calidad|excelencia",r"satisfaccion|complace",
    r"confianza|credibilidad",r"sostenible|responsable",r"compromiso|apoya|apoyar",
    r"patrocin(io|a|ador|an|ando)|auspic(ia|io|iador)",
    r"gana(r|dor|dora|ndo)?|triunf(a|o|ar|ando)",
    r"destaca(r|do|da|ndo)?",r"supera(r|ndo|cion)?",r"record|hito|milestone",
    r"avanza(r|do|da|ndo)?",r"benefici(a|o|ando|ar|ando)",r"importante(s)?",
    r"prioridad",r"bienestar",r"garantizar",r"seguridad",r"atencion",
    r"expres(o|ó|ando)",r"señala(r|do|ando)",r"ratific(a|o|ando|ar)",
]
NEG_VARIANTS = [
    r"demanda|denuncia|sanciona|multa|investiga|critica",
    r"cae|baja|pierde|crisis|quiebra|default",
    r"fraude|escandalo|irregularidad",
    r"fall(a|o|os)|interrumpe|suspende|cierra|renuncia|huelga",
    r"filtracion|ataque|phishing|hackeo|incumple|boicot|queja|reclamo|deteriora",
    r"problema(s|tica|ico)?|dificultad(es)?",r"retras(o|a|ar|ado)",
    r"perdida(s)?|deficit",
    r"conflict(o|os)?|disputa(s)?",r"rechaz(a|o|ar|ado)",r"negativ(o|a|os|as)",
    r"preocupa(cion|nte|do)?",r"alarma(nte)?|alerta",r"riesgo(s)?|amenaza(s)?",
]
CRISIS_KEYWORDS = re.compile(
    r"\b(crisis|emergencia|desastre|deslizamiento|inundaci[oó]n|afectaciones|"
    r"damnificados|tragedia|zozobra|alerta)\b", re.IGNORECASE)
RESPONSE_VERBS = re.compile(
    r"\b(atiend(e|en|iendo)|activ(a|o|ando)|decret(a|o|ando)|"
    r"responde(r|iendo)|trabaj(a|ando)|lidera(ndo)?|enfrenta(ndo)?|"
    r"gestiona(ndo)?|declar(o|a|ando)|anunci(a|o|ando))\b", re.IGNORECASE)
POS_PATTERNS = [re.compile(rf"\b(?:{p})\b", re.IGNORECASE) for p in POS_VARIANTS]
NEG_PATTERNS = [re.compile(rf"\b(?:{p})\b", re.IGNORECASE) for p in NEG_VARIANTS]

_PATRON_TITULAR = re.compile(
    r"^(nuevo|nueva|anuncia|lanza|presenta|inaugura|llega|abre|inicia|"
    r"logra|alcanza|supera|confirma|destaca|revela|señala|advierte|"
    r"lanzamiento|anuncio|apertura|inicio|presentacion|presentación)\b",
    re.IGNORECASE
)
_PATRON_ESTADO = re.compile(
    r"\b(calma|caos|urgente|hoy|ya|ahora|ayer|mañana|nuevo|nueva|"
    r"gran|grande|importante|especial|exclusivo)\s*$",
    re.IGNORECASE
)

_TILDE_MAP = {
    "regulacion":"regulación","regulaciones":"regulaciones",
    "innovacion":"innovación","innovaciones":"innovaciones",
    "tecnologia":"tecnología","tecnologias":"tecnologías",
    "tecnologica":"tecnológica","tecnologico":"tecnológico",
    "educacion":"educación","gestion":"gestión",
    "administracion":"administración","informacion":"información",
    "comunicacion":"comunicación","comunicaciones":"comunicaciones",
    "operacion":"operación","operaciones":"operaciones",
    "inversion":"inversión","inversiones":"inversiones",
    "expansion":"expansión","adquisicion":"adquisición",
    "adquisiciones":"adquisiciones","fusion":"fusión",
    "fusiones":"fusiones","transicion":"transición",
    "transformacion":"transformación","digitalizacion":"digitalización",
    "automatizacion":"automatización","modernizacion":"modernización",
    "optimizacion":"optimización","implementacion":"implementación",
    "evaluacion":"evaluación","planificacion":"planificación",
    "organizacion":"organización","atencion":"atención",
    "produccion":"producción","construccion":"construcción",
    "distribucion":"distribución","exportacion":"exportación",
    "importacion":"importación","comercializacion":"comercialización",
    "negociacion":"negociación","negociaciones":"negociaciones",
    "participacion":"participación","colaboracion":"colaboración",
    "asociacion":"asociación","integracion":"integración",
    "relacion":"relación","relaciones":"relaciones",
    "situacion":"situación","condicion":"condición",
    "condiciones":"condiciones","solucion":"solución",
    "soluciones":"soluciones","prevencion":"prevención",
    "proteccion":"protección","fiscalizacion":"fiscalización",
    "sancion":"sanción","sanciones":"sanciones",
    "investigacion":"investigación","investigaciones":"investigaciones",
    "accion":"acción","acciones":"acciones",
    "direccion":"dirección","decision":"decisión",
    "decisiones":"decisiones","eleccion":"elección",
    "elecciones":"elecciones","votacion":"votación",
    "aprobacion":"aprobación","legislacion":"legislación",
    "reclamacion":"reclamación","reclamaciones":"reclamaciones",
    "obligacion":"obligación","obligaciones":"obligaciones",
    "inflacion":"inflación","tributacion":"tributación",
    "financiera":"financiera","financiero":"financiero",
    "economica":"económica","economico":"económico",
    "economia":"economía","credito":"crédito",
    "creditos":"créditos","prestamo":"préstamo",
    "prestamos":"préstamos","interes":"interés",
    "comision":"comisión","comisiones":"comisiones",
    "politica":"política","politicas":"políticas",
    "politico":"político","publica":"pública",
    "publico":"público","estrategia":"estrategia",
    "estrategica":"estratégica","estrategico":"estratégico",
    "logistica":"logística","analisis":"análisis",
    "diagnostico":"diagnóstico","indice":"índice",
    "vehiculo":"vehículo","vehiculos":"vehículos",
    "electrico":"eléctrico","electrica":"eléctrica",
    "energia":"energía","energetica":"energética",
    "petroleo":"petróleo","mineria":"minería",
    "agricola":"agrícola","biologica":"biológica",
    "ecologica":"ecológica","inclusion":"inclusión",
    "exclusion":"exclusión","pension":"pensión",
    "pensiones":"pensiones","jubilacion":"jubilación",
    "compensacion":"compensación","remuneracion":"remuneración",
    "contratacion":"contratación","capacitacion":"capacitación",
    "formacion":"formación","certificacion":"certificación",
    "habilitacion":"habilitación","autorizacion":"autorización",
    "concesion":"concesión","licitacion":"licitación",
    "migracion":"migración","poblacion":"población",
    "recaudacion":"recaudación","asignacion":"asignación",
    "corporacion":"corporación","fundacion":"fundación",
    "institucion":"institución","instituciones":"instituciones",
    "region":"región","unico":"único","unica":"única",
    "ultimo":"último","ultima":"última","proximo":"próximo",
    "basico":"básico","basica":"básica","historico":"histórico",
    "historica":"histórica","medico":"médico","medica":"médica",
    "farmaceutica":"farmacéutica","clinica":"clínica",
    "numero":"número","telefono":"teléfono","telefonia":"telefonía",
    "movil":"móvil","moviles":"móviles","codigo":"código",
    "informatica":"informática","electronica":"electrónica",
    "robotica":"robótica","ciberseguridad":"ciberseguridad",
    "trafico":"tráfico","transito":"tránsito","aereo":"aéreo",
    "maritimo":"marítimo","turistica":"turística",
    "turistico":"turístico","gastronomia":"gastronomía",
    "academica":"académica","academico":"académico",
    "pedagogica":"pedagógica","cientifica":"científica",
    "cientifico":"científico","juridica":"jurídica",
    "juridico":"jurídico","constitucion":"constitución",
    "resolucion":"resolución","notificacion":"notificación",
    "programacion":"programación","actualizacion":"actualización",
    "verificacion":"verificación","validacion":"validación",
    "liquidacion":"liquidación","facturacion":"facturación",
    "evasion":"evasión","corrupcion":"corrupción",
    "deforestacion":"deforestación","contaminacion":"contaminación",
    "conservacion":"conservación","restauracion":"restauración",
    "rehabilitacion":"rehabilitación","renovacion":"renovación",
    "ampliacion":"ampliación","inauguracion":"inauguración",
    "celebracion":"celebración","clasificacion":"clasificación",
    "eliminacion":"eliminación","motivacion":"motivación",
    "satisfaccion":"satisfacción","reputacion":"reputación",
    "disposicion":"disposición",
}

_ENIE_MAP = {
    "desempeno":"desempeño","desempenos":"desempeños",
    "empeno":"empeño","empenos":"empeños",
    "ensenanza":"enseñanza","ensenanzas":"enseñanzas",
    "diseno":"diseño","disenos":"diseños",
    "disenador":"diseñador","disenadora":"diseñadora","disenadores":"diseñadores",
    "nino":"niño","nina":"niña","ninos":"niños","ninas":"niñas","ninez":"niñez",
    "ano":"año","anos":"años",
    "danio":"daño","danios":"daños","dano":"daño","danos":"daños",
    "danino":"dañino","danina":"dañina",
    "montana":"montaña","montanas":"montañas",
    "espana":"España","espanol":"español","espanola":"española","espanoles":"españoles",
    "companero":"compañero","companera":"compañera","companeros":"compañeros","companeras":"compañeras",
    "compania":"compañía","companias":"compañías","acompanamiento":"acompañamiento",
    "cana":"caña","canas":"cañas",
    "banio":"baño","banios":"baños","bano":"baño","banos":"baños",
    "pena":"peña","penas":"peñas","penon":"peñón",
    "senor":"señor","senora":"señora","senores":"señores","senoras":"señoras",
    "senal":"señal","senales":"señales","senalizacion":"señalización",
    "pequeno":"pequeño","pequena":"pequeña","pequenos":"pequeños","pequenas":"pequeñas",
    "sueno":"sueño","suenos":"sueños",
    "dueno":"dueño","duena":"dueña","duenos":"dueños","duenas":"dueñas",
    "otono":"otoño","punio":"puño","punios":"puños","puno":"puño",
    "canon":"cañón","canones":"cañones",
    "manana":"mañana","mananas":"mañanas",
    "cabana":"cabaña","cabanas":"cabañas","banera":"bañera",
    "vinedo":"viñedo","vinedos":"viñedos",
    "rebano":"rebaño","rebanos":"rebaños",
    "extrano":"extraño","extrana":"extraña","extranos":"extraños","extranas":"extrañas",
    "enganio":"engaño","engano":"engaño","enganos":"engaños",
    "tamanio":"tamaño","tamano":"tamaño","tamanos":"tamaños",
    "muneca":"muñeca","munecas":"muñecas",
    "cunado":"cuñado","cunada":"cuñada","cunados":"cuñados",
    "albanil":"albañil","albaniles":"albañiles",
    "narino":"Nariño","quindio":"Quindío",
    "ibanez":"Ibáñez","nunez":"Núñez","munoz":"Muñoz",
    "ordonez":"Ordóñez","yanez":"Yáñez","castaneda":"Castañeda","penalosa":"Peñalosa",
    "vineta":"viñeta","vinetas":"viñetas",
    "banado":"bañado","banada":"bañada",
    "rinon":"riñón","rinones":"riñones",
    "panial":"pañal","paniales":"pañales","panal":"pañal","panales":"pañales",
    "arana":"araña","aranas":"arañas",
    "pestana":"pestaña","pestanas":"pestañas",
    "guino":"guiño","guinos":"guiños",
    "munequera":"muñequera","lenador":"leñador","lenadores":"leñadores",
    "resena":"reseña","resenas":"reseñas",
    "panuelo":"pañuelo","panuelos":"pañuelos",
    "companerismo":"compañerismo",
    "desengano":"desengaño","lenio":"leño","leno":"leño",
}

def corregir_tildes(texto: str) -> str:
    if not texto: return texto
    palabras = texto.split()
    resultado = []
    for p in palabras:
        low = p.lower()
        if low in _TILDE_MAP:
            c = _TILDE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        elif low in _ENIE_MAP:
            c = _ENIE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        else:
            resultado.append(p)
    return " ".join(resultado)


# ======================================
# CSS
# ======================================
def load_custom_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Google+Sans:wght@400;500;700&family=Google+Sans+Text:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
:root {
    --bg:#f8f9fa;--s1:#ffffff;--s2:#f1f3f4;--s3:#e8eaed;
    --border:#dadce0;--border2:#bdc1c6;--border-focus:#f97316;
    --text:#202124;--text2:#3c4043;--text3:#5f6368;--text4:#9aa0a6;
    --accent:#f97316;--accent2:#ea580c;--accent3:#c2410c;
    --accent-bg:#fff7ed;--accent-bg2:#ffedd5;--accent-bdr:#fed7aa;
    --green:#059669;--green2:#047857;--green-bg:#ecfdf5;--green-bdr:#a7f3d0;
    --red:#dc2626;--amber:#d97706;--blue:#1a73e8;
    --r:8px;--r2:12px;--r3:16px;--r4:20px;
    --shadow-sm:0 1px 2px rgba(60,64,67,0.1),0 1px 3px rgba(60,64,67,0.08);
    --shadow-md:0 1px 3px rgba(60,64,67,0.12),0 4px 8px rgba(60,64,67,0.08);
    --shadow-lg:0 2px 6px rgba(60,64,67,0.1),0 8px 24px rgba(60,64,67,0.1);
    --transition:all 0.2s cubic-bezier(0.4,0,0.2,1);
}
html,body,[data-testid="stApp"]{
    background:var(--bg)!important;color:var(--text)!important;
    font-family:'Google Sans Text','Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
    font-size:14px;-webkit-font-smoothing:antialiased;letter-spacing:0.01em;
}
#MainMenu,footer,header{visibility:hidden}.stDeployButton{display:none}
.block-container{padding-top:1rem!important;padding-bottom:0!important}
[data-testid="stAppViewBlockContainer"]{padding-top:1rem!important}
.app-header{background:var(--s1);border:1px solid var(--border);border-radius:var(--r3);padding:1rem 1.5rem;margin-bottom:1rem;display:flex;align-items:center;gap:1rem;box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.app-header::after{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,#f97316,#fb923c,#fdba74);}
.app-header-icon{width:40px;height:40px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;color:white;flex-shrink:0;box-shadow:0 2px 8px rgba(249,115,22,0.3);}
.app-header-text{flex:1}
.app-header-title{font-family:'Google Sans',sans-serif;font-size:1.25rem;font-weight:700;color:var(--text);letter-spacing:-0.01em;line-height:1.3}
.app-header-version{font-family:'Roboto Mono',monospace;font-size:0.65rem;color:var(--text3);letter-spacing:0.03em;margin-top:0.15rem}
.app-header-badge{background:var(--accent-bg);border:1px solid var(--accent-bdr);color:var(--accent2);font-family:'Roboto Mono',monospace;font-size:0.6rem;font-weight:500;padding:0.25rem 0.75rem;border-radius:100px;letter-spacing:0.04em;text-transform:uppercase;white-space:nowrap;}
[data-testid="stTabs"] [data-testid="stTabsList"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;padding:4px!important;gap:4px!important;box-shadow:var(--shadow-sm)!important;margin-bottom:0.75rem!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]{font-family:'Google Sans',sans-serif!important;font-size:0.88rem!important;font-weight:500!important;color:var(--text2)!important;border-radius:var(--r)!important;padding:0.45rem 1.2rem!important;border:none!important;background:transparent!important;transition:var(--transition)!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]:hover{background:var(--s2)!important;color:var(--text)!important}
[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"]{background:var(--accent-bg)!important;color:var(--accent2)!important;border:1px solid var(--accent-bdr)!important;font-weight:700!important;}
.metrics-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:0.6rem;margin:0.8rem 0}
.metric-card{background:var(--s1);border:1px solid var(--border);border-radius:var(--r2);padding:0.8rem 0.6rem;text-align:center;transition:var(--transition);box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.metric-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--r2) var(--r2) 0 0}
.metric-card.m-total::before{background:linear-gradient(90deg,#5f6368,#9aa0a6)}
.metric-card.m-unique::before{background:linear-gradient(90deg,#059669,#34d399)}
.metric-card.m-dup::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.metric-card.m-time::before{background:linear-gradient(90deg,#1a73e8,#4285f4)}
.metric-card.m-cost::before{background:linear-gradient(90deg,#f97316,#fb923c)}
.metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-lg)}
.metric-val{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;line-height:1;margin-bottom:0.3rem;letter-spacing:-0.01em}
.metric-lbl{font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);text-transform:uppercase;letter-spacing:0.08em;font-weight:500}
[data-testid="stForm"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r3)!important;padding:1.2rem 1.5rem!important;box-shadow:var(--shadow-md)!important;}
.sec-label{font-family:'Google Sans',sans-serif;font-size:0.72rem;font-weight:700;color:var(--text2);letter-spacing:0.08em;text-transform:uppercase;padding-bottom:0.3rem;border-bottom:2px solid var(--s3);margin:0.8rem 0 0.5rem;display:flex;align-items:center;gap:0.5rem;}
.sec-label::before{content:'';display:inline-block;width:3px;height:12px;background:linear-gradient(180deg,#f97316,#ea580c);border-radius:2px}
.upload-zone{display:grid;grid-template-columns:repeat(3,1fr);gap:0.6rem;margin:0.3rem 0}
.upload-zone-card{background:var(--s1);border:1.5px dashed var(--border);border-radius:var(--r2);padding:0.6rem 0.8rem;display:flex;align-items:center;gap:0.6rem;transition:var(--transition);}
.upload-zone-card:hover{border-color:var(--accent);border-style:solid;transform:translateY(-1px);box-shadow:var(--shadow-md)}
.upload-zone-icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1rem;flex-shrink:0;}
.upload-zone-icon.uz-dossier{background:#fff7ed;color:#f97316}
.upload-zone-icon.uz-region{background:#ecfdf5;color:#059669}
.upload-zone-icon.uz-internet{background:#eff6ff;color:#1a73e8}
.upload-zone-text{flex:1;min-width:0}
.upload-zone-title{font-family:'Google Sans',sans-serif;font-size:0.82rem;font-weight:700;color:var(--text);line-height:1.2}
.upload-zone-desc{font-size:0.7rem;color:var(--text3);line-height:1.3}
[data-testid="stFileUploader"]{background:var(--s1)!important;border:1.5px dashed var(--border)!important;border-radius:var(--r)!important;padding:0.4rem 0.6rem!important;transition:var(--transition)!important;min-height:auto!important;}
[data-testid="stFileUploader"]:hover{border-color:var(--accent)!important;border-style:solid!important;background:var(--accent-bg)!important;}
[data-testid="stFileUploader"] section{padding:0.2rem!important}
[data-testid="stFileUploader"] section>div{font-size:0.78rem!important;color:var(--text2)!important}
[data-testid="stFileUploader"] section small{font-size:0.7rem!important;color:var(--text3)!important}
[data-testid="stFileUploader"] button{background:var(--accent-bg)!important;border:1px solid var(--accent-bdr)!important;color:var(--accent2)!important;font-weight:500!important;font-size:0.75rem!important;border-radius:100px!important;padding:0.25rem 0.8rem!important;font-family:'Google Sans',sans-serif!important;transition:var(--transition)!important;}
[data-testid="stFileUploader"] button:hover{background:var(--accent)!important;color:white!important;border-color:var(--accent)!important}
[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:var(--r)!important;font-family:'Google Sans Text',sans-serif!important;font-size:0.9rem!important;padding:0.5rem 0.75rem!important;transition:var(--transition)!important;}
[data-testid="stTextInput"] input:focus,[data-testid="stTextArea"] textarea:focus{border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(249,115,22,0.12)!important;}
[data-testid="stTextInput"] input::placeholder,[data-testid="stTextArea"] textarea::placeholder{color:var(--text4)!important;font-size:0.85rem!important;}
label[data-testid="stWidgetLabel"] p{font-family:'Google Sans',sans-serif!important;color:var(--text2)!important;font-size:0.82rem!important;font-weight:500!important;margin-bottom:0.15rem!important;}
.stButton>button,[data-testid="stDownloadButton"]>button{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:100px!important;font-family:'Google Sans',sans-serif!important;font-weight:500!important;font-size:0.88rem!important;transition:var(--transition)!important;padding:0.5rem 1.2rem!important;box-shadow:none!important;}
.stButton>button:hover,[data-testid="stDownloadButton"]>button:hover{border-color:var(--accent)!important;color:var(--accent2)!important;background:var(--accent-bg)!important;box-shadow:var(--shadow-sm)!important;transform:translateY(-1px)!important;}
.stButton>button[kind="primary"],[data-testid="stDownloadButton"]>button[kind="primary"]{background:var(--accent)!important;border:none!important;color:#fff!important;font-weight:500!important;font-size:0.92rem!important;padding:0.6rem 1.5rem!important;box-shadow:0 1px 3px rgba(249,115,22,0.3),0 4px 12px rgba(249,115,22,0.15)!important;letter-spacing:0.01em!important;}
.stButton>button[kind="primary"]:hover,[data-testid="stDownloadButton"]>button[kind="primary"]:hover{background:var(--accent2)!important;box-shadow:0 2px 6px rgba(234,88,12,0.35),0 8px 24px rgba(234,88,12,0.18)!important;transform:translateY(-1px)!important;color:#fff!important;}
[data-testid="stRadio"] label{font-family:'Google Sans Text',sans-serif!important;color:var(--text)!important;font-size:0.88rem!important;font-weight:400!important;}
[data-testid="stRadio"]{margin-bottom:0!important}
[data-testid="stRadio"]>div{gap:0!important}
[data-testid="stStatus"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;font-family:'Roboto Mono',monospace!important;font-size:0.8rem!important;}
[data-testid="stAlert"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;color:var(--text2)!important;font-size:0.85rem!important;padding:0.6rem 0.8rem!important;}
.success-banner{background:linear-gradient(135deg,#ecfdf5,#d1fae5);border:1px solid var(--green-bdr);border-left:4px solid var(--green);border-radius:var(--r2);padding:0.8rem 1.2rem;margin:0.5rem 0 0.8rem;display:flex;align-items:center;gap:0.8rem;}
.success-icon{width:34px;height:34px;background:linear-gradient(135deg,#059669,#047857);border-radius:50%;display:flex;align-items:center;justify-content:center;color:white;font-size:1rem;flex-shrink:0;}
.success-title{font-family:'Google Sans',sans-serif;font-size:1rem;font-weight:700;color:#047857;margin-bottom:0.1rem}
.success-sub{font-size:0.8rem;color:var(--text2)}
.auth-wrap{max-width:380px;margin:8vh auto 0;text-align:center}
.auth-icon{width:60px;height:60px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:16px;display:inline-flex;align-items:center;justify-content:center;font-size:1.6rem;color:white;margin-bottom:1rem;box-shadow:0 4px 16px rgba(249,115,22,0.3);}
.auth-title{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;color:var(--text);margin-bottom:0.3rem}
.auth-sub{font-size:0.85rem;color:var(--text3);margin-bottom:2rem}
.cluster-info{background:var(--accent-bg);border:1px solid var(--accent-bdr);border-radius:var(--r);padding:0.5rem 0.8rem;margin:0.4rem 0;font-family:'Roboto Mono',monospace;font-size:0.68rem;color:var(--text2);line-height:1.6;}
.cluster-info b{color:var(--accent2);font-size:0.72rem}
[data-testid="stProgressBar"]>div>div{background:linear-gradient(90deg,#f97316,#fb923c,#fdba74)!important;border-radius:100px!important;height:5px!important;}
[data-testid="stDataFrame"]{border:1px solid var(--border)!important;border-radius:var(--r2)!important;box-shadow:var(--shadow-sm)!important;overflow:hidden!important;}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:var(--s2);border-radius:3px}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--accent)}
.footer{font-family:'Roboto Mono',monospace;font-size:0.6rem;color:var(--text4);text-align:center;padding:0.8rem 0 0.5rem;letter-spacing:0.04em;border-top:1px solid var(--s3);margin-top:1rem;}
.stElementContainer{margin-bottom:0!important}
[data-testid="stVerticalBlock"]>div{gap:0.3rem!important}
[data-testid="stHorizontalBlock"]>div{gap:0.4rem!important}
hr{border-color:var(--s3)!important;margin:0.5rem 0!important}
[data-testid="stSelectbox"]>div>div{font-family:'Google Sans Text',sans-serif!important;font-size:0.88rem!important;color:var(--text)!important;}
@media(max-width:768px){
    .metrics-grid{grid-template-columns:repeat(2,1fr)}
    .upload-zone{grid-template-columns:1fr}
    .app-header{flex-direction:column;text-align:center;gap:0.5rem;padding:1rem}
}
</style>
""", unsafe_allow_html=True)


# ======================================
# Umbrales adaptativos según tamaño del corpus
# ======================================
def _umbrales_adaptativos(n: int) -> dict:
    """
    Devuelve umbrales ajustados al tamaño del corpus.

    Con pocas noticias (2-10) el espacio de similitudes es pequeño y los
    algoritmos de clustering tienden a colapsar noticias distintas.  Se
    elevan los umbrales para exigir similitud real antes de agrupar.

    Rangos:
      n ≤  5  → modo "individual": cada noticia es su propio subtema
                 salvo que la similitud sea ≥ 0.92 (casi idénticas)
      n ≤ 10  → umbrales muy estrictos
      n ≤ 20  → umbrales estrictos
      n > 20  → umbrales base originales
    """
    if n <= 5:
        return dict(
            subtema=0.93,          # sólo noticias casi idénticas se agrupan
            tema=0.85,
            dedup_label=0.90,
            fusion_subtemas=0.92,
            fusion_intergrupo=0.95,
            min_pertenencia_subtema=0.80,
            min_pertenencia_tema=0.75,
            coherencia_etiqueta=0.50,
            sim_minima_agrupacion=0.93,  # verificación post-clustering
            sim_minima_keywords=0.93,    # _paso2b desactivado en la práctica
            max_iter_fusion=1,
            num_temas_max=n,             # nunca más temas que noticias
            usar_paso2b=False,           # deshabilitar keywords raras
            usar_fusion_iterativa=False, # deshabilitar fusión iterativa
        )
    elif n <= 10:
        return dict(
            subtema=0.88,
            tema=0.80,
            dedup_label=0.85,
            fusion_subtemas=0.87,
            fusion_intergrupo=0.91,
            min_pertenencia_subtema=0.72,
            min_pertenencia_tema=0.65,
            coherencia_etiqueta=0.42,
            sim_minima_agrupacion=0.88,
            sim_minima_keywords=0.88,
            max_iter_fusion=2,
            num_temas_max=min(n, 5),
            usar_paso2b=False,
            usar_fusion_iterativa=False,
        )
    elif n <= 20:
        return dict(
            subtema=0.83,
            tema=0.76,
            dedup_label=0.82,
            fusion_subtemas=0.82,
            fusion_intergrupo=0.88,
            min_pertenencia_subtema=0.66,
            min_pertenencia_tema=0.58,
            coherencia_etiqueta=0.38,
            sim_minima_agrupacion=0.84,
            sim_minima_keywords=0.84,
            max_iter_fusion=3,
            num_temas_max=min(n // 2, NUM_TEMAS_MAX),
            usar_paso2b=True,
            usar_fusion_iterativa=True,
        )
    else:
        return dict(
            subtema=UMBRAL_SUBTEMA,
            tema=UMBRAL_TEMA,
            dedup_label=UMBRAL_DEDUP_LABEL,
            fusion_subtemas=UMBRAL_FUSION_SUBTEMAS,
            fusion_intergrupo=UMBRAL_FUSION_INTERGRUPO,
            min_pertenencia_subtema=UMBRAL_MIN_PERTENENCIA_SUBTEMA,
            min_pertenencia_tema=UMBRAL_MIN_PERTENENCIA_TEMA,
            coherencia_etiqueta=UMBRAL_COHERENCIA_ETIQUETA,
            sim_minima_agrupacion=SIM_MINIMA_AGRUPACION_SUBTEMA,
            sim_minima_keywords=SIM_MINIMA_KEYWORDS_RARAS,
            max_iter_fusion=MAX_ITER_FUSION,
            num_temas_max=NUM_TEMAS_MAX,
            usar_paso2b=True,
            usar_fusion_iterativa=True,
        )


# ======================================
# Web Scraping de GlobalNews
# ======================================
_SCRAPE_CACHE_PATH = "/root/.hermes/scraping_cache.json"
_LLM_RESUMEN_CACHE_PATH = "/root/.hermes/resumen_tono_cache.json"

def _load_cache(path):
    if os.path.exists(path):
        try:
            with open(path) as f: return json.load(f)
        except: pass
    return {}

def _save_cache(path, data):
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w") as f: json.dump(data, f, ensure_ascii=False)
    except: pass

def extract_urls_from_xlsx(xlsx_bytes):
    """Mapea ref(W2, W3...) → URL Validar.aspx desde un xlsx."""
    zf = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    try:
        rels = zf.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8")
        root = ET.fromstring(rels)
        rid2url = {rel.get("Id"): html.unescape(rel.get("Target",""))
                    for rel in root if "Validar.aspx" in rel.get("Target","")}
        sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
        sroot = ET.fromstring(sheet)
        ns = {"s":"http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        rns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
        hl = sroot.find(".//s:hyperlinks", ns)
        return {ref.get("ref"): rid2url[ref.get(rns+"id")]
                for ref in hl if ref.get(rns+"id") in rid2url} if hl is not None else {}
    except:
        return {}

def url_to_direct(url):
    """Validar.aspx → news2 directo para scraping."""
    nm = re.search(r'[?&]n=(\d+)', url)
    um = re.search(r'[?&]u=([a-f0-9-]+)', url, re.IGNORECASE)
    cm = re.search(r'[?&]c=(\d+)', url)
    if not nm or not um: return None
    return "http://news2.globalnews.com.co/?accessNewsCode={}|{}|{}&mode=image".format(
        um.group(1), nm.group(1), cm.group(1) if cm else "1")

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-CO,es;q=0.9,en;q=0.8",
}

def _extract_text_from_html(html: str) -> str:
    """Extrae texto legible de un HTML de noticia GlobalNews."""
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        # Remover scripts, styles, nav, footer
        for tag in soup(["script", "style", "nav", "header", "footer", "aside"]):
            tag.decompose()
        body_text = soup.get_text(separator="\n")
        # Limpiar bloques de texto
        lines = [ln.strip() for ln in body_text.split("\n") if ln.strip()]
        text = "\n".join(lines)
        # Intentar extraer entre marcadores conocidos
        for start_marker in ["Imagen Resumen", "Imagen resumen", "Contenido"]:
            i1 = text.find(start_marker)
            if i1 >= 0:
                text = text[i1 + len(start_marker):]
                break
        for end_marker in ["Nube de palabras", "ANFENAVI", "Compartir en"]:
            i2 = text.find(end_marker)
            if i2 >= 0:
                text = text[:i2]
                break
        return re.sub(r'\s+', ' ', text).strip()
    except Exception:
        # Fallback: regex básico si no hay bs4
        tags_re = re.compile(r'<[^>]+>')
        text = tags_re.sub(" ", html)
        text = re.sub(r'\s+', ' ', text).strip()
        return text if len(text) > 50 else ""

def scrape_single(direct_url: str):
    """Scrapea una noticia con requests (sin Selenium)."""
    try:
        import requests
        resp = requests.get(direct_url, headers=_HEADERS, timeout=15)
        resp.raise_for_status()
        text = _extract_text_from_html(resp.text)
        return text if len(text) > 100 else None
    except Exception:
        return None

def scrape_all_news(urls_data, cache, pbar, pstatus):
    """Scrape multiples noticias con requests concurrentes."""
    results = {}
    total = len(urls_data)

    # Resolve cached and to-fetch
    to_fetch = []
    for rnum, url, nid in urls_data:
        if nid in cache:
            results[rnum] = cache[nid]
            continue
        du = url_to_direct(url)
        if not du:
            results[rnum] = None
            continue
        to_fetch.append((rnum, du, nid))

    if not to_fetch:
        _save_cache(_SCRAPE_CACHE_PATH, cache)
        return results

    # Fetch concurrently
    from concurrent.futures import ThreadPoolExecutor, as_completed
    completed = 0

    def _fetch_task(item):
        rnum, du, nid = item
        text = scrape_single(du)
        return rnum, text, nid

    with ThreadPoolExecutor(max_workers=10) as pool:
        future_map = {pool.submit(_fetch_task, item): item for item in to_fetch}
        for fut in as_completed(future_map):
            rnum, text, nid = fut.result()
            if text:
                results[rnum] = text
                cache[nid] = text
            else:
                results[rnum] = None
            completed += 1
            pct = (len(results) + completed) / max(total, 1)
            pstatus.text("Scraping {}/{}...".format(completed, len(to_fetch)))
            pbar.progress(pct)

    _save_cache(_SCRAPE_CACHE_PATH, cache)
    return results

# ======================================
# Caché Global de Embeddings
# ======================================
class EmbeddingCache:
    def __init__(self):
        self._cache: Dict[str, List[float]] = {}
        self._hits = 0
        self._misses = 0

    def _key(self, text):
        return hashlib.md5(text[:2000].encode('utf-8', errors='ignore')).hexdigest()

    def get(self, text):
        k = self._key(text)
        if k in self._cache:
            self._hits += 1
            return self._cache[k]
        self._misses += 1
        return None

    def put(self, text, emb):
        self._cache[self._key(text)] = emb

    def get_many(self, textos):
        results = [None] * len(textos)
        missing = []
        for i, t in enumerate(textos):
            c = self.get(t)
            if c is not None:
                results[i] = c
            else:
                missing.append(i)
        return results, missing

    def stats(self):
        total = self._hits + self._misses
        rate = (self._hits / total * 100) if total > 0 else 0
        return f"Cache: {self._hits} hits, {self._misses} misses ({rate:.0f}%)"

    def clear(self):
        self._cache.clear()
        self._hits = 0
        self._misses = 0


if '_emb_cache' not in st.session_state:
    st.session_state['_emb_cache'] = EmbeddingCache()


def get_embedding_cache():
    return st.session_state['_emb_cache']


# ======================================
# Utilidades
# ======================================

@st.cache_data(ttl=3600)
def _cargar_mapa_excel(url: str) -> pd.DataFrame:
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        return pd.read_excel(io.BytesIO(resp.content))
    except Exception as e:
        st.error(f"Error al cargar mapa desde GitHub: {e}")
        st.stop()


def check_password():
    if st.session_state.get("password_correct", False):
        return True
    st.markdown("""
    <div class="auth-wrap">
        <div class="auth-icon">◈</div>
        <div class="auth-title">Sistema de Análisis</div>
        <div class="auth-sub">Ingresa tus credenciales para continuar</div>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("pw"):
            pw = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            if st.form_submit_button("Ingresar", use_container_width=True, type="primary"):
                if pw == st.secrets.get("APP_PASSWORD", "INVALID"):
                    st.session_state["password_correct"] = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
    return False


def call_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return fn(*a, **kw)
        except Exception as e:
            if att == 2:
                raise e
            time.sleep(d)
            d *= 2


async def acall_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return await fn(*a, **kw)
        except Exception as e:
            if att == 2:
                raise e
            await asyncio.sleep(d)
            d *= 2


def norm_key(text):
    if text is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))


def capitalizar_etiqueta(tema):
    if not tema or not tema.strip():
        return "Sin tema"
    tema = tema.strip().lower()
    tema = corregir_tildes(tema)
    return tema[0].upper() + tema[1:]


def _frase_esta_completa(texto):
    if not texto or not texto.strip():
        return False
    palabras = texto.strip().split()
    if not palabras:
        return False
    ultima = palabras[-1].lower().rstrip(".,;:!?")
    return unidecode(ultima) not in _TRAILING_INCOMPLETE and len(ultima) > 1


def _recortar_frase_completa(texto, max_palabras=7):
    if not texto:
        return "Sin tema"
    palabras = texto.strip().split()
    if len(palabras) > max_palabras:
        palabras = palabras[:max_palabras]
    while palabras and unidecode(palabras[-1].lower().rstrip(".,;:!?")) in _TRAILING_INCOMPLETE:
        palabras.pop()
    if not palabras:
        return texto.strip().split()[0] if texto.strip() else "Sin tema"
    return " ".join(palabras)


def limpiar_tema(tema):
    if not tema:
        return "Sin tema"
    tema = tema.strip().strip('"\'')
    for px in ["subtema:", "tema:", "categoría:", "categoria:", "category:"]:
        if tema.lower().startswith(px):
            tema = tema[len(px):].strip()
    tema = _recortar_frase_completa(tema, max_palabras=7)
    return capitalizar_etiqueta(tema) if tema else "Sin tema"


def limpiar_tema_geografico(tema, marca, aliases):
    if not tema:
        return "Sin tema"

    tl = unidecode(tema.lower())

    for n in [marca] + [a for a in aliases if a]:
        patron = r'\b' + re.escape(unidecode(n.strip().lower())) + r'\b'
        tl = re.sub(patron, '', tl)

    frases_eliminar = [
        "en colombia", "de colombia", "del pais", "en el pais",
        "territorio nacional", "a nivel nacional", "en todo el pais",
    ]
    for frase in frases_eliminar:
        tl = re.sub(r'\b' + re.escape(frase) + r'\b', '', tl)

    tl = re.sub(r'\s+', ' ', tl).strip()

    if not tl:
        return "Sin tema"

    tokens_orig = tema.split()
    tokens_norm = unidecode(tema.lower()).split()
    norm_disponibles = tl.split()

    resultado_tokens = []
    for orig, norm in zip(tokens_orig, tokens_norm):
        if norm_disponibles and norm == norm_disponibles[0]:
            resultado_tokens.append(orig)
            norm_disponibles.pop(0)

    resultado = " ".join(resultado_tokens).strip()
    resultado = corregir_tildes(resultado) if resultado else ""
    return limpiar_tema(resultado) if resultado.strip() else "Sin tema"


def string_norm_label(s):
    if not s:
        return ""
    s = unidecode(s.lower())
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return " ".join(t for t in s.split() if t not in STOPWORDS_ES)


def _validar_estructura_subtema(etiqueta: str) -> bool:
    if not etiqueta or len(etiqueta.split()) < 2:
        return False
    if len(etiqueta.split()) > 7:
        return False
    if _PATRON_TITULAR.match(etiqueta):
        return False
    if _PATRON_ESTADO.search(etiqueta):
        return False

    palabras = etiqueta.split()
    if len(palabras) <= 4:
        nexos = {
            "de","del","para","sobre","en","con","por","ante","hacia",
            "entre","sin","al","las","los","una","uno","que","como",
            "y","o","a","e","u",
        }
        tiene_nexo = any(unidecode(p.lower().rstrip(".,;:!?")) in nexos for p in palabras[1:])
        if not tiene_nexo:
            return False

    return True


def extract_link(cell):
    if hasattr(cell, "hyperlink") and cell.hyperlink:
        return {"value": "Link", "url": cell.hyperlink.target}
    if isinstance(cell.value, str) and "=HYPERLINK" in cell.value:
        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value)
        if m:
            return {"value": "Link", "url": m.group(1)}
    return {"value": cell.value, "url": None}


def normalize_title_for_comparison(title):
    if not isinstance(title, str):
        return ""
    tmp = re.split(r"\s*[:|-]\s*", title, 1)
    return re.sub(r"\W+", " ", tmp[0]).lower().strip()


def clean_title_for_output(title):
    return re.sub(r"\s*\|\s*[\w\s]+$", "", str(title)).strip()


def corregir_texto(text):
    if not isinstance(text, str):
        return text
    text = re.sub(r"(<br>|\[\.\.\.\]|\s+)", " ", text).strip()
    m = re.search(r"[A-ZÁÉÍÓÚÑ]", text)
    if m:
        text = text[m.start():]
    if text and not text.endswith("..."):
        text = text.rstrip(".") + "..."
    return text


def normalizar_tipo_medio(tipo_raw):
    if not isinstance(tipo_raw, str):
        return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    return {
        "fm": "Radio", "am": "Radio", "radio": "Radio",
        "aire": "Televisión", "cable": "Televisión", "tv": "Televisión",
        "television": "Televisión", "televisión": "Televisión",
        "senal abierta": "Televisión", "señal abierta": "Televisión",
        "diario": "Prensa", "prensa": "Prensa",
        "revista": "Revistas", "revistas": "Revistas",
        "online": "Internet", "internet": "Internet",
        "digital": "Internet", "web": "Internet"
    }.get(t, str(tipo_raw).strip().title() or "Otro")


def generar_resumen_cliente(texto, titulo, medio, fecha, cliente, voceros=""):
    """Genera un resumen enfocado al cliente sobre el texto scrapeado de la noticia."""
    prompt_sistema = (
        u"Eres analista de medios para {cliente}. "
        u"Analiza la noticia y genera UN resumen enfocado al cliente como audiencia.\n\n"
        u"Responde SOLO con JSON: "
        u'{{"resumen_cliente": "2-3 parrafos con datos, voceros, cifras e impacto para {cliente} y el sector avicola. '
        u'Si la noticia menciona a {cliente} o voceros, destaca su perspectiva. '
        u'Si no, analiza el impacto POTENCIAL para el sector."}}'
    ).format(cliente=cliente)

    prompt_usuario = (
        u"CLIENTE: {cliente}\nVOCEROS: {voceros}\nTITULAR: {titulo}\nMEDIO: {medio}\nFECHA: {fecha}\n\n"
        u"TEXTO:\n{texto}"
    ).format(cliente=cliente, voceros=voceros or "No especificados",
             titulo=titulo or "", medio=medio or "", fecha=fecha or "", texto=texto[:4500])

    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role":"system","content":prompt_sistema},{"role":"user","content":prompt_usuario}],
            temperature=0.2, max_tokens=1000,
            response_format={"type":"json_object"}
        )
        u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', None) or {}
        if u:
            st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
            st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
        return json.loads(resp.choices[0].message.content).get("resumen_cliente", "")
    except:
        return ""


def texto_para_embedding(titulo, resumen, max_len=1800):
    t = str(titulo or "").strip()
    r = str(resumen or "").strip()
    return f"{t}. {t}. {t}. {r}"[:max_len]


def _validar_etiqueta_completa(etiqueta, titulos_grp=None, resumenes_grp=None, marca="", aliases=None, fallback_fn=None):
    if not etiqueta or etiqueta.strip().lower() in ("sin tema", "varios", "n/a"):
        if fallback_fn:
            return fallback_fn(titulos_grp or [])
        return "Cobertura informativa general"
    if _frase_esta_completa(etiqueta):
        return etiqueta
    recortada = _recortar_frase_completa(etiqueta, max_palabras=7)
    if _frase_esta_completa(recortada) and len(recortada.split()) >= 2:
        return capitalizar_etiqueta(recortada)
    if titulos_grp and len(titulos_grp) > 0:
        try:
            prompt = (
                f"La frase '{etiqueta}' está incompleta o es genérica. "
                f"Genera una frase temática COMPLETA en español de 4-6 palabras "
                f"con preposición (de/del/para/sobre/en):\n\n"
                + "\n".join(f"  · {t[:120]}" for t in titulos_grp[:4])
                + "\n\nREGLAS: frase nominal con preposición, terminar en sustantivo/adjetivo, "
                "tildes y ñ correctas, sin marcas ni ciudades.\n"
                "CORRECTO: 'Proyecto de terminal de transportes', 'Operación del Canal del Dique'\n"
                "INCORRECTO: 'Terminal transportes', 'Operación canal'\n"
                'JSON: {"subtema":"..."}'
            )
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=80,
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
            raw = json.loads(resp.choices[0].message.content).get("subtema", "")
            if raw:
                cleaned = limpiar_tema_geografico(limpiar_tema(raw), marca, aliases or [])
                if _frase_esta_completa(cleaned) and len(cleaned.split()) >= 2:
                    return capitalizar_etiqueta(cleaned)
        except:
            pass
    if fallback_fn:
        return fallback_fn(titulos_grp or [])
    return capitalizar_etiqueta(recortada) if recortada and len(recortada.split()) >= 2 else "Cobertura informativa general"


def dedup_labels(etiquetas, umbral=UMBRAL_DEDUP_LABEL):
    unique = list(dict.fromkeys(etiquetas))
    if len(unique) <= 1:
        return etiquetas
    normed = [string_norm_label(u) for u in unique]
    n = len(unique)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for i in range(n):
        if not normed[i]:
            continue
        for j in range(i + 1, n):
            if not normed[j] or find(i) == find(j):
                continue
            if SequenceMatcher(None, normed[i], normed[j]).ratio() >= umbral:
                union(i, j)
    for i in range(n):
        if not normed[i]:
            continue
        tokens_i = set(normed[i].split())
        if len(tokens_i) < 2:
            continue
        for j in range(i + 1, n):
            if not normed[j] or find(i) == find(j):
                continue
            tokens_j = set(normed[j].split())
            if len(tokens_j) < 2:
                continue
            interseccion = tokens_i & tokens_j
            menor = min(len(tokens_i), len(tokens_j))
            if menor > 0 and len(interseccion) / menor >= 0.6:
                union(i, j)
    le = get_embeddings_batch(unique)
    vp = [(i, le[i]) for i in range(n) if le[i] is not None]
    if len(vp) >= 2:
        vi, vv = zip(*vp)
        sm = cosine_similarity(np.array(vv))
        for pi in range(len(vi)):
            for pj in range(pi + 1, len(vi)):
                if sm[pi][pj] >= umbral:
                    if find(vi[pi]) != find(vi[pj]):
                        union(vi[pi], vi[pj])
    freq = Counter(etiquetas)
    grupos = defaultdict(list)
    for i in range(n):
        grupos[find(i)].append(i)
    canon = {}
    for root, members in grupos.items():
        cands = [unique[m] for m in members]
        vc = [c for c in cands if c.lower() not in ("sin tema", "varios") and _frase_esta_completa(c)]
        va = [c for c in cands if c.lower() not in ("sin tema", "varios")]
        if vc:
            canon[root] = max(vc, key=lambda c: (freq[c], len(c)))
        elif va:
            best = max(va, key=lambda c: (freq[c], len(c)))
            r = _recortar_frase_completa(best)
            canon[root] = r if _frase_esta_completa(r) else best
        else:
            canon[root] = cands[0]
    lm = {unique[i]: canon[find(i)] for i in range(n)}
    return [capitalizar_etiqueta(lm.get(e, e)) for e in etiquetas]


def _fusionar_subtemas_semanticos(subtemas, textos_por_subtema, marca, aliases, umbral=UMBRAL_FUSION_SUBTEMAS):
    unique_subs = list(dict.fromkeys(subtemas))
    if len(unique_subs) <= 1:
        return subtemas
    repr_texts = []
    for sub in unique_subs:
        txts = textos_por_subtema.get(sub, [])
        palabras = []
        for t in txts[:20]:
            for w in string_norm_label(str(t)).split():
                if len(w) > 3:
                    palabras.append(w)
        top_kw = " ".join(w for w, _ in Counter(palabras).most_common(10))
        repr_texts.append(f"{sub}. {sub}. {sub}. {top_kw}"[:600])
    emb_repr = get_embeddings_batch(repr_texts)
    valid = [(i, emb_repr[i]) for i in range(len(unique_subs)) if emb_repr[i] is not None]
    if len(valid) < 2:
        return subtemas
    v_idx, v_emb = zip(*valid)
    sim = cosine_similarity(np.array(v_emb))
    n = len(v_idx)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for i in range(n):
        for j in range(i + 1, n):
            if find(i) == find(j):
                continue
            if sim[i][j] >= umbral:
                union(i, j)
    grupos = defaultdict(list)
    for i in range(n):
        grupos[find(i)].append(v_idx[i])
    freq = Counter(subtemas)
    lm = {}
    for root, members in grupos.items():
        cands = [unique_subs[m] for m in members]
        if len(cands) == 1:
            lm[cands[0]] = cands[0]
            continue
        vc = [c for c in cands if c.lower() not in ("sin tema", "varios") and _frase_esta_completa(c)]
        best = max(vc, key=lambda c: (freq.get(c, 0), len(c))) if vc else max(cands, key=lambda c: (freq.get(c, 0), len(c)))
        if len(cands) <= 3:
            unified = _unificar_subtemas_llm(cands, textos_por_subtema, marca, aliases)
            if unified and _frase_esta_completa(unified):
                best = unified
        for c in cands:
            lm[c] = capitalizar_etiqueta(best)
    return [lm.get(s, s) for s in subtemas]


def _unificar_subtemas_llm(subtemas_a_unificar, textos_por_subtema, marca, aliases):
    subs_str = "\n".join(f"  · {s}" for s in subtemas_a_unificar)
    all_kw = []
    for sub in subtemas_a_unificar:
        for t in textos_por_subtema.get(sub, [])[:5]:
            for w in string_norm_label(str(t)).split():
                if len(w) > 3:
                    all_kw.append(w)
    kw_str = " · ".join(w for w, _ in Counter(all_kw).most_common(8))
    prompt = (
        f"Estos subtemas son variaciones del MISMO tema. "
        f"Genera UN subtema unificado (4-6 palabras) como frase nominal completa:\n\n"
        f"{subs_str}\n\nKeywords: {kw_str}\n\n"
        "REGLAS: frase coherente con preposición (de/del/para/sobre/en), "
        "sin marcas ni ciudades, tildes y ñ correctas, terminar en sustantivo/adjetivo.\n"
        "CORRECTO: 'Regulación de tarifas eléctricas', 'Apertura de nuevas sucursales'\n"
        "INCORRECTO: 'Tarifas energía', 'Apertura sucursales', 'Actividad corporativa'\n"
        'JSON: {"subtema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=60,
            temperature=0.05,
            response_format={"type": "json_object"}
        )
        u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
        if u:
            st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
            st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
        raw = json.loads(resp.choices[0].message.content).get("subtema", "")
        if raw:
            return limpiar_tema_geografico(limpiar_tema(raw), marca, aliases)
    except:
        pass
    return None


def get_embeddings_batch(textos, batch_size=100):
    if not textos:
        return []
    cache = get_embedding_cache()
    resultados, missing = cache.get_many(textos)
    if not missing:
        return resultados
    mt = [textos[i][:2000] if textos[i] else "" for i in missing]
    for i in range(0, len(mt), batch_size):
        batch = mt[i:i + batch_size]
        bidx = missing[i:i + batch_size]
        try:
            resp = call_with_retries(openai.Embedding.create, input=batch, model=OPENAI_MODEL_EMBEDDING)
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_embedding'] += (u.get('total_tokens') if isinstance(u, dict) else getattr(u, 'total_tokens', 0)) or 0
            for j, d in enumerate(resp["data"]):
                oi = bidx[j]
                emb = d["embedding"]
                resultados[oi] = emb
                cache.put(textos[oi], emb)
        except:
            for j, t in enumerate(batch):
                oi = bidx[j]
                try:
                    r = openai.Embedding.create(input=[t], model=OPENAI_MODEL_EMBEDDING)
                    emb = r["data"][0]["embedding"]
                    resultados[oi] = emb
                    cache.put(textos[oi], emb)
                except:
                    pass
    return resultados


class DSU:
    def __init__(self, n):
        self.p = list(range(n))
        self.rank = [0] * n

    def find(self, i):
        path = []
        while self.p[i] != i:
            path.append(i)
            i = self.p[i]
        for node in path:
            self.p[node] = i
        return i

    def union(self, i, j):
        ri, rj = self.find(i), self.find(j)
        if ri == rj:
            return
        if self.rank[ri] < self.rank[rj]:
            ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]:
            self.rank[ri] += 1

    def grupos(self, n):
        c = defaultdict(list)
        for i in range(n):
            c[self.find(i)].append(i)
        return dict(c)


def agrupar_textos_similares(textos, umbral):
    if not textos:
        return {}
    embs = get_embeddings_batch(textos)
    valid = [(i, e) for i, e in enumerate(embs) if e is not None]
    if len(valid) < 2:
        return {}
    idxs, M = zip(*valid)
    labels = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral, metric="cosine", linkage="average"
    ).fit(np.array(M)).labels_
    g = defaultdict(list)
    for k, lbl in enumerate(labels):
        g[lbl].append(idxs[k])
    return dict(enumerate(g.values()))


def agrupar_por_titulo_similar(titulos):
    gid, grupos, used = 0, {}, set()
    norm = [normalize_title_for_comparison(t) for t in titulos]
    for i in range(len(norm)):
        if i in used or not norm[i]:
            continue
        grp = [i]
        used.add(i)
        for j in range(i + 1, len(norm)):
            if j in used or not norm[j]:
                continue
            if SequenceMatcher(None, norm[i], norm[j]).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                grp.append(j)
                used.add(j)
        if len(grp) >= 2:
            grupos[gid] = grp
            gid += 1
    return grupos


def seleccionar_representante(indices, textos):
    embs = get_embeddings_batch([textos[i] for i in indices])
    validos = [(indices[k], e) for k, e in enumerate(embs) if e is not None]
    if not validos:
        return indices[0], textos[indices[0]]
    idxs, M = zip(*validos)
    centro = np.mean(M, axis=0, keepdims=True)
    best = int(np.argmax(cosine_similarity(np.array(M), centro)))
    return idxs[best], textos[idxs[best]]


_SENT_SPLIT = re.compile(r'(?<=[.!?;])\s+|(?<=\n)')


def _split_sentences(text):
    parts = _SENT_SPLIT.split(text)
    sents = [p.strip() for p in parts if len(p.strip()) > 15]
    return sents if sents else [text[:600]]


# ======================================
# TONO
# ======================================
class ClasificadorTono:
    def __init__(self, marca, aliases):
        self.marca = marca.strip()
        self.aliases = [a.strip() for a in (aliases or []) if a.strip()]
        self._all_names = [self.marca] + self.aliases
        patterns = [re.escape(unidecode(n.lower())) for n in self._all_names]
        self.brand_re = re.compile(r"\b(" + "|".join(patterns) + r")\b", re.IGNORECASE) if patterns else re.compile(r"(a^b)")

    def _extraer_oraciones_marca(self, texto):
        oraciones = _split_sentences(texto)
        resultado = []
        for i, sent in enumerate(oraciones):
            if self.brand_re.search(unidecode(sent.lower())):
                partes = []
                if i > 0:
                    partes.append(oraciones[i - 1])
                partes.append(sent)
                if i < len(oraciones) - 1:
                    partes.append(oraciones[i + 1])
                resultado.append(" ".join(partes).strip())
        return list(dict.fromkeys(resultado))[:5]

    def _es_sujeto(self, oracion):
        on = unidecode(oracion.lower())
        m = self.brand_re.search(on)
        return m and m.start() < len(on) * 0.6

    def _sentimiento_oracion(self, oracion):
        on = unidecode(oracion.lower())
        bf = self.brand_re.search(on)
        if not bf:
            return 0, 0
        contexto = on[max(0, bf.start() - 80):bf.end() + 80]
        neg_near = bool(re.search(
            r'\b(no|sin|nunca|jamás|niega|rechaza|desmiente|tampoco|ni|denuncia|demanda|sanciona|multa|critica|ataca|fraude|corrupcion|irregularidad)\b',
            contexto, re.IGNORECASE
        ))
        verbo_pos = bool(re.search(
            r'\b(lanz|inaugur|estren|anunc|cre[ao]|construy|abr[ei]|inici|implement|desarroll|inviert|expand|fortalec|mejor|benefici|gan|crec|lidere?|lider[ao]s?|lideran[do]?|lidere?mos|aliad|celebr|reconoc|premi|solucion[ao]|resuelv|atiend|respond|present[a]|firm[a]|entreg[a]|inici[a]|refuer[a-z]*|consolid[a-z]*|destac[a-z]*|avanz[a-z]*|promuev[a-z]*|impuls[a-z]*)\b',
            contexto, re.IGNORECASE
        ))
        verbo_neg = bool(re.search(
            r'\b(cae|perdi|fall|suspend|cerr|renunc|huelg|ataqu|hacke|boicot|reclam|perdid|deficit|conflict|disput|rechaz|proble|riesg|traged|cris|emerg|desastr|inundac|desliz|damnif|deterior|irregular|corrupc|evas|sancion|multa|denunc|demand|investig|critic)\b',
            contexto, re.IGNORECASE
        ))
        if CRISIS_KEYWORDS.search(on) and RESPONSE_VERBS.search(on):
            if self._es_sujeto(oracion):
                return 3, 0
        ph = sum(1 for p in POS_PATTERNS if p.search(on))
        nh = sum(1 for p in NEG_PATTERNS if p.search(on))
        if verbo_pos:
            ph += 2
        if verbo_neg:
            nh += 2
        w = 1.2 if self._es_sujeto(oracion) else 0.3
        if neg_near:
            return int(nh * w * 1.3), int(ph * w)
        return int(ph * w), int(nh * w)

    def _reglas(self, oraciones):
        tp, tn = 0, 0
        sujeto_pos = 0
        sujeto_neg = 0
        for s in oraciones:
            p, n = self._sentimiento_oracion(s)
            tp += p
            tn += n
            if self._es_sujeto(s):
                if p > n:
                    sujeto_pos += 1
                elif n > p:
                    sujeto_neg += 1
        if sujeto_pos >= 1 and sujeto_neg == 0:
            return "Positivo"
        if sujeto_neg >= 1 and sujeto_pos == 0:
            return "Negativo"
        if tp >= 2 and tp > tn:
            return "Positivo"
        if tn >= 2 and tn > tp:
            return "Negativo"
        if sujeto_pos > sujeto_neg:
            return "Positivo"
        if sujeto_neg > sujeto_pos:
            return "Negativo"
        return None

    async def _llm(self, oraciones, texto):
        fragmentos = "\n".join(f"  → {s[:300]}" for s in oraciones[:4])
        prompt = (
            f"Evalúa el sentimiento EXCLUSIVAMENTE hacia '{self.marca}'"
            f" (alias: {', '.join(self.aliases) if self.aliases else 'N/A'}) "
            f"en los fragmentos donde se le menciona.\n"
            f"REGLAS CLAVE:\n"
            f"- Evalúa SOLO cómo se habla de '{self.marca}', NO el tono general de la noticia.\n"
            f"- Si la noticia es negativa pero '{self.marca}' es mencionada positivamente → Positivo.\n"
            f"- Si la noticia es positiva pero '{self.marca}' es criticada → Negativo.\n"
            f"- Si '{self.marca}' solo se menciona de paso sin juicio → Neutro.\n"
            f"- Competidor negativo NO hace a '{self.marca}' positivo → Neutro.\n"
            f"FRAGMENTOS CON MENCIÓN:\n{fragmentos}\n"
            f'JSON: {{"tono":"Positivo|Negativo|Neutro"}}'
        )
        try:
            resp = await acall_with_retries(
                openai.ChatCompletion.acreate,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=50,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
            tono = str(json.loads(resp.choices[0].message.content).get("tono", "Neutro")).strip().title()
            return {"tono": tono if tono in ("Positivo", "Negativo", "Neutro") else "Neutro"}
        except:
            return {"tono": "Neutro"}

    async def _clasificar(self, texto, sem):
        async with sem:
            om = self._extraer_oraciones_marca(texto)
            if not om:
                return {\"tono\": \"Neutro\"}
            r = self._reglas(om)
            if r:
                return {\"tono\": r}
            # Fallback: override para textos con evidencia clara positiva
            full_text = unidecode(texto.lower())
            strong_pos = [
                'liderando', 'reconocimiento', 'reconocido', 'premiado', 'celebra',
                'exito', 'éxito', 'mejores resultados', 'sostenibilidad,', 'innovador',
                'refuerza', 'impulso', 'impulsa', 'destaca por'
            ]
            if any(sp in full_text for sp in strong_pos):
                return {\"tono\": \"Positivo\"}
            return await self._llm(om, texto)

    async def procesar_lote_async(self, textos, pbar, resumenes, titulos):
        n = len(textos)
        txts = textos.tolist()
        pbar.progress(0.05, "Agrupando para tono...")
        txts_emb = [texto_para_embedding(str(titulos.iloc[i]), str(resumenes.iloc[i])) for i in range(n)]
        dsu = DSU(n)
        for g in [agrupar_textos_similares(txts_emb, SIMILARITY_THRESHOLD_TONO), agrupar_por_titulo_similar(titulos.tolist())]:
            for _, idxs in g.items():
                for j in idxs[1:]:
                    dsu.union(idxs[0], j)
        grupos = dsu.grupos(n)
        reps = {cid: seleccionar_representante(idxs, txts)[1] for cid, idxs in grupos.items()}
        sem = asyncio.Semaphore(CONCURRENT_REQUESTS)
        cids = list(reps.keys())
        tasks = [self._clasificar(reps[c], sem) for c in cids]
        rl = []
        for i, f in enumerate(asyncio.as_completed(tasks)):
            rl.append(await f)
            pbar.progress(0.1 + 0.85 * (i + 1) / len(tasks), f"Tono {i + 1}/{len(tasks)}")
        rpg = {cids[i]: r for i, r in enumerate(rl)}
        final = [None] * n
        for cid, idxs in grupos.items():
            r = rpg.get(cid, {"tono": "Neutro"})
            for i in idxs:
                final[i] = r
        pbar.progress(1.0, "Tono completado")
        return final


def analizar_tono_con_pkl(textos, pkl_file):
    try:
        pipeline = joblib.load(pkl_file)
        TM = {1: "Positivo", "1": "Positivo", 0: "Neutro", "0": "Neutro", -1: "Negativo", "-1": "Negativo"}
        return [{"tono": TM.get(p, str(p).title())} for p in pipeline.predict(textos)]
    except Exception as e:
        st.error(f"Error pkl: {e}")
        return None


# ======================================
# SUBTEMAS  (con umbrales adaptativos)
# ======================================
class ClasificadorSubtema:
    def __init__(self, marca, aliases):
        self.marca = marca
        self.aliases = aliases or []
        self._cache = {}
        # Se inicializa en procesar_lote con los umbrales del corpus
        self._umbrales: dict = {}

    def _paso1(self, titulos, resumenes, dsu):
        def nt(t, n):
            return ' '.join(re.sub(r'[^a-z0-9\s]', '', unidecode(str(t).lower())).split()[:n])

        bt, br = defaultdict(list), defaultdict(list)
        for i, (ti, re_) in enumerate(zip(titulos, resumenes)):
            a, b = nt(ti, 40), nt(re_, 15)
            if a:
                bt[hashlib.md5(a.encode()).hexdigest()].append(i)
            if b:
                br[hashlib.md5(b.encode()).hexdigest()].append(i)
        for bk in (bt, br):
            for idxs in bk.values():
                for j in idxs[1:]:
                    dsu.union(idxs[0], j)

    def _paso2(self, titulos, dsu):
        norm = [normalize_title_for_comparison(t) for t in titulos]
        n = len(norm)
        for i in range(n):
            if not norm[i]:
                continue
            for j in range(i + 1, n):
                if not norm[j] or dsu.find(i) == dsu.find(j):
                    continue
                if SequenceMatcher(None, norm[i], norm[j]).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    dsu.union(i, j)

    def _paso2b_keywords(self, titulos, dsu, ae):
        """
        Agrupa por keywords raras — pero SOLO si la similitud semántica real
        entre los dos textos supera sim_minima_keywords.
        Esto evita que una palabra compartida por azar agrupe noticias distintas.
        """
        sim_min = self._umbrales.get('sim_minima_keywords', SIM_MINIMA_KEYWORDS_RARAS)
        stop = {
            'el','la','los','las','un','una','unos','unas','de','del','al',
            'en','con','por','para','que','se','su','sus','es','son','fue',
            'como','mas','pero','sin','sobre','entre','tras','esta','este',
            'esto','hay','ser','han','ha','ya','muy','otro','otra','otros',
            'otras','todo','toda','todos','todas','puede','desde','hasta',
            'donde','cuando','quien','cual','cada','nos','les','ante','bajo',
            'nueva','nuevo','nuevos','nuevas','forma','hace','asi','sera',
            'segun','tiene','fueron','sido','hacer','dice','dijo','tambien',
        }
        titulo_words = []
        for t in titulos:
            ws = set()
            for w in re.findall(r'[a-z]+', unidecode(str(t).lower())):
                if len(w) >= 5 and w not in stop:
                    ws.add(w)
            titulo_words.append(ws)
        word_freq = Counter()
        for ws in titulo_words:
            for w in ws:
                word_freq[w] += 1
        n = len(titulos)
        max_freq = max(2, int(n * 0.03))
        rare_index = defaultdict(list)
        for i, ws in enumerate(titulo_words):
            for w in ws:
                if 2 <= word_freq[w] <= max_freq:
                    rare_index[w].append(i)
        # Verificar similitud semántica real antes de unir
        for idxs in rare_index.values():
            for a in range(len(idxs)):
                for b in range(a + 1, len(idxs)):
                    ia, ib = idxs[a], idxs[b]
                    if dsu.find(ia) == dsu.find(ib):
                        continue
                    ea, eb = ae[ia], ae[ib]
                    if ea is None or eb is None:
                        continue
                    sim = cosine_similarity(
                        np.array(ea).reshape(1, -1),
                        np.array(eb).reshape(1, -1)
                    )[0][0]
                    if sim >= sim_min:
                        dsu.union(ia, ib)

    def _paso3(self, et, ae, dsu, pbar, ps):
        """
        Clustering semántico con verificación post-hoc de similitud mínima.
        Después del clustering, cualquier par en el mismo grupo cuya similitud
        sea menor que sim_minima_agrupacion se separa (el miembro más lejano
        del centroide queda en su propio grupo).
        """
        umbral_cluster = self._umbrales.get('subtema', UMBRAL_SUBTEMA)
        sim_min = self._umbrales.get('sim_minima_agrupacion', SIM_MINIMA_AGRUPACION_SUBTEMA)
        n = len(et)
        if n < 2:
            return
        B = 500
        if n <= B:
            pbar.progress(ps, "Clustering semántico...")
            ok = [(k, e) for k, e in enumerate(ae) if e is not None]
            if len(ok) < 2:
                return
            io_, M = zip(*ok)
            sim_matrix = cosine_similarity(np.array(M))

            # Con pocos elementos (n ≤ 10) usamos enlace completo para
            # que el umbral sea la similitud MÍNIMA del grupo, no la promedio.
            linkage = 'complete' if n <= 10 else 'average'
            labels = AgglomerativeClustering(
                n_clusters=None, distance_threshold=1 - umbral_cluster,
                metric='precomputed', linkage=linkage
            ).fit(1 - sim_matrix).labels_
            g = defaultdict(list)
            for k, lbl in enumerate(labels):
                g[lbl].append(io_[k])

            # Post-hoc: verificar que cada miembro supera sim_min con el centroide
            for cl in g.values():
                if len(cl) < 2:
                    continue
                vecs = np.array([ae[i] for i in cl if ae[i] is not None])
                if len(vecs) < 2:
                    continue
                centroid = np.mean(vecs, axis=0)
                sims_al_centroid = cosine_similarity(vecs, centroid.reshape(1, -1)).flatten()
                # Solo une si TODOS superan sim_min; si alguno no, lo dejamos solo
                todos_ok = all(s >= sim_min for s in sims_al_centroid)
                if todos_ok:
                    for j in cl[1:]:
                        dsu.union(cl[0], j)
                else:
                    # Une solo los que superan sim_min con el más central
                    mejor_idx = int(np.argmax(sims_al_centroid))
                    repr_vec = np.array(ae[cl[mejor_idx]]).reshape(1, -1)
                    for k_local, i_global in enumerate(cl):
                        if ae[i_global] is None:
                            continue
                        sim_vs_repr = cosine_similarity(
                            np.array(ae[i_global]).reshape(1, -1), repr_vec
                        )[0][0]
                        if sim_vs_repr >= sim_min:
                            dsu.union(cl[mejor_idx], i_global)
            pbar.progress(ps + 0.18, "Clustering completado")
            return

        # Corpus grande: batches
        tb = max(1, (n + B - 1) // B)
        for bn_, bs in enumerate(range(0, n, B)):
            bi = list(range(bs, min(bs + B, n)))
            ok = [(idx, ae[idx]) for idx in bi if ae[idx] is not None]
            if len(ok) < 2:
                continue
            io_, M = zip(*ok)
            sim_matrix = cosine_similarity(np.array(M))
            labels = AgglomerativeClustering(
                n_clusters=None, distance_threshold=1 - umbral_cluster,
                metric='precomputed', linkage='average'
            ).fit(1 - sim_matrix).labels_
            g = defaultdict(list)
            for k, lbl in enumerate(labels):
                g[lbl].append(io_[k])
            for cl in g.values():
                if len(cl) < 2:
                    continue
                vecs = np.array([ae[i] for i in cl if ae[i] is not None])
                if len(vecs) < 2:
                    continue
                centroid = np.mean(vecs, axis=0)
                sims = cosine_similarity(vecs, centroid.reshape(1, -1)).flatten()
                mejor_idx = int(np.argmax(sims))
                repr_vec = np.array(ae[cl[mejor_idx]]).reshape(1, -1)
                for k_local, i_global in enumerate(cl):
                    if ae[i_global] is None:
                        continue
                    s = cosine_similarity(np.array(ae[i_global]).reshape(1, -1), repr_vec)[0][0]
                    if s >= sim_min:
                        dsu.union(cl[mejor_idx], i_global)
            pbar.progress(ps + 0.15 * (bn_ + 1) / tb, f"Clustering {bn_ + 1}/{tb}...")

        pbar.progress(ps + 0.16, "Unificando...")
        usar_fusion = self._umbrales.get('usar_fusion_iterativa', True)
        if usar_fusion:
            self._fusion(et, ae, dsu, pbar, ps + 0.16)

    def _fusion(self, textos, ae, dsu, pbar, ps):
        n = len(textos)
        umbral_inter = self._umbrales.get('fusion_intergrupo', UMBRAL_FUSION_INTERGRUPO)
        max_iter = self._umbrales.get('max_iter_fusion', MAX_ITER_FUSION)
        sim_min = self._umbrales.get('sim_minima_agrupacion', SIM_MINIMA_AGRUPACION_SUBTEMA)
        for it in range(max_iter):
            grupos = dsu.grupos(n)
            if len(grupos) < 2:
                break
            centroids, vg = [], []
            for gid, idxs in grupos.items():
                vecs = [ae[i] for i in idxs[:50] if ae[i] is not None]
                if vecs:
                    centroids.append(np.mean(vecs, axis=0))
                    vg.append(gid)
            if len(vg) < 2:
                break
            sim = cosine_similarity(np.array(centroids))
            # Usar el máximo de umbral_inter y sim_min para fusión
            umbral_efectivo = max(umbral_inter, sim_min)
            pairs = sorted(
                [(sim[i][j], i, j) for i in range(len(vg)) for j in range(i + 1, len(vg))
                 if sim[i][j] >= umbral_efectivo],
                reverse=True
            )
            fus = 0
            for _, i, j in pairs:
                ri, rj = grupos[vg[i]][0], grupos[vg[j]][0]
                if dsu.find(ri) != dsu.find(rj):
                    dsu.union(ri, rj)
                    fus += 1
            pbar.progress(min(ps + 0.04 * (it + 1), 0.52), f"Fusión {it + 1}: {fus}")
            if fus == 0:
                break

    def _extraer_keywords_titulos(self, titulos_grp: list, top_n: int = 6) -> list:
        palabras = []
        for t in titulos_grp[:10]:
            for w in string_norm_label(t).split():
                if len(w) > 3:
                    palabras.append(w)
        return [w for w, _ in Counter(palabras).most_common(top_n)]

    def _generar_etiqueta(self, textos_grp, titulos_grp, resumenes_grp):
        tn = sorted(set(normalize_title_for_comparison(t) for t in titulos_grp if t))
        ck = hashlib.md5(("|".join(tn[:12]) + f"#{len(titulos_grp)}").encode()).hexdigest()
        if ck in self._cache:
            return self._cache[ck]

        tm = list(dict.fromkeys(t[:130] for t in titulos_grp if t))[:6]
        rm = [str(r)[:200] for r in resumenes_grp[:3] if r and len(str(r)) > 20]

        kw_list = self._extraer_keywords_titulos(titulos_grp, top_n=8)
        palabras_res = []
        for r in resumenes_grp[:5]:
            for w in string_norm_label(str(r)).split():
                if len(w) > 4:
                    palabras_res.append(w)
        kw_res = [w for w, _ in Counter(palabras_res).most_common(4)
                  if w not in {unidecode(k.lower()) for k in kw_list}]
        kw_todos = kw_list + kw_res
        kw = ", ".join(kw_todos[:10])

        ctx_resumenes = (
            "\nRESÚMENES (para contexto):\n"
            + "\n".join(f"  · {r}" for r in rm)
        ) if rm else ""

        if len(kw_list) >= 3:
            ejemplo_dinamico = (
                f"'{kw_list[0].title()} de {kw_list[1].title()}' o "
                f"'{kw_list[0].title()} del {kw_list[2].title()}'"
            )
        elif len(kw_list) >= 2:
            ejemplo_dinamico = f"'{kw_list[0].title()} de {kw_list[1].title()}'"
        elif len(kw_list) == 1:
            ejemplo_dinamico = f"'{kw_list[0].title()} en la región'"
        else:
            ejemplo_dinamico = "'Proyecto de terminal de transportes'"

        prompt = (
            "Eres editor jefe de un periódico. "
            "Genera UN subtema periodístico (4-7 palabras) que sea una FRASE NOMINAL "
            "— sin sujeto ni verbo conjugado — para este grupo de noticias.\n\n"
            "TÍTULOS:\n" + "\n".join(f"  · {t}" for t in tm)
            + ctx_resumenes
            + f"\n\nPALABRAS CLAVE: {kw}\n\n"
            "REGLAS OBLIGATORIAS:\n"
            "  1. FRASE NOMINAL PURA: empieza con sustantivo, usa preposición para unir conceptos.\n"
            "     NUNCA empieces con cargo/persona ('Alcalde', 'Gobernador', 'Ministro').\n"
            "     NUNCA incluyas verbo conjugado ('presenta', 'anuncia', 'lanza', 'inaugura').\n"
            f"     CORRECTO: {ejemplo_dinamico}\n"
            "     INCORRECTO: 'Alcalde presenta proyecto terminal', "
            "'Gobernador anuncia inversión', 'Alcaldía lanza plan'\n"
            "  2. USA preposiciones (de, del, para, sobre, en, por) para conectar conceptos.\n"
            "  3. SÉ ESPECÍFICO: describe el asunto real, no el actor.\n"
            "  4. Ciudades y regiones SÍ pueden aparecer si son relevantes al tema.\n"
            "  5. Sin nombre de marcas privadas. Tildes y ñ correctas.\n\n"
            "EJEMPLOS CORRECTOS: 'Proyecto de terminal de transportes', "
            "'Operación del Canal del Dique', 'Plan de infraestructura vial', "
            "'Regulación de tarifas eléctricas', 'Inversión en salud pública'\n"
            "EJEMPLOS INCORRECTOS: 'Alcalde presenta proyecto', 'Gobernador lanza plan', "
            "'Tarifas energía', 'Gestión corporativa', 'Actividad legislativa'\n\n"
            'JSON: {"subtema":"..."}'
        )

        _VERBOS_FRASES = re.compile(
            r'\b(presenta|presentan|anuncia|anuncian|lanza|lanzan|inaugura|inauguran|'
            r'realiza|realizan|desarrolla|desarrollan|ejecuta|ejecutan|gestiona|gestionan|'
            r'impulsa|impulsan|promueve|promueven|lidera|lideran|encabeza|encabezan|'
            r'aprueba|aprueban|firma|firman|suscribe|suscriben|invierte|invierten|'
            r'construye|construyen|instala|instalan|entrega|entregan|recibe|reciben|'
            r'solicita|solicitan|visita|visitan|atiende|atienden|destaca|destacan|'
            r'señala|señalan|indica|indican|expresa|expresan|afirma|afirman|'
            r'propone|proponen|pide|piden|exige|exigen|apoya|apoyan|'
            r'informa|informan|reporta|reportan|advierte|advierten)\b',
            re.IGNORECASE
        )

        def _tiene_verbo_conjugado(s):
            return bool(_VERBOS_FRASES.search(s))

        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=60,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0

            raw = json.loads(resp.choices[0].message.content).get("subtema", "Varios")
            et = limpiar_tema_geografico(limpiar_tema(raw), self.marca, self.aliases)

            if not et or et.strip().lower() == "sin tema":
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)

            if _tiene_verbo_conjugado(et):
                et = self._refinar(tm, kw, rm, forzar_preposicion=True, prohibir_verbos=True)

            def _es_robotico(s):
                palabras = s.split()
                if len(palabras) <= 3:
                    nexos = {"de", "del", "para", "sobre", "en", "con", "por",
                             "ante", "hacia", "entre", "sin", "al", "las", "los",
                             "una", "uno", "que", "como", "y", "o", "a", "e", "u"}
                    tiene_nexo = any(unidecode(p.lower()) in nexos for p in palabras[1:])
                    if not tiene_nexo:
                        return True
                return False

            genericas = {"gestión", "gestion", "actividades", "acciones", "noticias",
                         "información", "informacion", "eventos", "varios", "sin tema",
                         "actividad corporativa", "gestion corporativa"}
            es_gen = string_norm_label(et) in {string_norm_label(g) for g in genericas}
            es_rob = _es_robotico(et)

            if es_gen or es_rob or len(et.split()) < 3:
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)

            if not _validar_estructura_subtema(et):
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)
                if not _validar_estructura_subtema(et):
                    et = self._fallback(titulos_grp)

            et = _validar_etiqueta_completa(
                et, titulos_grp=titulos_grp, resumenes_grp=resumenes_grp,
                marca=self.marca, aliases=self.aliases, fallback_fn=self._fallback
            )
        except:
            et = self._fallback(titulos_grp)

        et = capitalizar_etiqueta(et)
        self._cache[ck] = et
        return et

    def _refinar(self, titulos, kw, resumenes=None, forzar_preposicion=False, prohibir_verbos=False):
        ctx = (
            "\nContexto de resúmenes: " + " | ".join(r[:100] for r in resumenes[:3])
        ) if resumenes else ""
        kw_parts = [w.strip() for w in kw.split(",") if w.strip()]

        if len(kw_parts) >= 3:
            ej_bueno = (
                f"'{kw_parts[0].title()} de {kw_parts[1].title()}', "
                f"'{kw_parts[0].title()} en {kw_parts[2].title()}'"
            )
        elif len(kw_parts) >= 2:
            ej_bueno = f"'{kw_parts[0].title()} de {kw_parts[1].title()}'"
        elif len(kw_parts) == 1:
            ej_bueno = f"'{kw_parts[0].title()} en la región'"
        else:
            ej_bueno = "'Proyecto de terminal de transportes'"

        if len(kw_parts) >= 2:
            ej_malo = f"'{kw_parts[0].title()} {kw_parts[1].title()}' (sin preposición)"
        else:
            ej_malo = "'Actividad corporativa', 'Gestión institucional'"

        instruccion_prep = (
            "  OBLIGATORIO: usa una preposición (de, del, para, sobre, en, por) "
            "entre los conceptos. NUNCA dos sustantivos pegados sin nexo.\n"
        ) if forzar_preposicion else ""

        instruccion_verbo = (
            "  PROHIBIDO: verbos conjugados ('presenta', 'anuncia', 'lanza', 'inaugura', etc.). "
            "Solo frases nominales (sustantivos + preposiciones).\n"
            "  NUNCA empieces con cargo ('Alcalde', 'Gobernador', 'Ministro', 'Director').\n"
        ) if prohibir_verbos else ""

        prompt = (
            "Eres editor jefe. Genera UN subtema periodístico (4-7 palabras) "
            "como frase nominal sin verbo conjugado.\n\n"
            f"Títulos: {' | '.join(titulos[:5])}{ctx}\n"
            f"Keywords: {kw}\n\n"
            f"{instruccion_prep}"
            f"{instruccion_verbo}"
            f"CORRECTO: {ej_bueno}, 'Tarifas de energía eléctrica'\n"
            f"INCORRECTO: {ej_malo}, 'Alcalde presenta plan'\n"
            "Tildes y ñ correctas. Sin marcas privadas.\n"
            'JSON: {"subtema":"..."}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=60,
                temperature=0.2,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
            raw = json.loads(resp.choices[0].message.content).get("subtema", "Varios")
            et = limpiar_tema_geografico(limpiar_tema(raw), self.marca, self.aliases)
            if not _frase_esta_completa(et):
                et = _recortar_frase_completa(et)
                if not _frase_esta_completa(et):
                    return self._fallback(titulos)
            return et
        except:
            return self._fallback([])

    def _fallback(self, titulos):
        if not titulos:
            return "Cobertura informativa general"
        palabras = []
        for t in titulos[:5]:
            for w in string_norm_label(t).split():
                if len(w) > 4:
                    palabras.append(w)
        if palabras:
            top = [w for w, _ in Counter(palabras).most_common(3)]
            if len(top) >= 2:
                frase = f"{top[0]} de {top[1]}"
                if _frase_esta_completa(frase):
                    return capitalizar_etiqueta(frase)
                return capitalizar_etiqueta(f"{top[0]} {top[1]}")
            return capitalizar_etiqueta(top[0])
        return "Cobertura informativa general"

    def procesar_lote(self, col, pbar, res_puros, tit_puros):
        textos   = col.tolist()
        titulos  = tit_puros.tolist()
        resumenes = res_puros.tolist()
        n = len(textos)

        # ── Calcular umbrales adaptativos según tamaño del corpus ──────────
        self._umbrales = _umbrales_adaptativos(n)
        u = self._umbrales
        st.caption(
            f"📐 Corpus: **{n}** noticias · "
            f"Umbral subtema: **{u['subtema']}** · "
            f"Sim mínima: **{u['sim_minima_agrupacion']}** · "
            f"Paso2b: {'✓' if u['usar_paso2b'] else '✗'} · "
            f"Fusión iterativa: {'✓' if u['usar_fusion_iterativa'] else '✗'}"
        )

        et = [texto_para_embedding(titulos[i], resumenes[i]) for i in range(n)]

        pbar.progress(0.05, "Fase 1 · Idénticas...")
        dsu = DSU(n)
        self._paso1(titulos, resumenes, dsu)
        pbar.progress(0.12, "Fase 2 · Títulos...")
        self._paso2(titulos, dsu)

        pbar.progress(0.18, "Embeddings...")
        ae = get_embeddings_batch(et)

        # Paso 2b: keywords raras — solo si el corpus es suficientemente grande
        if u['usar_paso2b']:
            pbar.progress(0.15, "Fase 2b · Keywords raras (con validación semántica)...")
            self._paso2b_keywords(titulos, dsu, ae)

        pbar.progress(0.20, "Fase 3 · Clustering...")
        self._paso3(et, ae, dsu, pbar, 0.20)

        gf = dsu.grupos(n)
        ng = len(gf)
        pbar.progress(0.55, f"Fase 4 · Etiquetando {ng} grupos...")
        mapa = {}
        sg = sorted(gf.items(), key=lambda x: -len(x[1]))
        for k, (lid, idxs) in enumerate(sg):
            if k % 10 == 0:
                pbar.progress(0.55 + 0.25 * (k / max(ng, 1)), f"Etiquetando {k + 1}/{ng}...")
            if len(idxs) > MAX_GRUPO_ETIQUETA:
                subgrupos = [idxs[i:i + MAX_GRUPO_ETIQUETA] for i in range(0, len(idxs), MAX_GRUPO_ETIQUETA)]
                for sg_ in subgrupos:
                    e = self._generar_etiqueta(
                        [textos[i] for i in sg_],
                        [titulos[i] for i in sg_],
                        [resumenes[i] for i in sg_]
                    )
                    for i in sg_:
                        mapa[i] = e
            else:
                e = self._generar_etiqueta(
                    [textos[i] for i in idxs],
                    [titulos[i] for i in idxs],
                    [resumenes[i] for i in idxs]
                )
                for i in idxs:
                    mapa[i] = e

        subtemas = [mapa.get(i, "Varios") for i in range(n)]

        pbar.progress(0.80, "Fase 4b · Coherencia etiqueta↔texto...")
        umbral_coherencia = u['coherencia_etiqueta']
        subtemas_unicos = list(set(subtemas))
        embs_sub_lista = get_embeddings_batch(subtemas_unicos)
        emb_subtemas = {sub: emb for sub, emb in zip(subtemas_unicos, embs_sub_lista) if emb is not None}

        incoherentes = 0
        for i in range(n):
            sub = subtemas[i]
            emb_txt = ae[i]
            emb_sub = emb_subtemas.get(sub)
            if emb_txt is None or emb_sub is None:
                continue
            sim = cosine_similarity(
                np.array(emb_txt).reshape(1, -1),
                np.array(emb_sub).reshape(1, -1)
            )[0][0]
            if sim < umbral_coherencia:
                mejor_sub, mejor_sim = sub, sim
                for otro_sub, emb_otro in emb_subtemas.items():
                    if otro_sub == sub:
                        continue
                    sim_otro = cosine_similarity(
                        np.array(emb_txt).reshape(1, -1),
                        np.array(emb_otro).reshape(1, -1)
                    )[0][0]
                    if sim_otro > mejor_sim:
                        mejor_sim = sim_otro
                        mejor_sub = otro_sub
                if mejor_sub != sub and mejor_sim > umbral_coherencia:
                    subtemas[i] = mejor_sub
                else:
                    nueva = self._generar_etiqueta([textos[i]], [titulos[i]], [resumenes[i]])
                    subtemas[i] = capitalizar_etiqueta(nueva)
                incoherentes += 1

        if incoherentes:
            st.caption(f"ℹ️ Fase 4b: {incoherentes} noticias reclasificadas por baja coherencia etiqueta↔texto.")

        pbar.progress(0.82, "Fase 5 · Dedup...")
        subtemas = dedup_labels(subtemas, u['dedup_label'])

        pbar.progress(0.86, "Fase 5b · Fusión semántica...")
        textos_por_sub = defaultdict(list)
        for i, s in enumerate(subtemas):
            textos_por_sub[s].append(textos[i])
        n_antes = len(set(subtemas))
        subtemas = _fusionar_subtemas_semanticos(
            subtemas, textos_por_sub, self.marca, self.aliases, u['fusion_subtemas']
        )
        n_despues = len(set(subtemas))
        if n_antes != n_despues:
            pbar.progress(0.89, f"Fusión: {n_antes}→{n_despues}")

        pbar.progress(0.90, "Fase 6 · Consistencia...")
        subtemas = self._consistencia(subtemas, ae, pbar, u)

        indices_reclass = [i for i, s in enumerate(subtemas) if s == "_RECLASIFICAR"]
        if indices_reclass:
            pbar.progress(0.93, f"Fase 6b · Reclasificando {len(indices_reclass)} noticias aisladas...")
            for i in indices_reclass:
                et_ind = self._generar_etiqueta([textos[i]], [titulos[i]], [resumenes[i]])
                subtemas[i] = capitalizar_etiqueta(et_ind)

        pbar.progress(0.93, "Fase 7 · Completitud...")
        subtemas = self._validar_completitud_final(subtemas, textos, titulos, resumenes)

        pbar.progress(0.97, "Fase 8 · Dedup final...")
        subtemas = dedup_labels(subtemas, u['dedup_label'])
        textos_por_sub2 = defaultdict(list)
        for i, s in enumerate(subtemas):
            textos_por_sub2[s].append(textos[i])
        subtemas = _fusionar_subtemas_semanticos(
            subtemas, textos_por_sub2, self.marca, self.aliases, u['fusion_subtemas']
        )

        # ── Coherencia final: fusionar subtemas equivalentes ──
        # Ej: "Dipping de tendencias gastronomicas" vs
        #     "Tendencia global en sabores para pollo" → MISMO subtema
        pbar.progress(0.98, "Coherencia final de subtemas...")
        orig_unicos = list(dict.fromkeys(subtemas))
        if len(orig_unicos) >= 2:
            cano_map: Dict[str, str] = {s: s for s in orig_unicos}
            emb_subs = get_embeddings_batch(orig_unicos)
            valid_pairs = [
                (i, emb_subs[i]) for i in range(len(orig_unicos))
                if emb_subs[i] is not None
            ]
            if len(valid_pairs) >= 2:
                v_idx, v_emb = zip(*valid_pairs)
                sim_subs = cosine_similarity(np.array(v_emb))
                for pi in range(len(v_idx)):
                    for pj in range(pi + 1, len(v_idx)):
                        oi, oj = v_idx[pi], v_idx[pj]
                        ci, cj = cano_map[orig_unicos[oi]], cano_map[orig_unicos[oj]]
                        if ci == cj:
                            continue
                        merged = False
                        # Criterio 1: embedding coseno entre labels >= 0.87
                        if sim_subs[pi][pj] >= 0.87:
                            merged = True
                        # Criterio 2: Jaccard tokens >= 0.45
                        if not merged:
                            ti = set(string_norm_label(orig_unicos[oi]).split())
                            tj = set(string_norm_label(orig_unicos[oj]).split())
                            if len(ti) >= 2 and len(tj) >= 2:
                                if len(ti & tj) / len(ti | tj) >= 0.45:
                                    merged = True
                        if merged:
                            fi = sum(1 for ss in subtemas if cano_map.get(ss) == ci)
                            fj = sum(1 for ss in subtemas if cano_map.get(ss) == cj)
                            winner = ci if fi >= fj else cj
                            loser = cj if winner == ci else ci
                            for k in cano_map:
                                if cano_map[k] == loser:
                                    cano_map[k] = winner
            n_antes = len(set(subtemas))
            subtemas = [cano_map.get(s, s) for s in subtemas]
            n_despues = len(set(subtemas))
            if n_antes != n_despues:
                st.caption(
                    f"ℹ️ Coherencia final: {n_antes} → {n_despues}"
                    f" ({n_antes - n_despues} fusionado(s))"
                )

        subtemas = dedup_labels(subtemas, u['dedup_label'])
        subtemas = [capitalizar_etiqueta(s) for s in subtemas]
        nf = len(set(subtemas))
        pbar.progress(1.0, f"{nf} subtemas")
        st.info(f"Subtemas: **{nf}** · Grupos originales: **{ng}**")
        return subtemas

    def _validar_completitud_final(self, subtemas, textos, titulos, resumenes):
        por_subtema = defaultdict(list)
        for i, s in enumerate(subtemas):
            por_subtema[s].append(i)
        resultado = list(subtemas)
        for sub, idxs in por_subtema.items():
            if _frase_esta_completa(sub):
                continue
            recortada = _recortar_frase_completa(sub)
            if _frase_esta_completa(recortada) and len(recortada.split()) >= 2:
                for i in idxs:
                    resultado[i] = capitalizar_etiqueta(recortada)
                continue
            tit_grp = [titulos[i] for i in idxs[:6]]
            res_grp = [resumenes[i] for i in idxs[:3]]
            nueva = _validar_etiqueta_completa(
                sub, titulos_grp=tit_grp, resumenes_grp=res_grp,
                marca=self.marca, aliases=self.aliases, fallback_fn=self._fallback
            )
            for i in idxs:
                resultado[i] = capitalizar_etiqueta(nueva)
        return resultado

    def _consistencia(self, subtemas, ae, pbar, umbrales=None):
        if umbrales is None:
            umbrales = self._umbrales
        min_sub = umbrales.get('min_pertenencia_subtema', UMBRAL_MIN_PERTENENCIA_SUBTEMA)
        ps = defaultdict(list)
        for i, s in enumerate(subtemas):
            ps[s].append(i)
        r = list(subtemas)
        centroids = {}
        for sub, idxs in ps.items():
            vecs = [ae[i] for i in idxs if ae[i] is not None]
            if vecs:
                centroids[sub] = np.mean(vecs, axis=0)
        for sub in [s for s in centroids if len(ps[s]) >= 3]:
            idxs = ps[sub]
            if sub.lower() in ("sin tema", "varios") or len(idxs) < 3:
                continue
            vi = [(i, ae[i]) for i in idxs if ae[i] is not None]
            if len(vi) < 3:
                continue
            v_i, v_v = zip(*vi)
            M = np.array(v_v)
            sims = cosine_similarity(M, centroids[sub].reshape(1, -1)).flatten()
            thr = max(0.60, np.mean(sims) - 2 * np.std(sims))
            for k, (oi, sv) in enumerate(zip(v_i, sims)):
                if sv >= thr:
                    continue
                bs, bsim = sub, sv
                emb = ae[oi]
                for os_, oc in centroids.items():
                    if os_ == sub:
                        continue
                    s2 = cosine_similarity(np.array(emb).reshape(1, -1), oc.reshape(1, -1))[0][0]
                    if s2 > bsim and s2 > 0.75:
                        bsim = s2
                        bs = os_
                if bs != sub:
                    r[oi] = bs
                elif sv < min_sub:
                    r[oi] = "_RECLASIFICAR"
        return r


# ======================================
# TEMAS  (con umbrales adaptativos)
# ======================================
def _construir_representacion_grupo(subtema, textos_grupo, max_textos=30):
    palabras = []
    for t in textos_grupo[:max_textos]:
        for w in string_norm_label(str(t)).split():
            if len(w) > 3:
                palabras.append(w)
    kw_str = " ".join(w for w, _ in Counter(palabras).most_common(12))
    return f"{subtema}. {subtema}. {kw_str}"[:500]


def _validar_estructura_tema(tema: str) -> bool:
    if not tema or len(tema.split()) < 2:
        return False
    if len(tema.split()) > 4:
        return False
    if re.match(r'^[0-9]', tema):
        return False
    num_palabras = re.compile(
        r'^(uno|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez|'
        r'once|doce|veinte|cien|varios|cada)', re.IGNORECASE
    )
    if num_palabras.match(tema):
        return False
    if _PATRON_TITULAR.match(tema):
        return False
    if _PATRON_ESTADO.search(tema):
        return False
    return True


def _tema_es_igual_a_subtema(tema: str, subtemas_grupo: list) -> bool:
    if not tema or not subtemas_grupo:
        return False
    tn = string_norm_label(tema)
    for sub in subtemas_grupo:
        sn = string_norm_label(sub)
        if not tn or not sn:
            continue
        if SequenceMatcher(None, tn, sn).ratio() >= 0.80:
            return True
        if tn in sn or sn in tn:
            return True
    return False


def _generar_nombre_tema_llm(subtemas_grupo, textos_muestra, titulos_muestra):
    subs_list = "\n".join(f"  · {s}" for s in subtemas_grupo[:8])
    palabras = []
    for t in titulos_muestra[:15]:
        for w in string_norm_label(str(t)).split():
            if len(w) > 3:
                palabras.append(w)
    kw = ", ".join(w for w, _ in Counter(palabras).most_common(6))
    tit_muestra = "\n".join(f"  · {t[:100]}" for t in list(dict.fromkeys(titulos_muestra))[:5])

    prompt = (
        "Eres editor jefe de un periódico. "
        "Crea UNA sección editorial (2-4 palabras) que agrupe estos subtemas.\n\n"
        "SUBTEMAS:\n" + subs_list +
        "\n\nTÍTULOS DE REFERENCIA:\n" + tit_muestra +
        f"\n\nKEYWORDS: {kw}\n\n"
        "REGLAS ESTRICTAS:\n"
        "  1. Piensa en secciones de periódico: 'Política', 'Economía', "
        "'Tecnología', 'Seguridad', 'Justicia', 'Medio Ambiente'.\n"
        "  2. Más GENERAL y ABSTRACTO que los subtemas — nunca repitas "
        "un subtema ni copies fragmentos de titular.\n"
        "  3. NUNCA incluyas números, cantidades ni nombres propios.\n"
        "  4. 2-4 palabras. Sustantivo + adjetivo o sustantivo solo.\n"
        "  5. Tildes y ñ correctas.\n\n"
        "CORRECTO: 'Política', 'Gestión legislativa', 'Justicia penal', "
        "'Regulación financiera'\n"
        "INCORRECTO: 'Cinco congresistas con líos', 'Congresistas electos', "
        "'Investigación disciplinaria congreso', 'Nuevo acuerdo'\n\n"
        'JSON: {"tema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=40,
            temperature=0.05,
            response_format={"type": "json_object"}
        )
        u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
        if u:
            st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
            st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
        raw = json.loads(resp.choices[0].message.content).get("tema", "").strip().replace('"', '').replace('.', '')
        nombre = limpiar_tema(raw)
        if not _validar_estructura_tema(nombre):
            return None
        return nombre
    except:
        return None


def _regenerar_tema_diferente(subtemas_grupo, titulos_muestra, intento=0):
    subs_list = ", ".join(subtemas_grupo[:8])
    prompt = (
        f"Subtemas: {subs_list}\n\n"
        "Genera UNA categoría GENERAL (2-3 palabras), diferente a los subtemas. "
        "Piensa en sección de periódico (Economía, Política, Tecnología, "
        "Infraestructura, Cultura, Deportes…). "
        "Tildes y ñ correctas, terminar en sustantivo/adjetivo.\n"
        'JSON: {"tema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=50,
            temperature=0.2 + intento * 0.1,
            response_format={"type": "json_object"}
        )
        u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
        if u:
            st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
            st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
        return limpiar_tema(json.loads(resp.choices[0].message.content).get("tema", "").strip().replace('"', '').replace('.', ''))
    except:
        return None


def consolidar_temas(subtemas, textos, pbar):
    n = len(textos)
    u = _umbrales_adaptativos(n)

    pbar.progress(0.05, "Preparando temas...")
    df = pd.DataFrame({'subtema': subtemas, 'texto': textos})
    us = list(df['subtema'].unique())
    if len(us) <= 1:
        pbar.progress(1.0, "Un tema")
        return [capitalizar_etiqueta(s) for s in subtemas]

    # Con corpus muy pequeño: si hay tantos subtemas únicos como noticias,
    # NO agrupar — cada noticia ya tiene su propio subtema correcto.
    if n <= 5 and len(us) == n:
        pbar.progress(1.0, "Corpus pequeño: temas = subtemas")
        st.info(f"Temas: **{n}** (corpus pequeño — cada noticia tiene tema propio)")
        return [capitalizar_etiqueta(s) for s in subtemas]

    pbar.progress(0.10, "Representaciones...")
    textos_por_subtema = defaultdict(list)
    for i, sub in enumerate(subtemas):
        textos_por_subtema[sub].append(textos[i])
    repr_enriquecidas = [_construir_representacion_grupo(sub, textos_por_subtema[sub]) for sub in us]
    pbar.progress(0.20, "Embeddings contenido...")
    emb_repr = get_embeddings_batch(repr_enriquecidas)
    emb_labels = get_embeddings_batch(us)
    ae = get_embeddings_batch(textos)
    centroids_contenido = {}
    for sub in us:
        idxs = df.index[df['subtema'] == sub].tolist()[:50]
        vecs = [ae[i] for i in idxs if ae[i] is not None]
        if vecs:
            centroids_contenido[sub] = np.mean(vecs, axis=0)
    pbar.progress(0.35, "Similitudes...")
    vs = [s for s in us if s in centroids_contenido]
    if len(vs) < 2:
        pbar.progress(1.0, "Sin agrupación")
        return [capitalizar_etiqueta(s) for s in subtemas]
    idx_map = {s: i for i, s in enumerate(us)}
    M_content = np.array([centroids_contenido[s] for s in vs])
    sim_content = cosine_similarity(M_content)
    has_repr = all(emb_repr[idx_map[s]] is not None for s in vs)
    has_label = all(emb_labels[idx_map[s]] is not None for s in vs)
    if has_repr and has_label:
        sim_combined = (
            0.50 * sim_content
            + 0.35 * cosine_similarity(np.array([emb_repr[idx_map[s]] for s in vs]))
            + 0.15 * cosine_similarity(np.array([emb_labels[idx_map[s]] for s in vs]))
        )
    elif has_repr:
        sim_combined = (
            0.60 * sim_content
            + 0.40 * cosine_similarity(np.array([emb_repr[idx_map[s]] for s in vs]))
        )
    else:
        sim_combined = sim_content

    pbar.progress(0.45, "Clustering temas...")
    dist_matrix = np.clip(1 - sim_combined, 0, 2)
    np.fill_diagonal(dist_matrix, 0)

    umbral_tema = u['tema']
    num_temas_max = u['num_temas_max']

    # Con pocos subtemas usamos enlace completo (más conservador)
    linkage_temas = 'complete' if len(vs) <= 6 else 'average'
    cl = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral_tema,
        metric='precomputed', linkage=linkage_temas
    ).fit(dist_matrix)
    if len(set(cl.labels_)) > num_temas_max:
        cl = AgglomerativeClustering(
            n_clusters=num_temas_max, metric='precomputed', linkage=linkage_temas
        ).fit(dist_matrix)

    clusters = defaultdict(list)
    for i, lbl in enumerate(cl.labels_):
        clusters[lbl].append(vs[i])
    uc = [s for s in us if s not in vs]
    mt = {}
    tc = len(clusters)
    pbar.progress(0.50, f"Nombres {tc} temas...")
    for k, (cid, subtemas_cluster) in enumerate(clusters.items()):
        pbar.progress(0.50 + 0.35 * (k / max(tc, 1)), f"Tema {k + 1}/{tc}...")
        titulos_cluster = []
        textos_cluster = []
        for sub in subtemas_cluster:
            for idx in df.index[df['subtema'] == sub].tolist()[:10]:
                txt = str(textos[idx])
                partes = txt.split('. ')
                if partes:
                    titulos_cluster.append(partes[0][:120])
                textos_cluster.append(txt[:200])
        if len(subtemas_cluster) == 1:
            sub_unico = subtemas_cluster[0]
            nombre = _generar_nombre_tema_llm(subtemas_cluster, textos_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                p = sub_unico.split()
                nombre = _recortar_frase_completa(" ".join(p), max_palabras=3) if len(p) > 3 else sub_unico
                if _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                    nombre = sub_unico
        else:
            nombre = _generar_nombre_tema_llm(subtemas_cluster, textos_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster, intento=1)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                all_words = []
                for sub in subtemas_cluster:
                    for w in string_norm_label(sub).split():
                        if len(w) > 3:
                            all_words.append(w)
                nombre = (
                    capitalizar_etiqueta(" ".join(w for w, _ in Counter(all_words).most_common(2)))
                    if all_words else subtemas_cluster[0]
                )
        if not _frase_esta_completa(nombre):
            nombre = _recortar_frase_completa(nombre, max_palabras=4)
            if not _frase_esta_completa(nombre):
                freq = Counter(subtemas)
                nombre = _recortar_frase_completa(
                    max(subtemas_cluster, key=lambda s: freq.get(s, 0)), max_palabras=4
                )
        nombre = capitalizar_etiqueta(nombre)
        for sub in subtemas_cluster:
            mt[sub] = nombre
    for sub in uc:
        mt[sub] = capitalizar_etiqueta(sub)

    pbar.progress(0.87, "Validando pertenencia mínima a temas...")
    min_tema = u['min_pertenencia_tema']
    tf_inicial = [mt.get(sub, sub) for sub in subtemas]
    tema_agrupacion: Dict[str, list] = defaultdict(list)
    for i, tema in enumerate(tf_inicial):
        if ae[i] is not None:
            tema_agrupacion[tema].append(ae[i])
    tema_centroids: Dict[str, np.ndarray] = {
        t: np.mean(vecs, axis=0) for t, vecs in tema_agrupacion.items() if vecs
    }
    tf_validado: List[str] = []
    n_forzadas = 0
    for i, (sub, tema_asignado) in enumerate(zip(subtemas, tf_inicial)):
        emb = ae[i]
        if emb is not None and tema_asignado in tema_centroids:
            sim = cosine_similarity(
                np.array(emb).reshape(1, -1),
                tema_centroids[tema_asignado].reshape(1, -1)
            )[0][0]
            if sim < min_tema:
                tf_validado.append(capitalizar_etiqueta(_recortar_frase_completa(sub, max_palabras=4)))
                n_forzadas += 1
                continue
        tf_validado.append(capitalizar_etiqueta(tema_asignado))
    if n_forzadas:
        st.caption(f"ℹ️ {n_forzadas} noticias con baja pertenencia al tema agrupado → tema propio asignado.")

    pbar.progress(0.88, "Dedup temas...")
    tf_validado = dedup_labels(tf_validado, u['dedup_label'])

    pbar.progress(0.90, "Fusionando temas solapados...")
    mapa_fusion_temas = _fusionar_temas_contenidos(tf_validado)
    if mapa_fusion_temas:
        tf_validado = [mapa_fusion_temas.get(t, t) for t in tf_validado]
        n_fusionados = len(mapa_fusion_temas)
        st.caption(f"ℹ️ {n_fusionados} tema(s) fusionado(s) por contención o solapamiento semántico.")

    pbar.progress(0.92, "Validando tema ≠ subtema...")
    tf_validado = _post_validar_tema_vs_subtema(tf_validado, subtemas)
    pbar.progress(0.95, "Completitud...")
    tf_validado = [
        capitalizar_etiqueta(_recortar_frase_completa(t) if not _frase_esta_completa(t) else t)
        for t in tf_validado
    ]
    tf_validado = _unificar_tema_por_subtema(tf_validado, subtemas)
    st.info(f"Temas: **{len(set(tf_validado))}** (de {len(set(subtemas))} subtemas) · Máx: {num_temas_max}")
    pbar.progress(1.0, "Temas listos")
    return tf_validado


def _fusionar_temas_contenidos(temas: List[str]) -> Dict[str, str]:
    unique = list(dict.fromkeys(temas))
    if len(unique) < 2:
        return {}

    normed = {t: string_norm_label(t) for t in unique}
    mapa: Dict[str, str] = {}

    for i, ta in enumerate(unique):
        for tb in unique[i + 1:]:
            na, nb = normed[ta], normed[tb]
            if not na or not nb:
                continue
            if (f" {na} " in f" {nb} ") or nb == na or nb.startswith(na + " ") or nb.endswith(" " + na):
                canon = tb if len(tb) >= len(ta) else ta
                reemplazar = ta if canon == tb else tb
                mapa[reemplazar] = canon
            elif (f" {nb} " in f" {na} ") or na.startswith(nb + " ") or na.endswith(" " + nb):
                canon = ta if len(ta) >= len(tb) else tb
                reemplazar = tb if canon == ta else ta
                mapa[reemplazar] = canon

    umbral_relajado = 0.70
    candidatos = [(t, normed[t]) for t in unique if len(t.split()) <= 3 and t not in mapa]
    if len(candidatos) >= 2:
        textos_c = [t for t, _ in candidatos]
        embs = get_embeddings_batch(textos_c)
        validos = [(textos_c[i], embs[i]) for i in range(len(textos_c)) if embs[i] is not None]
        if len(validos) >= 2:
            etqs, vecs = zip(*validos)
            sim = cosine_similarity(np.array(vecs))
            for i in range(len(etqs)):
                for j in range(i + 1, len(etqs)):
                    if sim[i][j] >= umbral_relajado:
                        ta, tb = etqs[i], etqs[j]
                        if ta in mapa or tb in mapa:
                            continue
                        words_a = set(normed[ta].split())
                        words_b = set(normed[tb].split())
                        if words_a & words_b:
                            freq = Counter(temas)
                            canon = ta if freq.get(ta, 0) >= freq.get(tb, 0) else tb
                            reemplazar = tb if canon == ta else ta
                            mapa[reemplazar] = canon

    return mapa


def _post_validar_tema_vs_subtema(temas, subtemas):
    tema_a_subtemas = defaultdict(set)
    for t, s in zip(temas, subtemas):
        tema_a_subtemas[t].add(s)
    reemplazos = {}
    for tema, subs in tema_a_subtemas.items():
        if len(subs) == 1:
            sub_unico = list(subs)[0]
            tn = string_norm_label(tema)
            sn = string_norm_label(sub_unico)
            if tn and sn and SequenceMatcher(None, tn, sn).ratio() >= 0.80:
                nuevo = _regenerar_tema_diferente([sub_unico], [])
                if nuevo and not _tema_es_igual_a_subtema(nuevo, [sub_unico]) and _frase_esta_completa(nuevo):
                    reemplazos[tema] = capitalizar_etiqueta(nuevo)
    return [reemplazos.get(t, t) for t in temas] if reemplazos else temas


def _unificar_tema_por_subtema(temas, subtemas):
    sub_to_temas = defaultdict(list)
    for t, s in zip(temas, subtemas):
        sub_to_temas[s].append(t)
    sub_to_best = {}
    for sub, tema_list in sub_to_temas.items():
        sub_to_best[sub] = Counter(tema_list).most_common(1)[0][0]
    return [sub_to_best[s] for s in subtemas]


def analizar_temas_con_pkl(textos, pkl_file):
    try:
        pipeline = joblib.load(pkl_file)
        return [capitalizar_etiqueta(str(p)) for p in pipeline.predict(textos)]
    except Exception as e:
        st.error(f"Error pkl: {e}")
        return None


# ======================================
# Duplicados y Excel
# ======================================

def _normalizar_url(url: str) -> str:
    if not url:
        return ""
    url = url.strip().lower()
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    url = url.rstrip('/')
    return url


def _extraer_url_streaming(row, km):
    ls_key = km.get("link_streaming")
    if not ls_key:
        return None
    val = row.get(ls_key)
    if isinstance(val, dict):
        url = val.get("url")
        return url.strip() if url and url.strip() else None
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None


def detectar_duplicados_avanzado(rows, km):
    processed = deepcopy(rows)
    seen_url, seen_bcast = {}, {}
    seen_streaming: Dict[tuple, int] = {}
    tb = defaultdict(list)

    for i, row in enumerate(processed):
        if row.get("is_duplicate"):
            continue

        tipo    = normalizar_tipo_medio(str(row.get(km.get("tipodemedio", ""))))
        mencion = norm_key(row.get(km.get("menciones", "")))
        medio   = norm_key(row.get(km.get("medio", "")))

        streaming_url_raw = _extraer_url_streaming(row, km)
        if streaming_url_raw and mencion:
            streaming_url_norm = _normalizar_url(streaming_url_raw)
            if streaming_url_norm:
                sk = (streaming_url_norm, mencion)
                if sk in seen_streaming:
                    row["is_duplicate"] = True
                    row["idduplicada"] = processed[seen_streaming[sk]].get(
                        km.get("idnoticia", ""), ""
                    )
                    continue
                seen_streaming[sk] = i

        if tipo == "Internet":
            li  = row.get(km.get("link_nota", {})) or {}
            url = li.get("url") if isinstance(li, dict) else None
            if url and mencion:
                url_norm = _normalizar_url(url)
                k = (url_norm, mencion)
                if k in seen_url:
                    row["is_duplicate"] = True
                    row["idduplicada"] = processed[seen_url[k]].get(km.get("idnoticia", ""), "")
                    continue
                seen_url[k] = i
            if medio and mencion:
                tb[(medio, mencion)].append(i)

        elif tipo in ("Radio", "Televisión"):
            hora = str(row.get(km.get("hora", ""), "")).strip()
            if mencion and medio and hora:
                k = (mencion, medio, hora)
                if k in seen_bcast:
                    row["is_duplicate"] = True
                    row["idduplicada"] = processed[seen_bcast[k]].get(km.get("idnoticia", ""), "")
                else:
                    seen_bcast[k] = i

    for idxs in tb.values():
        if len(idxs) < 2:
            continue
        for i in range(len(idxs)):
            for j in range(i + 1, len(idxs)):
                a, b = idxs[i], idxs[j]
                if processed[a].get("is_duplicate") or processed[b].get("is_duplicate"):
                    continue
                ta  = normalize_title_for_comparison(processed[a].get(km.get("titulo", "")))
                tb_ = normalize_title_for_comparison(processed[b].get(km.get("titulo", "")))
                if ta and tb_ and SequenceMatcher(None, ta, tb_).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    if len(ta) < len(tb_):
                        processed[a]["is_duplicate"] = True
                        processed[a]["idduplicada"]  = processed[b].get(km.get("idnoticia", ""), "")
                    else:
                        processed[b]["is_duplicate"] = True
                        processed[b]["idduplicada"]  = processed[a].get(km.get("idnoticia", ""), "")

    return processed


def run_dossier_logic(sheet, xlsx_bytes=None, cliente="", voceros="", enable_scraping=False):
    """Paso 1: Limpieza y duplicados. Si enable_scraping=True, scrapea GlobalNews."""
    headers = [c.value for c in sheet[1] if c.value]
    nk = [norm_key(h) for h in headers]
    km = {n: n for n in nk}
    km.update({
        "titulo": norm_key("Titulo"),
        "resumen": norm_key("Resumen - Aclaracion"),
        "menciones": norm_key("Menciones - Empresa"),
        "medio": norm_key("Medio"),
        "tonoiai": norm_key("Tono IA"),
        "tema": norm_key("Tema"),
        "subtema": norm_key("Subtema"),
        "idnoticia": norm_key("ID Noticia"),
        "idduplicada": norm_key("ID duplicada"),
        "tipodemedio": norm_key("Tipo de Medio"),
        "hora": norm_key("Hora"),
        "fecha": norm_key("Fecha"),
        "link_nota": norm_key("Link Nota"),
        "link_streaming": norm_key("Link (Streaming - Imagen)"),
        "region": norm_key("Region")
    })
    rows, split_rows = [], []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue
        rows.append({nk[i]: c for i, c in enumerate(row) if i < len(nk)})
    for rc in rows:
        base = {
            k: (extract_link(v) if k in (km["link_nota"], km["link_streaming"]) else v.value)
            for k, v in rc.items()
        }
        if km.get("tipodemedio") in base:
            base[km["tipodemedio"]] = normalizar_tipo_medio(base.get(km["tipodemedio"]))
        ml = [m.strip() for m in str(base.get(km["menciones"], "")).split(";") if m.strip()]
        for m in ml or [None]:
            nr = deepcopy(base)
            if m:
                nr[km["menciones"]] = m
            split_rows.append(nr)
    for idx, row in enumerate(split_rows):
        row.update({"original_index": idx, "is_duplicate": False})
    processed = detectar_duplicados_avanzado(split_rows, km)
    for row in processed:
        if row["is_duplicate"]:
            row.update({km["tonoiai"]: "Duplicada", km["tema"]: "-", km["subtema"]: "-"})
    
    # ── Scraping opcional ──
    scraped_count = 0
    if enable_scraping and xlsx_bytes and cliente:
        scrape_cache = _load_cache(_SCRAPE_CACHE_PATH)
        resumenes_cache = _load_cache(_LLM_RESUMEN_CACHE_PATH)
        urls_map = extract_urls_from_xlsx(xlsx_bytes)
        
        # Map row numbers to URLs for column W
        url_rows = []
        ref_to_xlsx_row = {}
        ws_xls = sheet
        for ref, url in urls_map.items():
            cm = re.match(r'([A-Z]+)(\d+)', ref)
            if cm and cm.group(1) == "W":
                ref_to_xlsx_row[int(cm.group(2))] = url
        
        # Build scrape list: match processed rows to URLs
        for row in processed:
            if row.get("is_duplicate"): continue
            orig = row.get("original_index")
            xlsx_row = orig + 2  # 1-indexed + header
            if xlsx_row in ref_to_xlsx_row:
                nid = row.get(km.get("idnoticia", ""), "")
                url = ref_to_xlsx_row[xlsx_row]
                url_rows.append((xlsx_row, url, str(nid)))
        
        if url_rows:
            with st.status("Scraping de noticias...", expanded=True) as ss:
                pb = st.progress(0)
                scraped = scrape_all_news(url_rows, scrape_cache, pb, ss)
                n_ok = sum(1 for v in scraped.values() if v)
                ss.update(label="Scraping: {} noticias scrapeedas".format(n_ok), state="complete")
            
            # Generate client-focused summaries for scraped news
            for row in processed:
                if row.get("is_duplicate"): continue
                orig = row.get("original_index")
                xlsx_row = orig + 2
                texto = scraped.get(xlsx_row)
                if texto:
                    rk = km.get("resumen", "resumen")
                    ck = "rc_" + str(row.get(km.get("idnoticia", ""), ""))
                    if ck in resumenes_cache:
                        summary = resumenes_cache[ck]
                    else:
                        titulo = str(row.get(km.get("titulo", ""), ""))
                        medio = str(row.get(km.get("medio", ""), ""))
                        fecha = str(row.get(km.get("fecha", ""), ""))
                        summary = generar_resumen_cliente(texto, titulo, medio, fecha, cliente, voceros)
                        if summary: resumenes_cache[ck] = summary
                    if summary:
                        row[rk] = summary
                    scraped_count += 1
            
            if resumenes_cache:
                _save_cache(_LLM_RESUMEN_CACHE_PATH, resumenes_cache)
    
    return processed, km


def fix_links_by_media_type(row, km):
    tkey = km.get("tipodemedio")
    ln   = km.get("link_nota")
    ls   = km.get("link_streaming")
    if not (tkey and ln and ls):
        return
    tipo = row.get(tkey, "")
    rl   = row.get(ln) or {"value": "", "url": None}
    rs   = row.get(ls) or {"value": "", "url": None}
    hurl = lambda x: isinstance(x, dict) and bool(x.get("url"))
    if tipo in ("Radio", "Televisión"):
        row[ls] = {"value": "", "url": None}
    elif tipo == "Internet":
        row[ln], row[ls] = rs, rl
    elif tipo in ("Prensa", "Revistas"):
        if not hurl(rl) and hurl(rs):
            row[ln] = rs
        row[ls] = {"value": "", "url": None}


def generate_output_excel(rows, km):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ORDER = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio", "Region",
        "Seccion - Programa", "Titulo", "Autor - Conductor", "Nro. Pagina",
        "Dimension", "Duracion - Nro. Caracteres", "CPE", "Audiencia", "Tier",
        "Tono", "Tono IA", "Tema", "Subtema", "Link Nota",
        "Resumen - Aclaracion", "Link (Streaming - Imagen)", "Menciones - Empresa",
        "ID duplicada"
    ]
    NUM = {"ID Noticia", "Nro. Pagina", "Dimension", "Duracion - Nro. Caracteres", "CPE", "Tier", "Audiencia"}
    ws.append(ORDER)
    ls = NamedStyle(name="HL", font=Font(color="0000FF", underline="single"))
    if "HL" not in wb.style_names:
        wb.add_named_style(ls)
    for row in rows:
        tk = km.get("titulo")
        if tk and tk in row:
            row[tk] = clean_title_for_output(row.get(tk))
        rk = km.get("resumen")
        if rk and rk in row:
            row[rk] = corregir_texto(row.get(rk))
        out, links = [], {}
        for ci, h in enumerate(ORDER, 1):
            dk  = km.get(norm_key(h), norm_key(h))
            val = row.get(dk)
            cv  = None
            if h in NUM:
                try:
                    cv = float(val) if val is not None and str(val).strip() != "" else None
                except:
                    cv = str(val) if val is not None else None
            elif isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"):
                    links[ci] = val["url"]
            elif val is not None:
                cv = str(val)
            out.append(cv)
        ws.append(out)
        for ci, url in links.items():
            cell       = ws.cell(row=ws.max_row, column=ci)
            cell.hyperlink = url
            cell.style = "HL"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ======================================
# Proceso principal
# ======================================
async def run_full_process_async(df_file, bn, ba, tpkl, epkl, mode, xlsx_bytes=None, cliente="", voceros="", enable_scraping=False):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().clear()
    t0 = time.time()
    if "API" in mode:
        try:
            openai.api_key=st.secrets["OPENAI_API_KEY"]
            openai.aiosession.set(None)
        except:
            st.error("OPENAI_API_KEY no encontrado.")
            st.stop()
    with st.status("Paso 1 · Limpieza y duplicados", expanded=True) as s:
        wb_in = load_workbook(df_file, data_only=True)
        rows, km = run_dossier_logic(wb_in.active, xlsx_bytes=xlsx_bytes,
                                     cliente=cliente, voceros=voceros,
                                     enable_scraping=enable_scraping)
        s.update(label="✓ Paso 1", state="complete")
    with st.status("Paso 2 · Mapeos", expanded=True) as s:
        dfr  = _cargar_mapa_excel(st.secrets["REGION_MAP_URL"])
        rmap = {str(k).lower().strip(): v for k, v in pd.Series(dfr.iloc[:, 1].values, index=dfr.iloc[:, 0]).to_dict().items()}
        dfi  = _cargar_mapa_excel(st.secrets["INTERNET_MAP_URL"])
        imap = {str(k).lower().strip(): v for k, v in pd.Series(dfi.iloc[:, 1].values, index=dfi.iloc[:, 0]).to_dict().items()}
        for row in rows:
            mk = str(row.get(km.get("medio", ""), "")).lower().strip()
            row[km.get("region")] = rmap.get(mk, "N/A")
            if mk in imap:
                row[km.get("medio")]       = imap[mk]
                row[km.get("tipodemedio")] = "Internet"
            fix_links_by_media_type(row, km)
        s.update(label="✓ Paso 2", state="complete")
    gc.collect()
    ta = [r for r in rows if not r.get("is_duplicate")]
    if ta:
        df = pd.DataFrame(ta)
        df["_txt"] = df.apply(
            lambda r: texto_para_embedding(str(r.get(km["titulo"], "")), str(r.get(km["resumen"], ""))),
            axis=1
        )
        with st.status("Embeddings...", expanded=True) as s:
            _ = get_embeddings_batch(df["_txt"].tolist())
            s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")
        with st.status("Paso 3 · Tono", expanded=True) as s:
            pb = st.progress(0)
            if "PKL" in mode and tpkl:
                res = analizar_tono_con_pkl(df["_txt"].tolist(), tpkl)
                if res is None:
                    st.stop()
            elif "API" in mode:
                res = await ClasificadorTono(bn, ba).procesar_lote_async(
                    df["_txt"], pb, df[km["resumen"]], df[km["titulo"]]
                )
            else:
                res = [{"tono": "N/A"}] * len(ta)
            df[km["tonoiai"]] = [r["tono"] for r in res]
            s.update(label="✓ Paso 3 · Tono", state="complete")
        with st.status("Paso 4 · Clasificación", expanded=True) as s:
            pb = st.progress(0)
            if "Solo Modelos PKL" in mode:
                subtemas = ["N/A"] * len(ta)
                temas    = ["N/A"] * len(ta)
            else:
                subtemas = ClasificadorSubtema(bn, ba).procesar_lote(
                    df["_txt"], pb, df[km["resumen"]], df[km["titulo"]]
                )
                temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb)
            df[km["subtema"]] = subtemas
            if "PKL" in mode and epkl:
                tp = analizar_temas_con_pkl(df["_txt"].tolist(), epkl)
                if tp:
                    df[km["tema"]] = _unificar_tema_por_subtema(tp, subtemas)
            else:
                df[km["tema"]] = temas
            s.update(label="✓ Paso 4 · Clasificación", state="complete")
        rm2 = df.set_index("original_index").to_dict("index")
        for row in rows:
            if not row.get("is_duplicate"):
                row.update(rm2.get(row["original_index"], {}))
    gc.collect()
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    with st.status("Paso 5 · Informe", expanded=True) as s:
        st.session_state["output_data"]     = generate_output_excel(rows, km)
        st.session_state["output_filename"] = f"Informe_IA_{bn.replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.session_state["processing_complete"] = True
        st.session_state.update({
            "brand_name": bn, "brand_aliases": ba,
            "total_rows": len(rows), "unique_rows": len(ta), "duplicates": len(rows) - len(ta),
            "process_duration": f"{time.time() - t0:.0f}s",
            "process_cost": f"${ci + co + ce:.4f} USD",
            "cache_stats": get_embedding_cache().stats()
        })
        s.update(label=f"✓ Completado · {get_embedding_cache().stats()}", state="complete")


async def run_quick_async(df, tc, sc, bn, al):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().clear()
    df['_txt'] = df.apply(
        lambda r: texto_para_embedding(str(r.get(tc, "")), str(r.get(sc, ""))), axis=1
    )
    with st.status("Embeddings...", expanded=True) as s:
        _ = get_embeddings_batch(df['_txt'].tolist())
        s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")
    with st.status("Tono", expanded=True) as s:
        pb = st.progress(0)
        res = await ClasificadorTono(bn, al).procesar_lote_async(
            df["_txt"], pb, df[sc].fillna(''), df[tc].fillna('')
        )
        df['Tono IA'] = [r["tono"] for r in res]
        s.update(label="✓ Tono", state="complete")
    with st.status("Clasificación", expanded=True) as s:
        pb = st.progress(0)
        subtemas = ClasificadorSubtema(bn, al).procesar_lote(
            df["_txt"], pb, df[sc].fillna(''), df[tc].fillna('')
        )
        df['Subtema'] = subtemas
        temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb)
        df['Tema'] = temas
        s.update(label="✓ Clasificación", state="complete")
    df.drop(columns=['_txt'], inplace=True)
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    st.session_state['quick_cost'] = f"${ci + co + ce:.4f} USD"
    return df


def gen_quick_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df.to_excel(w, index=False, sheet_name='Analisis')
    return buf.getvalue()


def render_quick_tab():
    st.markdown('<div class="sec-label">Análisis rápido</div>', unsafe_allow_html=True)
    if 'quick_result' in st.session_state:
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Completado</div>'
            '<div class="success-sub">Listo para descargar</div></div></div>',
            unsafe_allow_html=True
        )
        st.metric("Costo", st.session_state.get('quick_cost', "$0.00"))
        st.dataframe(st.session_state.quick_result.head(10), use_container_width=True)
        st.download_button(
            "Descargar",
            data=gen_quick_excel(st.session_state.quick_result),
            file_name="Analisis_Rapido_IA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        if st.button("Nuevo análisis"):
            for k in ('quick_result', 'quick_df', 'quick_name', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()
        return
    if 'quick_df' not in st.session_state:
        st.markdown("Sube un Excel con columnas de título y resumen.")
        f = st.file_uploader("Excel", type=["xlsx"], label_visibility="collapsed", key="qu")
        if f:
            try:
                st.session_state.quick_df   = pd.read_excel(f)
                st.session_state.quick_name = f.name
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")
    else:
        st.success(f"**{st.session_state.quick_name}** cargado")
        with st.form("qf"):
            cols = st.session_state.quick_df.columns.tolist()
            c1, c2 = st.columns(2)
            tc = c1.selectbox("Col. título",  cols, 0)
            sc = c2.selectbox("Col. resumen", cols, 1 if len(cols) > 1 else 0)
            bn  = st.text_input("Marca",       placeholder="Ej: Bancolombia")
            bat = st.text_input("Alias (;)",   placeholder="Ej: Grupo Bancolombia;Ban")
            if st.form_submit_button("Analizar", use_container_width=True, type="primary"):
                if not bn:
                    st.error("Indica la marca.")
                else:
                    try:
                        openai.api_key = st.secrets["OPENAI_API_KEY"]
                        openai.aiosession.set(None)
                    except:
                        st.error("OPENAI_API_KEY no encontrada.")
                        st.stop()
                    al = [a.strip() for a in bat.split(";") if a.strip()]
                    with st.spinner("Procesando..."):
                        st.session_state.quick_result = asyncio.run(
                            run_quick_async(st.session_state.quick_df.copy(), tc, sc, bn, al)
                        )
                    st.rerun()
        if st.button("Otro archivo"):
            for k in ('quick_df', 'quick_name', 'quick_result', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()


# ======================================
# Main
# ======================================
def main():
    load_custom_css()
    if not check_password():
        return

    st.markdown("""
    <div class="app-header">
        <div class="app-header-icon">◈</div>
        <div class="app-header-text">
            <div class="app-header-title">Análisis de Noticias</div>
            <div class="app-header-version">v17.12 · clustering adaptativo por tamaño de corpus</div>
        </div>
        <div class="app-header-badge">IA</div>
    </div>""", unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["Análisis Completo", "Análisis Rápido"])

    with tab1:
        if not st.session_state.get("processing_complete", False):

            st.markdown('<div class="sec-label">Configuración</div>', unsafe_allow_html=True)
            cl, cr = st.columns([3, 2])
            with cl:
                bn  = st.text_input("Marca principal", placeholder="Ej: Bancolombia", key="bn")
                bat = st.text_input("Alias (separados por ;)", placeholder="Ej: Grupo Bancolombia;Ban", key="ba")
            with cr:
                mode = st.radio(
                    "Modo de análisis",
                    ["API de OpenAI", "Híbrido (PKL + API)", "Solo Modelos PKL"],
                    index=0, key="mode"
                )
                enable_sc = st.checkbox("Scrapear texto de noticias (GlobalNews)", key="enable_scraping",
                                       help="Extrae el texto real de cada noticia para generar resumen enfocado")
                if enable_sc:
                    st.text_input(
                        "Voceros (sep. por ;) — para resaltar en resumen",
                        key="voceros_scrape", placeholder="Ej: Gonzalo Moreno;Juan Perez")

            tpkl, epkl = None, None
            if "PKL" in mode:
                st.markdown('<div class="sec-label">Modelos PKL</div>', unsafe_allow_html=True)
                p1, p2 = st.columns(2)
                tpkl = p1.file_uploader(
                    "Modelo de Sentimiento (.pkl)",
                    type=["pkl"], key="tpkl",
                    help="Pipeline sklearn para clasificar tono: -1/0/1 o Negativo/Neutro/Positivo"
                )
                epkl = p2.file_uploader(
                    "Modelo de Temas (.pkl)",
                    type=["pkl"], key="epkl",
                    help="Pipeline sklearn para clasificar temas"
                )

            with st.form("main_form"):
                st.markdown('<div class="sec-label">Archivo de entrada</div>', unsafe_allow_html=True)
                st.markdown("""
                <div class="upload-zone" style="grid-template-columns:1fr">
                    <div class="upload-zone-card">
                        <div class="upload-zone-icon uz-dossier">📋</div>
                        <div class="upload-zone-text">
                            <div class="upload-zone-title">Dossier</div>
                            <div class="upload-zone-desc">Noticias a analizar · Región e Internet se cargan desde GitHub</div>
                        </div>
                    </div>
                </div>""", unsafe_allow_html=True)
                f1 = st.file_uploader("Dossier", type=["xlsx"], label_visibility="collapsed", key="f1")

                st.markdown(
                    f'<div class="cluster-info">'
                    f'<b>Parámetros base</b> · Sub={UMBRAL_SUBTEMA} · Tema={UMBRAL_TEMA} · Máx={NUM_TEMAS_MAX} '
                    f'· FusInter={UMBRAL_FUSION_INTERGRUPO} · FusSem={UMBRAL_FUSION_SUBTEMAS} '
                    f'· Dedup={UMBRAL_DEDUP_LABEL} · MinSub={UMBRAL_MIN_PERTENENCIA_SUBTEMA} '
                    f'· MinTema={UMBRAL_MIN_PERTENENCIA_TEMA} · MaxGrupo={MAX_GRUPO_ETIQUETA} · '
                    f'<b>Coherencia={UMBRAL_COHERENCIA_ETIQUETA}</b> · '
                    f'<b>SimMin={SIM_MINIMA_AGRUPACION_SUBTEMA}</b> (adaptativos según n)'
                    f'</div>',
                    unsafe_allow_html=True
                )

                if st.form_submit_button("▶ Iniciar análisis", use_container_width=True, type="primary"):
                    if not all([f1, bn.strip()]):
                        st.error("Completa todos los campos.")
                    else:
                        al = [a.strip() for a in bat.split(";") if a.strip()]
                        cur_mode = st.session_state.get("mode", "API de OpenAI")
                        cur_tpkl = st.session_state.get("tpkl")
                        cur_epkl = st.session_state.get("epkl")
                        # Leer bytes para scraping
                        f1.seek(0); xlsx_bytes = f1.read(); f1.seek(0)
                        enable_scrape = st.session_state.get("enable_scraping", False)
                        voceros_scrape = st.session_state.get("voceros_scrape", "")
                        asyncio.run(run_full_process_async(f1, bn, al, cur_tpkl, cur_epkl, cur_mode,
                                                         xlsx_bytes=xlsx_bytes if enable_scrape else None,
                                                         cliente=bn, voceros=voceros_scrape,
                                                         enable_scraping=enable_scrape))
                        st.rerun()
        else:
            total = st.session_state.total_rows
            uniq  = st.session_state.unique_rows
            dups  = st.session_state.duplicates
            dur   = st.session_state.process_duration
            cost  = st.session_state.get("process_cost", "$0.00")
            st.markdown(
                '<div class="success-banner"><div class="success-icon">✓</div>'
                '<div><div class="success-title">Análisis completado</div>'
                '<div class="success-sub">Informe listo para descargar</div></div></div>',
                unsafe_allow_html=True
            )
            st.markdown(f"""
            <div class="metrics-grid">
              <div class="metric-card m-total"><div class="metric-val" style="color:var(--text)">{total}</div><div class="metric-lbl">Total</div></div>
              <div class="metric-card m-unique"><div class="metric-val" style="color:var(--green)">{uniq}</div><div class="metric-lbl">Únicas</div></div>
              <div class="metric-card m-dup"><div class="metric-val" style="color:var(--amber)">{dups}</div><div class="metric-lbl">Duplicados</div></div>
              <div class="metric-card m-time"><div class="metric-val" style="color:var(--blue)">{dur}</div><div class="metric-lbl">Tiempo</div></div>
              <div class="metric-card m-cost"><div class="metric-val" style="color:var(--accent)">{cost}</div><div class="metric-lbl">Costo</div></div>
            </div>""", unsafe_allow_html=True)
            if 'cache_stats' in st.session_state:
                st.caption(f"📊 {st.session_state['cache_stats']}")
            c1, c2 = st.columns(2)
            c1.download_button(
                "⬇ Descargar informe",
                data=st.session_state.output_data,
                file_name=st.session_state.output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            if c2.button("Nuevo análisis", use_container_width=True):
                pwd = st.session_state.get("password_correct")
                st.session_state.clear()
                st.session_state.password_correct = pwd
                st.rerun()

    with tab2:
        render_quick_tab()

    st.markdown(
        '<div class="footer">v17.12 · Análisis de Noticias con IA · Johnathan Cortés ©</div>',
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
